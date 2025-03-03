import openpyxl
import sys
from os.path import abspath, dirname
import os
from pprint import pprint
from waapi import WaapiClient
import shutil

import comlib.cloudfeishu_h as cloudfeishu_h
import comlib.oi_h as oi_h
import comlib.config as config
import comlib.excel_h as excel_h
import comlib.waapi_h as waapi_h

# 新特性
# State自动赋值

"""根目录获取"""


def get_py_path():
    py_path = ""
    if hasattr(sys, 'frozen'):
        py_path = dirname(sys.executable)
    elif __file__:
        py_path = dirname(abspath(__file__))
    return py_path


# 打包版
root_path = get_py_path()

"""在线表获取"""
cloudfeishu_h.thread_download_excel(config.mus_sheet_token_dict, root_path)
excel_path = os.path.join(root_path, 'Excel')
file_name_list = excel_h.excel_get_path_list(excel_path)
# pprint(file_name_list)

"""功能数据初始化"""
# switch创建列表
switch_create_list = []
# unit创建列表
unit_create_list = []
# playlist创建列表
playlist_create_list = []

"""音乐资源名称列表"""
mus_name_list = []
value_desc_list = []

"""Mus媒体资源名称及文件夹映射"""
media_info_dict, _ = oi_h.get_type_file_name_and_path('.wav', os.path.join(root_path, 'New_Media'))
# pprint(media_info_dict)

"""音乐媒体资源路径"""
wwise_mus_original_path = os.path.join(config.original_path, 'Mus')

"""*****************功能检测区******************"""
"""字符串长度检查"""


def check_by_str_length(str1, length, name):
    if len(str1) > length:
        if "_" in str1:
            oi_h.print_error(name + "：描述后缀名过长，限制字符数为" + str(length))
        else:
            oi_h.print_error(name + "：通过_切分的某个单词过长，每个单词长度不应超过10个字符，需要再拆分")
        return False
    return True


"""分隔符的大小写和长度检查返回bool"""


def check_by_length_and_word_bool(name, word_list_len):
    word_list = name.split("_")
    if len(word_list) <= word_list_len:
        for word in word_list:
            """判定_的每个都不能超过10个"""
            if not check_by_str_length(word, config.one_word_len, name):
                return False
            """判定_的每个开头必须大写"""
            if len(word) > 0:
                # 只需要检查第一个字符，因为istitle会导致检查字符串只能首字母大写
                if word[0].istitle() == False and word[0].isdigit() == False:
                    is_title = False
                    oi_h.print_error(name + "：通过”_“分隔的每个单词开头都需要大写")
                    return False
    else:
        oi_h.print_error(name + "：通过”_“分隔的单词个数不能超过" + str(config.word_list_len))
        return False
    return True


"""检测字符串是否含有中文"""


def check_is_chinese(string):
    for ch in string:
        if u'\u4e00' <= ch <= u'\u9fff':
            return True
    return False


with WaapiClient() as client:
    """*****************Wwise功能区******************"""
    """设置obj的notes"""


    def set_obj_notes(obj_id, notes_value):
        args = {
            'object': obj_id,
            'value': notes_value
        }
        client.call("ak.wwise.core.object.setNotes", args)
        # print_warning(obj_name + "描述更改为：" + notes_value)


    """设置对象属性"""


    def set_obj_property(obj_id, property, value):
        args = {
            "object": obj_id,
            "property": property,
            "value": value
        }
        client.call("ak.wwise.core.object.setProperty", args)
        # print(name_2 + ":" + str(trigger_rate))


    """obj创建"""


    def create_wwise_obj(type, parent, name, notes):
        unit_object = None
        if name != "Mus":
            unit_object = client.call("ak.wwise.core.object.create",
                                      waapi_h.args_object_create(parent,
                                                                 type, name, notes))
            oi_h.print_warning(name + "：" + type + "已创建")

        return unit_object


    """获取Wwise中的Unit列表"""


    def get_wwise_type_list(root_path, type):
        # 获取Wwise所有的Audio Unit
        wwise_bus_name_all_list = client.call("ak.wwise.core.object.get",
                                              waapi_h.waql_by_type(type, root_path),
                                              options=config.options)['return']
        wwise_bus_name_name_dict = {item['name']: item['id'] for item in wwise_bus_name_all_list}
        # pprint(wwise_bus_name_name_dict)
        # {'Amb': '{4BB49221-4DC1-4A05-A57C-3D42FA570675}',
        #  'Amb_A01': '{415AC456-D390-43D5-A09E-B5D30B253D76}',
        #  'CG': '{BBB332B4-C230-4D60-9ABA-91A36D968C11}'}
        return wwise_bus_name_name_dict


    """mus创建"""


    def create_wwise_mus(unit_list, playlist_list, switch_list):

        # unit创建
        for mus_name in unit_list:
            # 查找已有unit列表
            unit_have_dict = get_wwise_type_list(config.wwise_mus_global_path, "WorkUnit")
            playlist_have_dict = get_wwise_type_list(config.wwise_mus_global_path, "MusicPlaylistContainer")
            switch_have_dict = get_wwise_type_list(config.wwise_mus_global_path, "MusicSwitchContainer")
            # pprint(unit_have_dict)
            # Wwise里没有unit的的
            if mus_name not in unit_have_dict:
                # 查找该unit需要放的父级是谁
                unit_parent = oi_h.find_longest_prefix_key(mus_name, unit_have_dict)
                if not unit_parent:
                    # WorkUnit创建
                    unit_object = create_wwise_obj("WorkUnit",
                                                   config.wwise_mus_global_path, mus_name,
                                                   "")
                    if unit_object:
                        # Switch父级创建
                        switch_object = create_wwise_obj("MusicSwitchContainer",
                                                         unit_object['id'], mus_name,
                                                         "")
                        print("1:" + mus_name)

                else:
                    # t5apprint(playlist_list)
                    unit_have_dict = get_wwise_type_list(config.wwise_mus_global_path, "WorkUnit")
                    unit_parent = oi_h.find_longest_prefix_key(mus_name, unit_have_dict)
                    if mus_name in switch_list:
                        if mus_name not in switch_have_dict:
                            print("2:" + mus_name)
                            # WorkUnit创建
                            unit_object = create_wwise_obj("WorkUnit",
                                                           switch_have_dict[unit_parent], mus_name,
                                                           "")

                            # Switch父级创建
                            switch_object = create_wwise_obj("MusicSwitchContainer",
                                                             unit_object['id'], mus_name,
                                                             "")

                        else:
                            print("3:" + mus_name)
                            unit_object = create_wwise_obj("WorkUnit",
                                                           switch_have_dict[mus_name], mus_name,
                                                           "")

                    elif mus_name in playlist_list:
                        if mus_name not in playlist_have_dict:
                            print("4:" + mus_name)
                            # WorkUnit创建
                            unit_object = create_wwise_obj("WorkUnit",
                                                           switch_have_dict[unit_parent], mus_name,
                                                           "")

                            # Switch父级创建
                            playlist_object = create_wwise_obj("MusicPlaylistContainer",
                                                               unit_object['id'], mus_name,
                                                               "")

        # switch创建
        for mus_name in switch_list:
            unit_have_dict = get_wwise_type_list(config.wwise_mus_global_path, "WorkUnit")
            switch_have_dict = get_wwise_type_list(config.wwise_mus_global_path, "MusicSwitchContainer")

            if mus_name not in switch_have_dict:
                switch_parent = oi_h.find_longest_prefix_key(mus_name, switch_have_dict)
                parent_id = ""
                if switch_parent in switch_have_dict:
                    parent_id = switch_have_dict[switch_parent]

                if parent_id:
                    switch_object = create_wwise_obj("MusicSwitchContainer",
                                                     parent_id, mus_name,
                                                     "")
        # playlist创建
        for mus_name in playlist_list:
            unit_have_dict = get_wwise_type_list(config.wwise_mus_global_path, "WorkUnit")
            switch_have_dict = get_wwise_type_list(config.wwise_mus_global_path, "MusicSwitchContainer")
            playlist_have_dict = get_wwise_type_list(config.wwise_mus_global_path, "MusicPlaylistContainer")
            if mus_name not in playlist_have_dict:
                # print(mus_name)
                # 找到的父级为unit
                unit_parent_1 = oi_h.find_longest_prefix_key(mus_name, unit_have_dict)
                unit_parent_2 = oi_h.find_longest_prefix_key(mus_name, switch_have_dict)
                parent_id = None
                # 前缀找不到尝试找同名的
                if not unit_parent_1:
                    if mus_name in unit_have_dict:
                        parent_id = unit_have_dict[mus_name]
                if unit_parent_1:

                    if unit_parent_2:
                        if len(unit_parent_2) >= len(unit_parent_1):
                            parent_id = switch_have_dict[unit_parent_2]
                        else:
                            parent_id = unit_have_dict[unit_parent_1]
                    else:
                        parent_id = unit_have_dict[unit_parent_1]
                elif unit_parent_2:
                    parent_id = switch_have_dict[unit_parent_2]

                if parent_id:
                    playlist_object = create_wwise_obj("MusicPlaylistContainer",
                                                       parent_id, mus_name,
                                                       "")


    """获取需要创建的mus列表"""


    def get_create_mus_name_list():
        # 遍历每一行
        for row in mus_config_sheet.iter_rows():
            mus_name = None
            # 遍历每一列
            for cell in row:
                cell_value, is_merge = excel_h.check_is_mergecell(cell, mus_config_sheet)
                if cell_value:
                    is_unit = False
                    # print(cell_value)
                    if not mus_name:
                        mus_name = cell_value
                        is_unit = True
                    else:
                        mus_name += "_" + cell_value
                        fill = cell.fill
                        if fill.fill_type == 'solid':
                            if fill.start_color.index == "FFD9F3FD":
                                is_unit = True
                    is_switch = False

                    if is_unit:
                        if mus_name not in unit_create_list:
                            unit_create_list.append(mus_name)
                    # 是否为switch容器检查
                    if is_merge:
                        is_switch = True
                    else:
                        if mus_config_sheet.cell(row=cell.row,
                                                 column=cell.column + 1).value:
                            is_switch = True
                    if is_switch:
                        if mus_name not in switch_create_list:
                            switch_create_list.append(mus_name)
                    else:
                        if mus_name not in playlist_create_list:
                            playlist_create_list.append(mus_name)

            # print()


    """*****************主程序处理******************"""
    """音乐结构创建"""

    # 拼接xlsx的路径
    file_path_xlsx = os.path.join(excel_path, config.excel_mus_config_path)
    # 获取xlsx的workbook
    wb = openpyxl.load_workbook(file_path_xlsx)
    mus_config_sheet = wb['Mus']
    get_create_mus_name_list()
    switch_create_list.sort(key=len)
    playlist_create_list.sort(key=len)
    unit_create_list.sort(key=len)
    # pprint(switch_create_list)
    # pprint(playlist_create_list)
    # pprint(unit_create_list)
    create_wwise_mus(unit_create_list, playlist_create_list, switch_create_list)

    """******************对象删除清理********************"""
    """对象删除"""


    def delete_obj(obj_id, obj_name, obj_type, warn=True):
        args = {
            "object": "%s" % obj_id
        }
        client.call("ak.wwise.core.object.delete", args)
        if warn:
            oi_h.print_warning(obj_name + "(" + obj_type + ")" + ":已删除")


    """删除状态"""


    def delete_check(wwise_obj_dict, excel_list, obj_type):
        for wwise_obj in wwise_obj_dict:
            if wwise_obj not in excel_list:
                obj_list = client.call("ak.wwise.core.object.get",
                                       waapi_h.waql_from_id(wwise_obj_dict[wwise_obj]),
                                       options=config.options)['return']
                # pprint(obj_list)
                if obj_list:
                    # 还需要notes不相同
                    if obj_list[0]['notes'] not in value_desc_list:
                        # State/StateGroup删除
                        delete_obj(wwise_obj_dict[wwise_obj], wwise_obj, obj_type)
                        # StateGroupUnit删除


    unit_have_dict = get_wwise_type_list(config.wwise_mus_global_path, "WorkUnit")
    switch_have_dict = get_wwise_type_list(config.wwise_mus_global_path, "MusicSwitchContainer")
    playlist_have_dict = get_wwise_type_list(config.wwise_mus_global_path, "MusicPlaylistContainer")

    # 查找state/switch，再跟表格比对有没有，没有就删除资源及事件引用
    delete_check(switch_have_dict, switch_create_list, "SwitchContainer")
    delete_check(playlist_have_dict, playlist_create_list, "MusicPlaylistContainer")
    delete_check(unit_have_dict, unit_create_list, "WorkUnit")

    oi_h.print_warning("结构已自动创建完成，请手动指派State")

    """******************音频资源导入********************"""
    """获取表格中的事件描述和状态所在的列"""


    def get_descrip_and_status_column():
        bpm_column = None
        require_name_column = None
        status_column = None
        if list(sheet.rows)[0]:
            for cell in list(sheet.rows)[0]:
                if cell.value:
                    if '资源名称' == str(cell.value):
                        require_name_column = cell.column
                        # print(require_name_column)
                    elif '资源描述' == str(cell.value):
                        status_column = cell.column
                        # print(status_column)
                    elif 'BPM' == str(cell.value):
                        bpm_column = cell.column

        return require_name_column, status_column, bpm_column


    # 记录所有资源名称
    sound_name_list = []

    """查找对象"""


    def find_obj(args):
        options = {
            'return': ['name', 'id', 'path', 'notes', 'originalWavFilePath', 'type', 'parent', "Tempo"]

        }
        obj_sub_list = client.call("ak.wwise.core.object.get", args, options=options)['return']
        if not obj_sub_list:
            obj_sub_id = ""
            obj_sub_path = ""
        else:
            obj_sub_id = obj_sub_list[0]['id']
            obj_sub_path = obj_sub_list[0]['path']
        return obj_sub_list, obj_sub_id, obj_sub_path


    """若要改名,mustrack需要单独处理"""


    def change_name_by_musictrack(name, old_name, wwise_new_path, sheet_name):
        # 源original的音乐资源路径
        old_path = os.path.join(wwise_mus_original_path, old_name + ".wav")
        new_path = os.path.join(wwise_mus_original_path, name + ".wav")
        # 媒体资源重命名
        os.rename(old_path, new_path)
        client.call("ak.wwise.core.audio.import",
                    waapi_h.args_mus_create(new_path, wwise_new_path, sheet_name))


    """修改Wwise内容的名字"""


    def change_name_by_wwise_content(obj_dict, new_name, obj_type, obj_parent_path):
        # 还需要重新修改媒体资源名字并导入
        if obj_type == "MusicTrack":
            # 旧的mustrack需要记录，然后删除
            notes = obj_dict['notes']
            wwise_old_path = obj_dict['path']
            wwise_old_name = obj_dict['name']
            delete_obj(obj_dict['id'], new_name, obj_type, False)

            # 生成新的mustrack
            new_obj = create_wwise_obj(obj_type, obj_parent_path, new_name, notes)
            if new_obj:
                new_obj_id = new_obj['id']
                obj_list = client.call("ak.wwise.core.object.get",
                                       waapi_h.waql_from_id(new_obj_id),
                                       options=config.options)['return']
                new_obj_path=obj_list[0]['path']
                print(new_obj_path)
                change_name_by_musictrack(new_name, wwise_old_name, new_obj_path, sheet_name)


        else:
            args = {
                "objects": [
                    {

                        "object": obj_dict['id'],
                        "name": new_name,
                    }
                ]
            }
            client.call("ak.wwise.core.object.set", args)
        oi_h.print_warning(obj_dict['name'] + "(" + obj_type + ")改名为：" + new_name)


    """通用内容创建"""


    def create_obj_content(obj_parent_path, obj_type,
                           obj_name, obj_desc, state_id):
        flag = 0
        obj_id = ""
        obj_path = ""
        obj_list = client.call("ak.wwise.core.object.get",
                               waapi_h.waql_by_type(obj_type, obj_parent_path),
                               options=config.options)['return']
        # obj已存在
        for obj_dict in obj_list:
            if obj_dict['name'] == obj_name:
                flag = 1
                # 设置notes
                set_obj_notes(obj_dict['id'], obj_desc)
                obj_id = obj_dict['id']
                obj_path = obj_dict['path']
                # 设置bpm
                if (obj_dict['type'] == "MusicSegment") and bpm_value and bpm_value != 0:
                    set_obj_property(obj_dict['id'], "OverrideClockSettings", True)
                    set_obj_property(obj_dict['id'], "Tempo", bpm_value)
                break
            # 改名了但描述一致
            elif obj_dict['notes'] == obj_desc:
                flag = 2
                # 改名
                change_name_by_wwise_content(obj_dict, obj_name, obj_type, obj_parent_path)

                obj_id = obj_dict['id']
                obj_path = obj_dict['path']
                break
        # obj不存在，需要创建
        # print(flag)
        # 不存在则创建
        if flag == 0:
            if obj_type == "Event":
                # 专属trig用
                args = {
                    # 选择父级
                    "parent": obj_parent_path,
                    # 创建类型名称
                    "type": obj_type,
                    "name": obj_name,
                    "notes": obj_desc,
                    "children": [
                        {
                            "name": "",
                            "type": "Action",
                            "@ActionType": 35,
                            "@Target": state_id
                        }]
                }
            else:
                # 创建的rnd属性
                args = {
                    # 选择父级
                    "parent": obj_parent_path,
                    # 创建类型名称
                    "type": obj_type,
                    "name": obj_name,
                    "notes": obj_desc,
                }
            obj_object = client.call("ak.wwise.core.object.create", args)
            # print(obj_object)
            obj_id = obj_object['id']
            _, _, obj_path = find_obj(
                {'waql': ' "%s" ' % obj_id})
            oi_h.print_warning(obj_object['name'] + ":" + obj_type + "已创建")

        return obj_id, obj_path


    """需要导入的音乐资源路径查找"""


    def get_mus_wav_path(mus_name):
        for media_info_name in media_info_dict:
            check_name = media_info_name.replace(".wav", "")
            if mus_name == check_name:
                return media_info_dict[media_info_name]


    """创建新的mus"""


    def create_mus_content(mus_name, mus_desc):
        # 如果没有资源但改名，需要同步引用更改
        flag = 0
        for playlist_have_name in playlist_have_dict:
            if playlist_have_name in mus_name:
                mus_segment_id, _ = create_obj_content(playlist_have_dict[playlist_have_name], "MusicSegment",
                                                       mus_name,
                                                       mus_desc, "")
                mus_track_id, mus_track_path = create_obj_content(mus_segment_id, "MusicTrack", mus_name,
                                                                  mus_desc, "")

                # 媒体资源导入
                media_path = get_mus_wav_path(mus_name)
                if media_path:
                    client.call("ak.wwise.core.audio.import",
                                waapi_h.args_mus_create(media_path, mus_track_path, sheet_name))
                    oi_h.print_warning((mus_name + ":媒体资源已导入"))

                flag = 1
                break
        if flag == 0:
            oi_h.print_error(mus_name + "：无相应前缀的playlist父级，请检查命名是否有误或先创建结构")


    """***************音乐资源导入主程序***************"""

    # 提取规则：只提取xlsx文件
    for i in file_name_list:
        if ".xlsx" in i:
            # 拼接xlsx的路径
            file_path_xlsx = os.path.join(excel_path, i)
            # 获取xlsx的workbook
            wb = openpyxl.load_workbook(file_path_xlsx)
            # 获取xlsx的所有sheet
            sheet_names = wb.sheetnames
            # 加载所有工作表
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                word_list_len = sheet.max_column
                value, value_desc, bpm = get_descrip_and_status_column()
                if value:
                    for cell in list(sheet.columns)[value - 1]:
                        if (cell.value) and (not check_is_chinese(cell.value)) and (
                                cell.value != "Mus_Login") and (check_by_length_and_word_bool(
                            cell.value, 10)):
                            """名称查重"""
                            if cell.value not in mus_name_list:
                                mus_name_list.append(cell.value)
                                # pprint(cell.value)
                                """资源描述"""
                                value_desc_value = sheet.cell(row=cell.row,
                                                              column=value_desc).value
                                """资源描述检查"""
                                if value_desc_value:
                                    if value_desc_value not in value_desc_list:
                                        value_desc_list.append(value_desc_value)
                                        """bpm获取"""
                                        bpm_value, _ = excel_h.check_is_mergecell(sheet.cell(row=cell.row,
                                                                                             column=bpm), sheet)
                                        """音频资源创建"""
                                        create_mus_content(cell.value,
                                                           value_desc_value)
                                    else:
                                        oi_h.print_error(value_desc_value + "：表格中有重复项描述，请检查")
                                else:
                                    oi_h.print_error(cell.value + "：描述不能为空，请补充")

                            else:
                                oi_h.print_error(cell.value + "：表格中有重复项描述，请检查")

    """同步表中删除的内容"""
    musegment_have_dict = get_wwise_type_list(config.wwise_mus_global_path, "MusicSegment")

    # 查找state/switch，再跟表格比对有没有，没有就删除资源及事件引用
    delete_check(musegment_have_dict, mus_name_list, "MusicSegment")

    # 清除复制的媒体资源
    # shutil.rmtree(os.path.join(root_path, 'New_Media'))
    # os.mkdir(os.path.join(root_path, 'New_Media'))
    oi_h.delete_type_files(os.path.join(root_path, "New_Media"), '.wav')



os.system("pause")
