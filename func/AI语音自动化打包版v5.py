import datetime

print(f"AIAudioTasks Start {datetime.datetime.now()}")
# xml读取库
import shutil
import os
import sys
from os.path import abspath, dirname
import openpyxl
import re
# xml写入库
from xml.dom.minidom import Document, parse
from pprint import pprint

# 自定义库
import comlib.excel_h as excel_h
import comlib.config as config
import comlib.oi_h as oi_h
import comlib.csv_h as csv_h
import comlib.exe_h as exe_h
import comlib.cloudfeishu_h as cloudfeishu_h

# 新特性：效率提升，多线程下载在线表、激活特定语言的导出
"""****************数据获取******************"""

# 获取文件所在目录
py_path = ""
if hasattr(sys, 'frozen'):
    py_path = dirname(sys.executable)
elif __file__:
    py_path = dirname(abspath(__file__))

# 飞书在线文档下载
media_sheet_token_dict = config.es_sheet_token_dict
c_vo_sheet_name = 'C类（操控演员表演，接管镜头）剧情语音'
d_vo_sheet_name = 'D类（操控演员表演，不接管镜头）剧情语音'
c_vo_sheet_token = media_sheet_token_dict[c_vo_sheet_name]
d_vo_sheet_token = media_sheet_token_dict[d_vo_sheet_name]
print(c_vo_sheet_name + "：音频资源表更新")
cloudfeishu_h.download_cloud_sheet(c_vo_sheet_token,
                                   os.path.join(py_path, "Excel", c_vo_sheet_name) + '.xlsx')
print(d_vo_sheet_name + "：音频资源表更新")
cloudfeishu_h.download_cloud_sheet(d_vo_sheet_token,
                                   os.path.join(py_path, "Excel", d_vo_sheet_name) + '.xlsx')

# 获取wsources文件
external_input_path = os.path.join(py_path, config.external_input_path)

# es输出文件夹
external_output_path = config.external_output_path
external_output_win_path = config.external_output_win_path
external_output_android_path = config.external_output_android_path
external_output_ios_path = config.external_output_ios_path

es_xml_path = os.path.join(py_path, config.es_xml_path)

# 获取媒体资源文件列表
wav_path = os.path.join(py_path, "New_Media")

# excel表标题列
excel_es_title_list = config.excel_es_title_list

# 获取当前目录下要遍历的表路径列表
file_name_list = excel_h.excel_get_path_list(os.path.join(py_path, "Excel"))

_, wem_list = oi_h.get_type_file_name_and_path('.wem', external_output_path)

# 获取写死的cookie数据表
es_vo_data_dict = config.es_vo_data_dict

"""****************分表数据*****************"""
genexcel_path = os.path.join(py_path, "GenExcel")
gencsv_path = os.path.join(py_path, "GenCSV")
genexcel_source_path = os.path.join(py_path, "excel_souce.xlsx")
gencsv_source_path = os.path.join(py_path, "csv_source.csv")

# 获取源相应的excel表信息
sheet_dt_source, wb_dt_source = excel_h.excel_get_sheet(genexcel_source_path, 'Sheet1')
# 标题列索引获取
Merge_title_dict = excel_h.excel_get_all_sheet_title_column(sheet_dt_source)

"""**************写xml**************"""
# 创建一个对象
doc = Document()
# 创建根节点
root = doc.createElement("ExternalSourcesList")
root.setAttribute("SchemaVersion", "1")

# 如果保存在Wwise工程下，此Root为Wwise工程的相对路径，则不需要输入绝对路径，仅需输入.wav的名字
# 否则取文件的绝对路径
# root.setAttribute("Root", "UngeneratedExternalSources")
# 添加载Document对象上
doc.appendChild(root)

"""*****************功能检测区******************"""
"""复制excel和csv表"""


def copy_excel_and_csv(file_name):
    target_path = os.path.join(genexcel_path, file_name + ".xlsx")
    # 检查目标文件是否存在
    if not os.path.exists(target_path):
        # 检查源文件是否存在
        if os.path.exists(genexcel_source_path):
            # 复制源文件到目标文件
            shutil.copyfile(genexcel_source_path, target_path)

    target_path = os.path.join(gencsv_path, file_name + ".csv")
    # 检查目标文件是否存在
    if not os.path.exists(target_path):
        # 检查源文件是否存在
        if os.path.exists(gencsv_source_path):
            # 复制源文件到目标文件
            shutil.copyfile(gencsv_source_path, target_path)


"""新建excel表和csv表"""


def create_new_excel_and_csv():
    for i in file_name_list:
        if ".xlsx" in i:
            plot_type = ""
            if c_vo_sheet_name in i:
                plot_type = "C"
            elif d_vo_sheet_name in i:
                plot_type = "D"
            if plot_type:
                # 拼接xlsx的路径
                file_path_xlsx = os.path.join(os.path.join(py_path, "Excel"), i)
                # 获取xlsx的workbook
                wb = openpyxl.load_workbook(file_path_xlsx)
                # 获取xlsx的所有sheet
                sheet_names = wb.sheetnames
                # 加载所有工作表
                for sheet_name in sheet_names:
                    dt_plot_name = "DT_AudioPlotLanguage_" + plot_type + "_" + sheet_name
                    copy_excel_and_csv(dt_plot_name)
            else:
                print("此类型剧情未添加，无法创建excel表和csv表")


"""语音ID生成"""


def create_es_id(name):
    # 以'_'分隔字符串，组成列表
    parts = name.split('_')

    # 确保列表有足够的元素
    if len(parts) < 5:
        raise ValueError(name + "：输入字符串格式不正确")

    # 根据第3个元素确定数字
    third_element = parts[2]
    if third_element == 'C':
        number = '1'
    elif third_element == 'D':
        number = '2'
    else:
        raise ValueError(name + "以_切分思维第三个元素非C或D，请检查")

    # 拼接数字及第4和第5个元素
    id_str = number + parts[3] + parts[4]
    es_id = int(id_str)

    return es_id


"""xml创建元素"""


def xml_create_element(media_path):
    source = doc.createElement("Source")
    source.setAttribute("Path", media_path)
    if "CG" in media_path:
        source.setAttribute("Conversion", "CG")
    elif "VO" in media_path:
        source.setAttribute("Conversion", "VO")
    root.appendChild(source)


"""写入excel_DT_AudioPlotInfo_path表"""


def write_dt_excel(vo_id, cell_sound, dt_sheet, cell_es_type):
    if cell_es_type in config.es_event_id_dict:
        event_id = config.es_event_id_dict[cell_es_type]
        # 插入为空的行
        insert_row = dt_sheet.max_row + 1
        value_dict = {
            "Row": vo_id,
            'PlotID': vo_id,
            'Extern Source Name': cell_sound.value + ".wem",
            'Extern Source ID': vo_id,
            'Event ID': event_id,

        }
        for cell in list(dt_sheet.rows)[0]:
            if cell.value in value_dict:
                dt_sheet.cell(row=insert_row, column=cell.column).value = value_dict[cell.value]
    else:
        oi_h.print_error(cell_sound.value + ":ES Type请映射请补充在config")


"""写入Merge excel表"""


def write_DT_Merge_excel(vo_id, cell_sound, cell_es_type, external_cookiedata, sheet_DT_Merge, is_ai):
    if cell_es_type in config.es_event_id_dict:
        # 插入为空的行
        insert_row = sheet_DT_Merge.max_row + 1

        # 先查找是否有该id，有则选取该row
        for cell_id in list(sheet_DT_Merge.columns)[0]:
            if cell_id.value == vo_id:
                insert_row = cell_id.row
                break

        event_id = config.es_event_id_dict[cell_es_type]
        value_dict = {
            "Row": vo_id,
            'PlotID': vo_id,
            'Media Info Id': vo_id,
            'Media Name': cell_sound.value + ".wem",
            'Event ID': event_id,
            'ExternalSourceCookie': int(external_cookiedata[1]),
            'ExternalSourceName': str(external_cookiedata[0]),
            'CodecID': 4,
            'bIsStreamed': 'TRUE',
            'bUseDeviceMemory': "FALSE",
            'MemoryAlignment': 0,
            'PrefetchSize': 0,
            'bIsAI': is_ai
        }
        for cell in list(sheet_DT_Merge.rows)[0]:
            if cell.value in value_dict:
                sheet_DT_Merge.cell(row=insert_row, column=cell.column).value = value_dict[cell.value]
    else:
        oi_h.print_error(cell_sound.value + ":ES Type请映射请补充在config")


"""删除相应的.wem文件"""


def delete_cancel_wem(media_name):
    for wem_dict in wem_list:
        for wem_name in wem_dict:
            # print(wem_name)
            if media_name in wem_name:
                if os.path.isfile(wem_dict[wem_name]):
                    os.remove(wem_dict[wem_name])
                    oi_h.print_warning("wem文件删除：" + str(wem_dict[wem_name]))


"""移除标记为cancel的所有内容"""


def delete_cancel_content():
    # 提取规则：只提取xlsx文件
    for i in file_name_list:
        if ".xlsx" in i:
            plot_type = ""
            if c_vo_sheet_name in i:
                plot_type = "C"
            elif d_vo_sheet_name in i:
                plot_type = "D"
            if plot_type:
                # 拼接xlsx的路径
                file_path_xlsx = os.path.join(os.path.join(py_path, "Excel"), i)
                # 获取xlsx的workbook
                wb = openpyxl.load_workbook(file_path_xlsx)
                # 获取xlsx的所有sheet
                sheet_names = wb.sheetnames
                # 加载所有工作表
                for sheet_name in sheet_names:
                    # 统计当前ES资源表的文件名，用于删除资源（dt表中有此列表中没有的名字将会被删除)
                    sheet_file_name_list = []
                    # 统计当前ES资源表中被标记为cancel的文件名（dt表中有此列表中的名字将会被删除)
                    cancel_file_name_list = []
                    sheet = wb[sheet_name]
                    dt_plot_name = "DT_AudioPlotLanguage_" + plot_type + "_" + sheet_name
                    dt_plot_excel_path = os.path.join(genexcel_path, dt_plot_name + '.xlsx')
                    dt_plot_csv_path = os.path.join(gencsv_path, dt_plot_name + '.csv')
                    if os.path.exists(dt_plot_excel_path):
                        sheet_DT_Merge, wb_DT_Merge = excel_h.excel_get_sheet(dt_plot_excel_path, 'Sheet1')
                        if not excel_h.is_sheet_empty(sheet):
                            title_colunmn_dict = excel_h.excel_get_sheet_title_column(sheet, excel_es_title_list)
                            if title_colunmn_dict:
                                # 获取音效名下的内容
                                for cell_sound in list(sheet.columns)[
                                    excel_h.sheet_title_column("文件名", title_colunmn_dict) - 1]:
                                    if cell_sound.value and (cell_sound.value != "文件名"):
                                        # 统计当前ES资源表的文件名
                                        if cell_sound.value not in sheet_file_name_list:
                                            sheet_file_name_list.append(cell_sound.value)
                                            if sheet.cell(row=cell_sound.row,
                                                          column=excel_h.sheet_title_column("State",
                                                                                            title_colunmn_dict)).value == 'cancel':
                                                # 统计当前ES资源表中标记为cancel的文件名
                                                if cell_sound.value not in cancel_file_name_list:
                                                    cancel_file_name_list.append(cell_sound.value)
                        # 进行删除操作
                        delete_cancel_content_action(sheet_file_name_list, cancel_file_name_list, sheet_DT_Merge,
                                                     wb_DT_Merge,
                                                     dt_plot_excel_path,
                                                     dt_plot_csv_path)


"""具体操作删除内容"""


def delete_cancel_content_action(sheet_file_name_list, cancel_file_name_list, sheet_DT_Merge, wb_DT_Merge,
                                 excel_DT_Merge_path,
                                 csv_DT_Merge_path):
    # pprint(table_name_list)
    # 记录要删除的行的索引
    merge_del_list = []

    # 删除excel表的相应内容
    # es资源表中没有的添加到merge_del_list
    for row in range(2, sheet_DT_Merge.max_row + 1):
        cell = sheet_DT_Merge.cell(row=row, column=Merge_title_dict['Media Name'])
        # pprint(cell.value)
        flag = 0
        if cell.value:
            for sheet_file_name in sheet_file_name_list:
                # 在media_info表中找到该ID，则标记
                if sheet_file_name in cell.value:
                    is_cancel = False
                    # cancel列表中有的添加到merge_del_list
                    for cancel_file_name in cancel_file_name_list:
                        if cancel_file_name in cell.value:
                            if cell.row not in merge_del_list:
                                merge_del_list.append(cell.row)
                                oi_h.print_warning(cell.value + ":ES资源表中文件被标记为cancel，在ID表中删除相关信息")
                                delete_cancel_wem(cell.value)
                                is_cancel = True
                                flag = 2
                                break
                    if is_cancel == False:
                        flag = 1
                        break
            # 未找到media name的则删除
            if flag == 0:
                if cell.row not in merge_del_list:
                    merge_del_list.append(cell.row)
                    oi_h.print_warning(cell.value + ":ES资源表中该文件名不存在，在ID表中删除相关信息")
                    delete_cancel_wem(cell.value)

    # 列表多行删除，为避免行数出问题，需倒序删除
    # 从最后一行开始删除
    for row in sorted(merge_del_list, reverse=True):
        sheet_DT_Merge.delete_rows(row)
    # 保存excel表的信息
    wb_DT_Merge.save(excel_DT_Merge_path)
    # excel转csv
    csv_h.excel_to_csv(excel_DT_Merge_path, csv_DT_Merge_path)


"""检查是否以CG_External_或VO_External_开头"""


# 返回值为True或False

def check_prefix(string):
    pattern = r'^(CG_External_|VO_External_)'
    return re.match(pattern, string) is not None


"""自动化生成ES表"""


def auto_gen_es_dt():
    # 读excel表
    # 提取规则：只提取xlsx文件
    for i in file_name_list:
        if ".xlsx" in i:
            plot_type = ""
            if c_vo_sheet_name in i:
                plot_type = "C"
            elif d_vo_sheet_name in i:
                plot_type = "D"
            if plot_type:
                # 拼接xlsx的路径
                file_path_xlsx = os.path.join(os.path.join(py_path, "Excel"), i)
                # 获取xlsx的workbook
                wb = openpyxl.load_workbook(file_path_xlsx)
                # 获取xlsx的所有sheet
                sheet_names = wb.sheetnames
                # 加载所有工作表
                for sheet_name in sheet_names:
                    sheet = wb[sheet_name]
                    dt_plot_name = "DT_AudioPlotLanguage_" + plot_type + "_" + sheet_name
                    dt_plot_excel_path = os.path.join(genexcel_path, dt_plot_name + '.xlsx')
                    dt_plot_csv_path = os.path.join(gencsv_path, dt_plot_name + '.csv')
                    if os.path.exists(dt_plot_excel_path):
                        sheet_DT_Merge, wb_DT_Merge = excel_h.excel_get_sheet(dt_plot_excel_path, 'Sheet1')
                        if not excel_h.is_sheet_empty(sheet):
                            title_colunmn_dict = excel_h.excel_get_sheet_title_column(sheet, excel_es_title_list)
                            if title_colunmn_dict:
                                # 获取音效名下的内容
                                for cell_sound in list(sheet.columns)[
                                    excel_h.sheet_title_column("文件名", title_colunmn_dict) - 1]:
                                    if cell_sound.value and (cell_sound.value != "文件名"):
                                        is_wwise_have = 0
                                        # 查找字典里有无对应的ES容器
                                        for external_sound in es_vo_data_dict:
                                            # for external_sound_dict in external_sound_list:
                                            cell_es_type = sheet.cell(
                                                row=cell_sound.row,
                                                column=excel_h.sheet_title_column("External_Type",
                                                                                  title_colunmn_dict)).value
                                            if external_sound == cell_es_type:
                                                # 查找mediainfo表的id有没有，没有则添加，有则修改内容
                                                vo_id = create_es_id(cell_sound.value)

                                                # is_ai生成
                                                is_ai = 'FALSE'
                                                if sheet.cell(row=cell_sound.row,
                                                              column=excel_h.sheet_title_column("State",
                                                                                                title_colunmn_dict)).value == 'AI':
                                                    is_ai = 'TRUE'

                                                # 将数据写入DT excel表
                                                write_DT_Merge_excel(vo_id, cell_sound,
                                                                     cell_es_type,
                                                                     es_vo_data_dict[
                                                                         external_sound],
                                                                     sheet_DT_Merge, is_ai)

                                                is_wwise_have = 1
                                                break
                                        if is_wwise_have == 0:
                                            oi_h.print_error(
                                                cell_sound.value + "：在Wwise中未建立相应容器，请检查填写是否正确或Wwise中容器是否存在")

                        # 这里保存xlsx和csv的信息
                        # 保存excel表的信息
                        wb_DT_Merge.save(dt_plot_excel_path)

                        # 将excel转为csv
                        csv_h.excel_to_csv(dt_plot_excel_path, dt_plot_csv_path)

                    else:
                        print(dt_plot_excel_path + "：excel表路径有问题，请检查")
                wb.save(file_path_xlsx)


"""自动化生成ES数据"""


def auto_gen_es_file(file_wav_dict):
    # 查找要导入的媒体文件里有没有对应的
    for file_wav_name in file_wav_dict:
        # 读excel表
        # 提取规则：只提取xlsx文件
        for i in file_name_list:
            if ".xlsx" in i:
                plot_type = ""
                if c_vo_sheet_name in i:
                    plot_type = "C"
                elif d_vo_sheet_name in i:
                    plot_type = "D"
                if plot_type:
                    # 拼接xlsx的路径
                    file_path_xlsx = os.path.join(os.path.join(py_path, "Excel"), i)
                    # 获取xlsx的workbook
                    wb = openpyxl.load_workbook(file_path_xlsx)
                    # 获取xlsx的所有sheet
                    sheet_names = wb.sheetnames
                    # 加载所有工作表
                    for sheet_name in sheet_names:
                        sheet = wb[sheet_name]
                        dt_plot_name = "DT_AudioPlotLanguage_" + plot_type + "_" + sheet_name
                        dt_plot_excel_path = os.path.join(genexcel_path, dt_plot_name + '.xlsx')
                        dt_plot_csv_path = os.path.join(gencsv_path, dt_plot_name + '.csv')
                        if os.path.exists(dt_plot_excel_path):
                            sheet_DT_Merge, wb_DT_Merge = excel_h.excel_get_sheet(dt_plot_excel_path, 'Sheet1')
                            if not excel_h.is_sheet_empty(sheet):
                                title_colunmn_dict = excel_h.excel_get_sheet_title_column(sheet, excel_es_title_list)
                                if title_colunmn_dict:
                                    # 获取音效名下的内容
                                    for cell_sound in list(sheet.columns)[
                                        excel_h.sheet_title_column("文件名", title_colunmn_dict) - 1]:
                                        if cell_sound.value and (cell_sound.value != "文件名"):
                                            """查找文件"""
                                            if cell_sound.value in file_wav_name:
                                                if sheet.cell(row=cell_sound.row,
                                                              column=excel_h.sheet_title_column("State",
                                                                                                title_colunmn_dict)).value == 'AI':
                                                    # print(file_wav_dict[file_wav_name])
                                                    # 在表格中找到此音效才将其写入xml
                                                    xml_create_element(file_wav_dict[file_wav_name])
                                                    oi_h.print_warning(file_wav_name + "：已导入")
                                                    break


"""分目录移动文件夹"""


def move_file_to_different_catalog(source_directory_path):
    # 遍历源目录中的所有文件
    for filename in os.listdir(source_directory_path):

        # 构建文件的完整路径
        file_path = os.path.join(source_directory_path, filename)
        # print(file_path)

        # 只处理文件，忽略子目录
        if os.path.isfile(file_path) and filename.endswith('.wem'):
            # print(file_path)
            # 提取文件名中的标识部分
            parts = filename.split('_')
            if len(parts) > 3:
                identifier = f'{parts[2]}_{parts[3]}'

                # 构建目标文件夹路径
                target_directory = os.path.join(source_directory_path, f'AudioPlotLanguage_{identifier}')

                # print(target_directory)
                # 确保目标文件夹存在
                if not os.path.exists(target_directory):
                    os.makedirs(target_directory)

                # 构建目标文件的完整路径
                target_file = os.path.join(target_directory, filename)

                # 移动文件
                shutil.move(file_path, target_file)

                # print(f'Moved {filename} to {target_directory}')


"""*******************主程序*******************"""

table_name_list = []
# 新建excel和csv表
create_new_excel_and_csv()

print(f"开始创建excel和csv表 {datetime.datetime.now()}")
# 创建ES DT表
auto_gen_es_dt()
print(f"结束创建excel和csv表 {datetime.datetime.now()}")

# 语音ES生成
for language in config.language_list:
    wav_language_path = os.path.join(wav_path, language)
    # 要文件夹路径存在及文件夹中存在内容才启动以下操作
    if oi_h.is_folder_empty(wav_language_path):
        file_wav_language_dict, _ = oi_h.get_type_file_name_and_path('.wav', wav_language_path)
        # print(file_wav_language_dict)
        auto_gen_es_file(file_wav_language_dict)
        # # xml文件写入
        # pprint(file_wav_dict)
        with open(es_xml_path, 'w+') as f:
            # 按照格式写入
            f.write(doc.toprettyxml())
            f.close()
        # pprint(doc.toprettyxml())
        # 复制xml为wsources
        shutil.copy2(es_xml_path,
                     external_input_path)

        # gen_external(language)
        exe_h.gen_ai_language(external_input_path, language)

        # 将生成资源移动到每个dt表对应的路径
        move_file_to_different_catalog(os.path.join(external_output_win_path, language, 'AILanguage'))
        move_file_to_different_catalog(os.path.join(external_output_android_path, language, 'AILanguage'))
        move_file_to_different_catalog(os.path.join(external_output_ios_path, language, 'AILanguage'))

        # 初始化xml
        doc = Document()
        root = doc.createElement("ExternalSourcesList")
        root.setAttribute("SchemaVersion", "1")
        doc.appendChild(root)

        # 将删除的xml内容写入wsources
        shutil.copy2(es_xml_path,
                     external_input_path)

# 删除xml的内容
doc = parse(es_xml_path)
parent_node = doc.getElementsByTagName('ExternalSourcesList')[0]
while parent_node.hasChildNodes():
    parent_node.removeChild(parent_node.firstChild)

# 可选：如果需要，将更改保存回文件
with open(es_xml_path, 'w', encoding='utf-8') as f:
    doc.writexml(f)

# 将删除的xml内容写入wsources
shutil.copy2(es_xml_path, external_input_path)

print(f"创建wem资源完毕 {datetime.datetime.now()}")

# 文件清理
delete_cancel_content()

# else:
#     print(1)

# 清除复制的媒体资源
shutil.rmtree(wav_path)
os.mkdir(wav_path)
os.mkdir(os.path.join(wav_path, "Chinese"))
os.mkdir(os.path.join(wav_path, "English"))
os.mkdir(os.path.join(wav_path, "Japanese"))
os.mkdir(os.path.join(wav_path, "Korean"))
# os.mkdir(os.path.join(wav_path, "SFX"))

print(f"无效文件清理 {datetime.datetime.now()}")

# dt表导入UE
ue_csv_dt_es_path = os.path.join(py_path, "ue_csv_dt_es.py")
exe_h.run_unreal_editor_with_script(ue_csv_dt_es_path)
# os.system("pause")
print(f"在ue中导入dt表 {datetime.datetime.now()}")
