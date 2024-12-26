import openpyxl
import json
import sys
from os.path import abspath, dirname
import os
from pprint import pprint
import re
from waapi import WaapiClient
import shutil
from openpyxl.cell import MergedCell
import module.cloudfeishu.cloudfeishu_h as cloudfeishu_h
import module.oi.oi_h as oi_h
import comlib.config as config
import module.excel.excel_h as excel_h
import module.waapi.waapi_h as waapi_h
from pathlib import Path
import 命名规范检查

"""根目录获取"""
# 获取当前脚本的文件名及路径
script_name_with_extension = __file__.split('/')[-1]
# 去掉扩展名
script_name = script_name_with_extension.rsplit('.', 1)[0]
# 替换为files
root_path = script_name.replace("func", "files")
# print(f"当前脚本的名字是: {root_path}")c

file_name_list = excel_h.excel_get_path_list(root_path)
# pprint(file_name_list)

"""表格获取输入"""
# 所有音频资源表的映射
media_sheet_token_dict = config.media_sheet_token_dict
media_sheet_token_keys = list(media_sheet_token_dict.keys())
# 将一个list转为dict，list的index+1作为键，元素作为值
digi_meidia_dict = {index + 1: value for index, value in enumerate(media_sheet_token_keys)}


# print(digi_meidia_dict)


# python输入检查
# 1.必须为数字和以,为间隔，例如1，4，2，6
# 2.通过,分隔的数字不能重复
def is_valid_input(input_string):
    # 删除空格
    input_string = input_string.replace(" ", "")

    # 检查是否仅包含数字和逗号
    if not all((c.isdigit() or c == ',') for c in input_string):
        print("输入错误，请确保输入的只有数字且以,隔开")
        return False

    # 拆分字符串并转换为整数列表
    try:
        numbers = list(map(int, input_string.split(',')))
    except ValueError:
        print("输入错误，请确保输入的只有数字且以,隔开")
        return False

    # 检查是否有重复
    if len(numbers) != len(set(numbers)):
        print("输入错误，请输无重复的数字")
        return False

    for numbers in numbers:
        if numbers > len(media_sheet_token_keys):
            print("输入错误，请输入上述表中有的数字")
            return False

    return True


print("--------------------------")
print("输入数字对应资源表类型如下：")
print("输入值若为0，代表更新所有资源表内容")
print("若需获取多个资源表，以英文字符,(逗号）隔开，输入完毕后按下回车即可")
print("输入案例：1,3,5,6")

for i in range(len(media_sheet_token_keys)):
    print("{}、".format(i + 1) + str(media_sheet_token_keys[i]))
print("--------------------------")

temp = ''
flag = 1
while flag:
    temp = input("请输入要更新资源的音效表数字:")
    if not is_valid_input(temp):
        print("")
        print("请重新输入⬇")
        continue
    flag = 0

input_digi_list = list(map(int, temp.split(',')))
for i in input_digi_list:
    if i in digi_meidia_dict:
        media_sheet_name = digi_meidia_dict[i]
        if media_sheet_name in media_sheet_token_dict:
            print(media_sheet_name + "：音频资源表更新")
            sheet_token = media_sheet_token_dict[media_sheet_name]
            cloudfeishu_h.download_cloud_sheet(sheet_token, os.path.join(root_path, media_sheet_name))

"""在线表获取"""
# 规范检查表
# for excel_media_token in config.excel_media_token_list:
#     _, excel_name = cloudfeishu_h.get_excel_token(excel_media_token)
#     # pprint(excel_name)
#     for file_name in file_name_list:
#         if excel_name in file_name:
#             print(excel_name)
#             cloudfeishu_h.download_cloud_sheet(excel_media_token, os.path.join(root_path, file_name))
#             break

"""收集unit的路径"""
audio_unit_list = []
event_unit_list = []
audio_mixer_list = []

"""*****************功能检测区******************"""
"""event结构找父级测试"""


# python检查 `list1` 中每个以“_”拆分后的字符串组成的列表的所有元素是否全部存在于 `list2` 中某个以“_”拆分后的字符串组成的列表中,
# 若存在，则继续判断选出“list1”中最长的字符串，与“list2”的字符串生成一字典
# 将list2改为dict2的键
def find_event_parent_match(dict1, string):
    # 将字符串拆分为列表
    string_list = string.split('_')

    # 过滤符合条件的键
    valid_keys = [
        key for key in dict1.keys()
        if all(part in string_list for part in key.split('_'))
    ]

    if valid_keys:
        # 找出最长的键
        longest_key = max(valid_keys, key=len)
        return longest_key
    else:
        return None


"""获取event_unit_list中缺失的unit"""


# 遍历 `list1`：对于 `list1` 中的每个元素，将其用下划线分隔，并转换为集合 `parts1`。
# 遍历 `list2`：对于 `list2` 中的每个元素，同样进行分隔并转换为集合 `parts2`。
# 检查子集关系：使用 `issubset` 方法检查 `parts1` 是否是 `parts2` 的子集。
# 更新结果列表：如果 `parts1` 是 `parts2` 的子集，则将 `list2` 中的该元素添加到 `result` 列表中。
# 输出结果：最终输出更新后的列表 `result`。
def get_missing_event_unit(list1, list2):
    for element1 in list1[:]:  # 使用 list1[:] 复制列表以避免修改时的迭代问题
        parts1 = set(element1.split('_'))
        world_list1 = element1.split('_')
        flag = 0
        # 至少要有两层_的才运算
        if len(world_list1) > 1:
            # 若为怪物，至少要三层_才运算
            if ("Mon_Mob" in element1) or ("Mon_Boss" in element1) or ("VO_Game" in element1):
                if len(world_list1) > 2:
                    flag = 1
            else:
                flag = 1
        if flag == 1:
            for element2 in list2:
                parts2 = set(element2.split('_'))
                if parts1.issubset(parts2):  # 检查是否有共同元素
                    if element2 not in list1:
                        list1.append(element2)
                    # break  # 如果已经匹配到一个，跳出内层循环


"""获取最长共同前缀"""


def longest_common_prefix(s1, s2):
    """Helper function to find the longest common prefix of two strings."""
    min_length = min(len(s1), len(s2))
    for i in range(min_length):
        if s1[i] != s2[i]:
            return s1[:i]
    return s1[:min_length]


"""在字典中查找最长共同前缀的值，且字典值长度需小于target"""


def find_longest_prefix_key(target, dictionary):
    max_prefix_length = 0
    best_key = None

    for key in dictionary.keys():
        if key in target:
            if len(key) < len(target):
                prefix = longest_common_prefix(target, key)
                if len(prefix) > max_prefix_length:
                    max_prefix_length = len(prefix)
                    best_key = key

    return best_key


"""获取表格中的事件描述和状态所在的列"""


def get_descrip_and_status_column():
    require_module_column = None
    second_module_column = None
    require_name_column = None
    status_column = None
    sample_name_column = None

    if list(sheet.rows)[0]:
        for cell in list(sheet.rows)[0]:
            if cell.value:
                if 'Require Module' in str(cell.value):
                    require_module_column = cell.column
                    # print(require_module_column)
                elif 'Second Module' in str(cell.value):
                    second_module_column = cell.column
                    # print(second_module_column)
                elif 'Require Name' in str(cell.value):
                    require_name_column = cell.column
                    # print(require_name_column)
                elif 'Status' in str(cell.value):
                    status_column = cell.column
                    # print(status_column)
                elif 'Sample Name' in str(cell.value):
                    sample_name_column = cell.column
                    # print(status_column)

    return require_module_column, second_module_column, require_name_column, status_column, sample_name_column


with WaapiClient() as client:
    """*****************Wwise功能区******************"""
    """获取Wwise中的Unit列表"""


    def get_wwise_type_list(root_path, type):
        # 获取Wwise所有的Audio Unit
        wwise_audio_unit_all_list = client.call("ak.wwise.core.object.get",
                                                waapi_h.waql_by_type(type, root_path),
                                                options=config.options)['return']
        wwise_audio_unit_name_dict = {item['name']: item['id'] for item in wwise_audio_unit_all_list}
        # pprint(wwise_audio_unit_name_dict)
        # {'Amb': '{4BB49221-4DC1-4A05-A57C-3D42FA570675}',
        #  'Amb_A01': '{415AC456-D390-43D5-A09E-B5D30B253D76}',
        #  'CG': '{BBB332B4-C230-4D60-9ABA-91A36D968C11}'}
        return wwise_audio_unit_name_dict


    """obj创建"""


    def create_wwise_obj(type, parent, name, notes):
        unit_object = client.call("ak.wwise.core.object.create",
                                  waapi_h.args_object_create(parent,
                                                             type, name, notes))
        oi_h.print_warning(name + "：" + type + "已创建")
        return unit_object


    """eventunit结构创建"""


    def create_wwise_event_unit(event_unit_list):
        unit_object = {}
        # 查找Event中是否有该unit（第一层级）
        for audio_unit in event_unit_list:
            wwise_audio_unit_name_dict = get_wwise_type_list(config.wwise_event_path, "WorkUnit")
            # 若Wwise中无相应Unit则需要创建
            if audio_unit not in wwise_audio_unit_name_dict:
                event_unit_parent = find_event_parent_match(wwise_audio_unit_name_dict, audio_unit)
                if event_unit_parent:
                    # print(audio_unit + ":" + event_unit_parent)
                    # WorkUnit创建
                    unit_object = create_wwise_obj("WorkUnit",
                                                   wwise_audio_unit_name_dict[event_unit_parent], audio_unit, "")
                # 若未找到父级则在根目录创建
                else:
                    unit_object = create_wwise_obj("WorkUnit",
                                                   config.wwise_event_path, audio_unit, "")


    """audiounit结构创建"""


    # 有2种方式
    # 1：audio下，unit和actormixer都需要创建
    # 2：audio下，只需要创建actormixer
    def create_wwise_audio_unit(unit_list, flag):
        # 查找Wwise中是否有该unit
        for audio_unit in unit_list:
            wwise_audio_unit_name_dict = get_wwise_type_list(config.wwise_sfx_path, "ActorMixer")
            # 若Wwise中无相应Unit则需要创建
            if audio_unit not in wwise_audio_unit_name_dict:
                # 查找该unit需要放的父级是谁
                audio_unit_parent = find_longest_prefix_key(audio_unit, wwise_audio_unit_name_dict)
                # 根目录创建
                if not audio_unit_parent:
                    # print(audio_unit)
                    audio_unit_parent_obj = create_wwise_obj("WorkUnit",
                                                             config.wwise_sfx_path, audio_unit, "")
                    audio_unit_parent = audio_unit_parent_obj['id']
                else:
                    # 2：audio下，只需要创建actormixer
                    if flag == 2:
                        # ActorMixer创建
                        unit_object = create_wwise_obj("ActorMixer",
                                                       wwise_audio_unit_name_dict[audio_unit_parent], audio_unit, "")
                    elif flag == 1:
                        # WorkUnit创建
                        unit_object = create_wwise_obj("WorkUnit",
                                                       wwise_audio_unit_name_dict[audio_unit_parent], audio_unit, "")
                        # 1：audio下，unit和actormixer都需要创建
                        if unit_object:
                            # ActorMixer创建
                            unit_object = create_wwise_obj("ActorMixer",
                                                           unit_object['id'], audio_unit, "")


    """*****************主程序处理******************"""

    # 获取Wwise所有的Event Unit
    wwise_event_unit_all_list = client.call("ak.wwise.core.object.get",
                                            waapi_h.waql_by_type("WorkUnit", config.wwise_event_path),
                                            options=config.options)['return']
    wwise_event_unit_name_dict = {item['name']: item['id'] for item in wwise_event_unit_all_list}
    # pprint(wwise_event_unit_name_dict)

    # 记录所有资源名称
    sound_name_list = []

    # 提取规则：只提取xlsx文件
    for i in file_name_list:
        if ".xlsx" in i:
            # 拼接xlsx的路径
            file_path_xlsx = os.path.join(root_path, i)
            # 获取xlsx的workbook
            wb = openpyxl.load_workbook(file_path_xlsx)
            # 获取xlsx的所有sheet
            sheet_names = wb.sheetnames
            # 加载所有工作表
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                word_list_len = sheet.max_column

                require_module_column, second_module_column, require_name_column, status_column, sample_name_column = get_descrip_and_status_column()
                if sample_name_column:
                    for cell_sound in list(sheet.columns)[sample_name_column - 1]:
                        if status_column:
                            if sheet.cell(row=cell_sound.row,
                                          column=status_column).value in config.status_list:
                                # pprint(cell_sound.value)

                                if 命名规范检查.check_basic(cell_sound.value, audio_unit_list, event_unit_list,
                                                            audio_mixer_list, word_list_len):
                                    # 通过命名规范检查
                                    # pprint(cell_sound.value)
                                    pass

    """*****************work unit创建******************"""
    # 对列表进行排序
    audio_unit_list.sort(key=len)
    event_unit_list.sort(key=len)
    audio_mixer_list.sort(key=len)

    # 获取event_unit_list中缺失的unit
    get_missing_event_unit(event_unit_list, audio_unit_list)

    # # 1：audio下，unit和actormixer都需要创建
    create_wwise_audio_unit(audio_unit_list, 1)
    # # event下，unit创建
    create_wwise_event_unit(event_unit_list)
    # 2：audio下，只需要创建actormixer
    create_wwise_audio_unit(audio_mixer_list, 2)

    # pprint("audio_unit_list：")
    # pprint(audio_unit_list)
    # print()

    # pprint("event_unit_list：")
    # pprint(event_unit_list)
    # print()
    #
    # pprint("audio_mixer_list：")
    # pprint(audio_mixer_list)
