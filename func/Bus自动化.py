import os

from waapi import WaapiClient
import openpyxl
import re
from pprint import pprint

import config
import module.cloudfeishu.cloudfeishu_h as cloudfeishu_h
import module.waapi.waapi_h as waapi_h
import module.oi.oi_h as oi_h
import module.excel.excel_h as excel_h
import 命名规范检查

"""根目录获取"""
# 获取当前脚本的文件名及路径
script_name_with_extension = __file__.split('/')[-1]
# 去掉扩展名
script_name = script_name_with_extension.rsplit('.', 1)[0]
# 替换为files
root_path = script_name.replace("func", "files")
# print(f"当前脚本的名字是: {root_path}")

"""文件获取"""
# 规范检查表
excel_bus_path = os.path.join(root_path, config.excel_bus_path)
# print(wb)

"""在线表获取"""
# 规范检查表
check_name_token = config.bus_sheet_token
cloudfeishu_h.download_cloud_sheet(check_name_token, excel_bus_path)

"""功能数据初始化"""
# bus创建列表
bus_create_list = []
# bus unit创建列表
unit_create_list = []

# 获取xlsx的workbook
wb = openpyxl.load_workbook(excel_bus_path)

with WaapiClient() as client:
    """*****************Wwise功能区******************"""


    # 设置对象引用
    def set_obj_refer(obj_id, refer, value):
        args = {
            "object": obj_id,
            "reference": refer,
            "value": value
        }
        client.call("ak.wwise.core.object.setReference", args)


    """修改Wwise内容的名字"""


    def change_name_by_wwise_content(obj_id, name, old_name, obj_type):
        args = {
            "objects": [
                {

                    "object": obj_id,
                    "name": name,
                }
            ]
        }
        client.call("ak.wwise.core.object.set", args)
        oi_h.print_warning(old_name + "(" + obj_type + ")改名为：" + name)


    """设置obj的notes"""


    def set_obj_notes(obj_id, notes_value):
        args = {
            'object': obj_id,
            'value': notes_value
        }
        client.call("ak.wwise.core.object.setNotes", args)
        # print_warning(obj_name + "描述更改为：" + notes_value)


    """设置对象属性"""


    def set_obj_property(obj_id, property, value):
        args = {
            "object": obj_id,
            "property": property,
            "value": value
        }
        client.call("ak.wwise.core.object.setProperty", args)
        # print(name_2 + ":" + str(trigger_rate))


    """obj创建"""


    def create_wwise_obj(type, parent, name, notes):
        unit_object = client.call("ak.wwise.core.object.create",
                                  waapi_h.args_object_create(parent,
                                                             type, name, notes))
        oi_h.print_warning(name + "：" + type + "已创建")
        return unit_object


    """获取Wwise中的Unit列表"""


    def get_wwise_type_list(root_path, type):
        # 获取Wwise所有的Audio Unit
        wwise_bus_name_all_list = client.call("ak.wwise.core.object.get",
                                              waapi_h.waql_by_type(type, root_path),
                                              options=config.options)['return']
        wwise_bus_name_name_dict = {item['name']: item['id'] for item in wwise_bus_name_all_list}
        # pprint(wwise_bus_name_name_dict)
        # {'Amb': '{4BB49221-4DC1-4A05-A57C-3D42FA570675}',
        #  'Amb_A01': '{415AC456-D390-43D5-A09E-B5D30B253D76}',
        #  'CG': '{BBB332B4-C230-4D60-9ABA-91A36D968C11}'}
        return wwise_bus_name_name_dict


    """list中以_拆分的元素都必须在list2以_拆分的元素中"""


    # 遍历 `list1`：对于 `list1` 中的每个元素，将其用下划线分隔
    # 遍历 `list2`：对于 `list2` 中的每个元素，同样进行分隔
    # list2中的元素必须都得在list1中找到

    def check_word(list1, list2):
        list_1 = list1.split('_')
        list_2 = list2.split('_')
        for name2 in list_2:
            if name2 not in list_1:
                pprint(list1 + "：拼写有误，请检查")
                return False
        return True


    """获取最长共同前缀"""


    def longest_common_prefix(s1, s2):
        """Helper function to find the longest common prefix of two strings."""
        min_length = min(len(s1), len(s2))
        for i in range(min_length):
            if s1[i] != s2[i]:
                return s1[:i]
        return s1[:min_length]


    """在字典中查找最长共同前缀的值，且字典值长度需小于target"""


    def find_longest_prefix_key(target, dictionary):
        max_prefix_length = 0
        best_key = None

        for key in dictionary.keys():
            if key in target:
                if len(key) < len(target):
                    prefix = longest_common_prefix(target, key)
                    if len(prefix) > max_prefix_length:
                        max_prefix_length = len(prefix)
                        best_key = key
        if best_key:
            if check_word(target, best_key):
                return best_key
            else:
                return None
        return best_key


    """obj子级创建"""


    # obj_type：要创建的对象类型
    # parent_type：对象父级类型
    # obj_list：要创建对象的名称列表
    # have_obj_path：查找已有对象wwise路径

    def create_wwise_sub_obj(obj_type, parent_type, obj_list, have_obj_path):
        for obj_name in obj_list:
            bus_have_dict = get_wwise_type_list(have_obj_path, parent_type)
            aux_parent = find_longest_prefix_key(obj_name, bus_have_dict)
            if aux_parent:
                create_wwise_obj(obj_type,
                                 bus_have_dict[aux_parent], obj_name,
                                 bus_config_dict[obj_name][value_desc_column - 2])

                # print(bus_config_dict[obj_name][value_desc_column - 2])
            else:
                oi_h.print_error(obj_name + "：无相应父级，请检查命名是否正确或先创建父级bus")


    """obj子级创建"""


    # obj_type：要创建的对象类型
    # parent_type：对象父级类型
    # obj_list：要创建对象的名称列表
    # have_obj_path：查找已有对象wwise路径

    def create_wwise_obj_by_excel(obj_type, parent_type, obj_name, have_obj_path, desc_value):
        bus_have_dict = get_wwise_type_list(have_obj_path, parent_type)
        aux_parent = find_longest_prefix_key(obj_name, bus_have_dict)
        if aux_parent:
            # create_wwise_obj(obj_type,
            #                  bus_have_dict[aux_parent], obj_name,
            #                  desc_value)
            create_obj_content(bus_have_dict[aux_parent], obj_type,
                               obj_name, desc_value)

            # print(bus_config_dict[obj_name][value_desc_column - 2])
        else:
            oi_h.print_error(obj_name + "：无相应父级，请检查命名是否正确或先创建父级bus")


    """Bus创建"""


    def create_wwise_bus(unit_list, bus_list):
        if bus_list:
            for bus_name in bus_list:
                # 需要创建unit的
                if (bus_name in unit_list) and unit_list:
                    # 查找已有unit列表
                    unit_have_dict = get_wwise_type_list(config.wwise_bus_path, "WorkUnit")
                    # pprint(unit_have_dict)
                    # 查找已有bus列表
                    bus_have_dict = get_wwise_type_list(config.wwise_bus_path, "Bus")
                    # Wwise里没有unit的的
                    if bus_name not in unit_have_dict:
                        # 查找该unit需要放的父级是谁
                        bus_name_parent = find_longest_prefix_key(bus_name, unit_have_dict)
                        # 根目录创建
                        if not bus_name_parent:
                            oi_h.print_error(bus_name + "：未找到可以存放的父级Unit，请创建")
                        else:

                            # WorkUnit创建
                            unit_object = create_wwise_obj("WorkUnit",
                                                           bus_have_dict[bus_name_parent], bus_name,
                                                           "")
                            # Bus创建
                            bus_object = create_wwise_obj("Bus",
                                                          unit_object['id'], bus_name,
                                                          "")
                # 需要创建bus的
                else:
                    # 查找已有bus列表
                    bus_have_dict = get_wwise_type_list(config.wwise_bus_path, "Bus")
                    if bus_name not in bus_have_dict:
                        # 查找该unit需要放的父级是谁
                        bus_name_parent = find_longest_prefix_key(bus_name, bus_have_dict)
                        # Bus创建
                        bus_object = create_wwise_obj("Bus",
                                                      bus_have_dict[bus_name_parent], bus_name,
                                                      "")


    """获取需要创建的Bus列表"""


    def get_create_bus_name_list():
        # 遍历每一行
        for row in sheet.iter_rows():
            bus_name = None
            # 遍历每一列
            for cell in row:
                cell_value, _ = excel_h.check_is_mergecell(cell, sheet)
                if cell_value:
                    is_unit = False
                    # print(cell_value)
                    if not bus_name:
                        bus_name = cell_value
                        is_unit = True
                    else:
                        bus_name += "_" + cell_value
                        fill = cell.fill
                        if fill.fill_type == 'solid':
                            if fill.start_color.index == "FFD9F3FD":
                                is_unit = True
                    if bus_name not in bus_create_list:
                        bus_create_list.append(bus_name)
                    if is_unit:
                        if bus_name not in unit_create_list:
                            unit_create_list.append(bus_name)

            # print()


    """通用内容创建"""


    def create_obj_content(obj_parent_path, obj_type,
                           obj_name, obj_desc):
        flag = 0
        obj_id = ""
        # 查找此obj是否存在
        obj_list = client.call("ak.wwise.core.object.get",
                               waapi_h.waql_by_type(obj_type, obj_parent_path),
                               options=config.options)['return']
        # obj已存在
        for obj_dict in obj_list:
            if obj_dict['name'] == obj_name:
                flag = 1
                # 设置notes
                set_obj_notes(obj_dict['id'], obj_desc)
                obj_id = obj_dict['id']
                break
            # 改名了但描述一致
            elif obj_dict['notes'] == obj_desc:
                flag = 2
                # 改名
                change_name_by_wwise_content(obj_dict['id'], obj_name, obj_dict['name'],
                                             obj_type)
                obj_id = obj_dict['id']
                break
        # obj不存在，需要创建
        # 不存在则创建
        if flag == 0:
            obj_object = None
            if obj_type == "Effect":
                args = waapi_h.args_effect_create(obj_parent_path, 8454147,
                                                  obj_name)
                obj_list = client.call("ak.wwise.core.object.set", args)
                # pprint(obj_list)
                obj_object = obj_list['objects'][0]['children'][0]
            else:
                args = waapi_h.args_object_create(obj_parent_path, obj_type,
                                                  obj_name, obj_desc)
                obj_object = client.call("ak.wwise.core.object.create", args)
                # print(obj_object)
            obj_id = obj_object['id']
            oi_h.print_warning(obj_object['name'] + ":" + obj_type + "已创建")
        set_obj_property(obj_id, "OverrideColor", False)
        return obj_id


    """ducking metter指派"""


    def set_source_meter_and_rtpc(rtpc_name):
        flag = 0
        rtpc_id = None
        meter_id = None
        # rtpc创建
        for rtpc_unit_name in rtpc_unit_have_dict:
            pattern = f"^{re.escape(rtpc_unit_name)}"
            if re.match(pattern, rtpc_name):
                flag = 1
                rtpc_id = create_obj_content(rtpc_unit_have_dict[rtpc_unit_name], "GameParameter",
                                             rtpc_name, "")
        if flag == 0:
            oi_h.print_error(rtpc_name + "：未找到相应的rtpc unit父级，请检查")

        # effect创建
        for effect_unit_name in effect_ducking_unit_have_dict:
            # print(effect_unit_name)
            pattern = f"^{re.escape(effect_unit_name)}"
            if re.match(pattern, rtpc_name):
                flag = 1
                # meter创建
                meter_id = create_obj_content(effect_ducking_unit_have_dict[effect_unit_name], "Effect",
                                              rtpc_name, "")

        if flag == 0:
            oi_h.print_error(rtpc_name + "：未找到相应的meter unit父级，请检查")

        return rtpc_id, meter_id


    """设置目标的rtpc"""


    #   'ControlInput': {'id': '{5D509B17-5562-4145-83B0-414A6107374F}',
    #                    'name': 'Aux_Ducking_Mus'},

    def set_target_rtpc(target_bus_id, rtpc_id, target_bus_name):
        # 先查找目标上的RTPC是否存在，若存在则不创建
        rtpc_list = client.call("ak.wwise.core.object.get",
                                waapi_h.waql_find_RTPC(target_bus_id),
                                options=config.options)['return']

        if rtpc_list:
            for rtpc_dict in rtpc_list:
                if rtpc_dict['ControlInput']['id'] == rtpc_id:
                    return
        args = waapi_h.args_rtpc_ducking_create(target_bus_id, rtpc_id)
        create_rtpc_list = client.call("ak.wwise.core.object.set", args)
        oi_h.print_warning(target_bus_name + "：ducking目标已设置相应rtpc")


    """创建Ducking所需要的内容"""


    def set_ducking():
        if duck_column:
            for cell in list(sheet.columns)[duck_column - 1]:
                if (cell.value) and (
                        not 命名规范检查.check_is_chinese(cell.value)):
                    if cell.value not in ducking_name_list:
                        if cell.value in bus_have_dict:
                            # 源查找
                            source_name, _ = excel_h.check_is_mergecell(sheet.cell(row=cell.row,
                                                                                   column=require_name_column), sheet)
                            if source_name:
                                rtpc_name = source_name + "_Ducking"
                                # print(source_name)
                                # 设置源的rtpc和meter
                                rtpc_id, meter_id = set_source_meter_and_rtpc(rtpc_name)
                                # 设置目标的rtpc
                                set_target_rtpc(bus_have_dict[cell.value], rtpc_id, cell.value)
                            else:
                                oi_h.print_error(cell.value + "：要ducking的bus无source bus的名称，请检查")




                        else:
                            oi_h.print_error(cell.value + "：要ducking的bus在wwise中未创建，请检查名称或先创建bus")
                    else:
                        oi_h.print_error(cell.value + "：有共同作用对象，需要发送aux bus")


    """获取需要创建的Bus列表"""


    def get_create_sub_bus_name_list():
        if require_name_column:
            for cell in list(sheet.columns)[require_name_column - 1]:
                if (cell.value) and (
                        not 命名规范检查.check_is_chinese(cell.value)):
                    if 命名规范检查.check_by_length_and_word_bool(
                            cell.value, 10):
                        # print(cell.value)
                        """资源描述"""
                        value_desc_value = sheet.cell(row=cell.row,
                                                      column=value_desc_column).value
                        """资源描述检查"""
                        if value_desc_value:
                            if value_desc_value not in value_desc_list:
                                value_desc_list.append(value_desc_value)

                                """名称查重及列表添加"""
                                if "Aux" in cell.value:
                                    if cell.value not in aux_name_list:
                                        # 对象创建
                                        create_wwise_obj_by_excel("AuxBus", "Bus", cell.value,
                                                                  config.wwise_bus_path, value_desc_value)


                                    else:
                                        oi_h.print_error(value_desc_value + "：表格中有重复项描述，请检查")

                                else:
                                    if cell.value not in bus_name_list:
                                        create_wwise_obj_by_excel("Bus", "Bus", cell.value,
                                                                  config.wwise_bus_path, value_desc_value)


                                    else:
                                        oi_h.print_error(value_desc_value + "：表格中有重复项描述，请检查")

                            else:
                                oi_h.print_error(cell.value + "：表格中有重复项描述，请检查")


    """获取表格中的事件描述和状态所在的列"""


    def get_descrip_and_status_column():
        duck_obj_column = None
        require_name_column = None
        status_column = None
        duck_column = None
        aux_column = None
        state_column = None
        break_column = None
        if list(sheet.rows)[0]:
            for cell in list(sheet.rows)[0]:
                if cell.value:
                    if '资源名称' == str(cell.value):
                        require_name_column = cell.column
                    elif '资源描述' == str(cell.value):
                        status_column = cell.column
                    elif 'ducking对象' == str(cell.value):
                        duck_column = cell.column
                    elif '发送至aux bus' == str(cell.value):
                        aux_column = cell.column
                    elif '添加StateGroup' == str(cell.value):
                        state_column = cell.column
                    elif '打断bus名称' == str(cell.value):
                        break_column = cell.column

        return require_name_column, status_column, duck_column, aux_column, state_column, break_column


    # """获取每行bus的配置，将bus名存为字典的键，将配置存为元组"""
    #
    #
    # def get_sub_bus_config():
    #     for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column,
    #                                values_only=True):
    #
    #         if row[0] and (
    #                 not 命名规范检查.check_is_chinese(row[0])) and 命名规范检查.check_by_length_and_word_bool(
    #             row[0], 10):
    #             key = row[0]
    #             # if key not in bus_config_dict:
    #             #     bus_config_dict[key] = row[1:]
    #
    #             # openpyxl遍历每行，将第1列值设为字典的键，2-8列设为字典的值，并将每行的字典添加到一个列表，
    #             # 行可能会有合并单元格，导致某个列有多个值，需要将多个值存在一个列表中，就是这个元组的元素是个列表
    #             # 初始化一个空列表，用于存储第2到第8列的值
    #             # 获取合并单元格范围
    #             merged_cells = sheet.merged_cells.ranges
    #             values = []
    #
    #             # 遍历第2到第8列
    #             for i in range(1, sheet.max_column):
    #                 cell_value = row[i]
    #                 cell = row[i]
    #                 cell_value = excel_h.check_is_mergecell(cell, sheet)
    #
    #                 # 如果不是合并单元格，直接将值添加到列表
    #                 if not isinstance(cell_value, list):
    #                     cell_value = [cell_value]
    #                 values.append(cell_value)
    #             if key not in bus_config_dict:
    #                 bus_config_dict[key] = values

    """*****************主程序处理******************"""

    # bus结构创建
    if wb["Bus结构创建"]:
        sheet = wb["Bus结构创建"]
        get_create_bus_name_list()
        # pprint(bus_create_list)
        # pprint(unit_create_list)
        # 列表排序
        bus_create_list.sort(key=len)
        unit_create_list.sort(key=len)
        create_wwise_bus(unit_create_list, bus_create_list)

    # 若自建Bus则置灰,颜色设为27
    # 查找已有bus列表
    bus_have_dict = get_wwise_type_list(config.wwise_bus_path, "Bus")
    for bus_have_name in bus_have_dict:
        if bus_have_name not in bus_create_list:
            if (bus_have_name not in config.bus_no_color_list) and (
                    not oi_h.check_by_re(r'^Aux', bus_have_name)):
                set_obj_property(bus_have_dict[bus_have_name], "OverrideColor", True)
                set_obj_property(bus_have_dict[bus_have_name], "Color", 0)

    # 用于记录要创建的子bus名称及属性
    bus_name_list = []
    # 用于记录要创建的子aux名称及属性
    aux_name_list = []
    # 用于记录需要ducking的bus
    ducking_name_list = []

    # 所有Bus的配置
    bus_config_dict = {}

    # 资源描述列表
    value_desc_list = []
    aux_have_dict = get_wwise_type_list(config.wwise_bus_path, "AuxBus")

    # rtpc unit字典
    rtpc_unit_have_dict = get_wwise_type_list(config.wwise_rtpc_path, "WorkUnit")
    # effect ducking unit字典
    effect_ducking_unit_have_dict = get_wwise_type_list(config.wwise_effect_ducking_path, "WorkUnit")
    # pprint(effect_ducking_unit_have_dict)
    # rtpc列表
    rtpc_have_dict = get_wwise_type_list(config.wwise_rtpc_path, "GameParameter")
    # 获取Meter
    meter_have_dict = get_wwise_type_list(config.wwise_effect_ducking_path, "Effect")

    # 子bus创建
    if wb["Bus创建"]:
        sheet = wb["Bus创建"]
        require_name_column, value_desc_column, duck_column, aux_column, state_column, break_column = get_descrip_and_status_column()
        # get_sub_bus_config()
        # 子bus创建
        get_create_sub_bus_name_list()
        # bus_name_list.sort(key=len)
        # aux_name_list.sort(key=len)
        # 子bus创建
        # create_wwise_sub_obj("AuxBus", "Bus", aux_name_list, config.wwise_bus_path)
        # create_wwise_sub_obj("Bus", "Bus", bus_name_list, config.wwise_bus_path)
        # pprint(bus_name_list)
        # pprint(aux_name_list)
        # pprint(bus_config_dict)
        # Ducking配置检索
        bus_have_dict = get_wwise_type_list(config.wwise_bus_path, "Bus")
        set_ducking()

    # """同步表中删除的内容"""
    # """对象删除"""
    #
    #
    # def delete_obj(obj_id, obj_name, obj_type):
    #     args = {
    #         "object": "%s" % obj_id
    #     }
    #     client.call("ak.wwise.core.object.delete", args)
    #     oi_h.print_warning(obj_name + "(" + obj_type + ")" + ":已删除")
    #
    #
    # """删除状态"""
    #
    #
    # def delete_check(wwise_obj_dict, excel_list, obj_type):
    #     for wwise_obj in wwise_obj_dict:
    #         # pprint(wwise_obj_dict['name'])
    #         if wwise_obj not in excel_list:
    #             # State/StateGroup删除
    #             delete_obj(wwise_obj_dict[wwise_obj], wwise_obj, obj_type)
    #             # StateGroupUnit删除
    #
    #
    # unit_have_dict = get_wwise_type_list(config.wwise_bus_path, "WorkUnit")
    # bus_have_dict = get_wwise_type_list(config.wwise_bus_path, "Bus")
    #
    # # 查找state/switch，再跟表格比对有没有，没有就删除资源及事件引用
    # delete_check(bus_have_dict, bus_create_list, "Bus")
    # delete_check(unit_have_dict, unit_create_list, "WorkUnit")
