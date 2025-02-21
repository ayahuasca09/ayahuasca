import openpyxl
import os
from pprint import pprint
import re
from waapi import WaapiClient
import shutil

import comlib.oi_h as oi_h
import comlib.config as config
import comlib.excel_h as excel_h

print("*****************媒体占位资源开始生成******************")

"""根目录获取"""
root_path = os.path.join(config.auto_sound_path, "媒体资源占位生成")
excel_path = os.path.join(root_path, "Excel")

file_name_list = excel_h.excel_get_path_list(excel_path)

"""*****************功能检测区******************"""
"""检测字符串是否含有中文"""


def check_is_chinese(string):
    for ch in string:
        if u'\u4e00' <= ch <= u'\u9fff':
            return True
    return False


"""随机容器后缀检查"""


def check_is_random(name):
    is_random = re.search(r"(_R\d{2,4})$", name)
    if is_random:
        return True
    else:
        return False


"""获取表格中的事件描述和状态所在的列"""


def get_descrip_and_status_column():
    require_module_column = None
    second_module_column = None
    require_name_column = None
    status_column = None
    sample_name_column = None

    if list(sheet.rows):
        if list(sheet.rows)[0]:
            for cell in list(sheet.rows)[0]:
                if cell.value:
                    if 'Require Module' in str(cell.value):
                        require_module_column = cell.column
                        # print(require_module_column)
                    elif 'Second Module' in str(cell.value):
                        second_module_column = cell.column
                        # print(second_module_column)
                    elif 'Require Name' in str(cell.value):
                        require_name_column = cell.column
                        # print(require_name_column)
                    elif 'Status' in str(cell.value):
                        status_column = cell.column
                        # print(status_column)
                    elif 'Sample Name' in str(cell.value):
                        sample_name_column = cell.column
                        # print(status_column)

    return require_module_column, second_module_column, require_name_column, status_column, sample_name_column


"""拼接事件描述"""


def get_event_descrip():
    event_descrip = ""
    if require_module_column:
        require_module_value, _ = excel_h.check_is_mergecell(
            sheet.cell(row=cell_sound.row, column=require_module_column), sheet)
        if require_module_value:
            event_descrip = require_module_value + "_"
    if second_module_column:
        second_module_value, _ = excel_h.check_is_mergecell(sheet.cell(row=cell_sound.row, column=second_module_column),
                                                            sheet)
        if event_descrip:
            if second_module_value:
                event_descrip = event_descrip + second_module_value + "_"
        else:
            if second_module_value:
                event_descrip = second_module_value + "_"

    if require_name_column:
        require_name_value, _ = excel_h.check_is_mergecell(sheet.cell(row=cell_sound.row, column=require_name_column),
                                                           sheet)
        if event_descrip:
            if require_name_value:
                event_descrip = event_descrip + require_name_value + "_"
        else:
            if require_name_value:
                event_descrip = require_name_value + "_"
    if not event_descrip:
        if cell_sound.value:
            oi_h.print_error(cell_sound.value + "：没有事件描述，请补充！")

    # 格式规范
    else:
        event_descrip = re.sub(r'^_', "", event_descrip)
        event_descrip = re.sub(r'_*$', "", event_descrip)
        event_descrip = re.sub(r'\s+', "", event_descrip)
    return event_descrip


with WaapiClient() as client:
    """*****************Wwise功能区******************"""
    """查找对象"""


    def find_obj(args):
        options = {
            'return': ['name', 'id', 'path', 'notes', 'originalWavFilePath', 'type']

        }
        obj_sub_list = client.call("ak.wwise.core.object.get", args, options=options)['return']
        if not obj_sub_list:
            obj_sub_id = ""
            obj_sub_path = ""
        else:
            obj_sub_id = obj_sub_list[0]['id']
            obj_sub_path = obj_sub_list[0]['path']
        return obj_sub_list, obj_sub_id, obj_sub_path


    """获取wwise对象的最长字符串父级"""


    def find_obj_parent(obj_name, waql):
        # 所有Unit获取
        parent_list, parent_id, _ = find_obj(waql)

        # 存储符合的对象的父级，最后找出最长最符合的
        parent_len = 0
        obj_parent_id = {}
        if parent_list != None:
            for parent_dict in parent_list:
                if parent_dict['name'] in obj_name:
                    if len(parent_dict['name']) > parent_len:
                        parent_len = len(parent_dict['name'])
                        obj_parent_id = parent_dict['id']
        else:
            oi_h.print_error(cell_sound.value + "相应的Unit目录结构未创建！需要创建")

        return obj_parent_id


    """媒体资源创建及导入"""


    def import_media(media_name, rnd_path):
        # 媒体资源复制
        # 复制到的目录
        copy_catalog = os.path.join(root_path, "New_Media")
        source_path = shutil.copy2(os.path.join(root_path, "Media_Temp.wav"),
                                   os.path.join(copy_catalog, media_name + ".wav"))

        # 在容器中创建媒体资源
        import_media_in_rnd(source_path, media_name, rnd_path)


    """修改Wwise内容的名字"""


    def change_name_by_wwise_content(obj_id, name, old_name, obj_type):
        args = {
            "objects": [
                {

                    "object": obj_id,
                    "name": name,
                }
            ]
        }
        client.call("ak.wwise.core.object.set", args)
        oi_h.print_warning(old_name + "(" + obj_type + ")改名为：" + name)


    """设置obj的notes"""


    def set_obj_notes(obj_id, notes_value):
        args = {
            'object': obj_id,
            'value': notes_value
        }
        client.call("ak.wwise.core.object.setNotes", args)
        # oi_h.print_warning(obj_name + "描述更改为：" + notes_value)


    """导入媒体资源"""


    def import_sound_sfx_and_media(media_name, rnd_path):
        # 媒体资源导入
        import_media(media_name, rnd_path)
        # Sound颜色设置
        _, sound_container_id, _ = find_obj(
            {'waql': 'from type Sound where name = "%s"' % media_name})
        if sound_container_id:
            set_sound_color_default(sound_container_id)

        oi_h.print_warning(media_name + "：Sound创建及Media导入")


    """创建新的随机容器"""


    def create_rnd_container(media_name, sys_name):
        # 查找该rnd是否已存在
        flag = 0
        # 去除随机容器数字
        rnd_name = re.sub(r"(_R\d{2,4})$", "", media_name)

        rnd_path = ""
        rnd_container_list, rnd_id, _ = find_obj(
            {'waql': ' "%s" select descendants where type = "RandomSequenceContainer" ' % os.path.join(
                config.wwise_sfx_path, sys_name)})
        # 找到重名的
        for rnd_container_dict in rnd_container_list:
            if rnd_container_dict['name'] == rnd_name:
                flag = 1
                rnd_path = rnd_container_dict['path']
                set_obj_notes(rnd_container_dict['id'], event_descrip)
                break
            # 没有重名但描述一致，判定为改名
            elif rnd_container_dict['notes'] == event_descrip:
                # pprint(rnd_name + "：RandomContainer已存在，将不再导入")
                flag = 2
                rnd_path = rnd_container_dict['path'].replace(rnd_container_dict['name'], rnd_name)
                change_name_by_wwise_content(rnd_container_dict['id'], rnd_name, rnd_container_dict['name'],
                                             "RandContainer")
                # Sound也要改名
                sound_container_list, sound_container_id, _ = find_obj(
                    {'waql': ' "%s" select children  ' % rnd_container_dict['id']})
                for sound_cotainer_dict in sound_container_list:
                    wav_tail = re.search(r"(_R\d{2,4})$", sound_cotainer_dict['name'])
                    sound_name = rnd_name
                    if wav_tail:
                        sound_name = rnd_name + wav_tail.group()
                    change_name_by_wwise_content(sound_cotainer_dict['id'], sound_name, sound_cotainer_dict['name'],
                                                 "Sound")
                    # Media路径更改
                    new_media_path = sound_cotainer_dict['originalWavFilePath'].replace(sound_cotainer_dict['name'],
                                                                                        sound_name)
                    # Media重命名
                    os.rename(sound_cotainer_dict['originalWavFilePath'],
                              new_media_path)
                    # 在容器中重新导入media
                    import_media_in_rnd(new_media_path, sound_name, rnd_path)
                    oi_h.print_warning(sound_cotainer_dict['originalWavFilePath'] + "(Media)改名为：" + new_media_path)

                    # 语音的话需要删除容器内的引用
                    if sys_name == "VO":
                        sound_language_list, _, _ = find_obj(
                            {'waql': ' "%s" select children  ' %
                                     sound_cotainer_dict['id']})
                        if sound_language_list:
                            for sound_language_dict in sound_language_list:
                                if sound_cotainer_dict['name'] in sound_language_dict['name']:
                                    args = {
                                        "object": "%s" % sound_language_dict['id']
                                    }
                                    client.call("ak.wwise.core.object.delete", args)
                                    oi_h.print_warning(
                                        sound_language_dict['name'] + "：容器引用删除")

                break

        # 不存在则创建
        if flag == 0:
            # 查找所在路径
            rnd_parent_id = find_obj_parent(rnd_name,
                                            {
                                                'waql': ' "%s" select descendants,this where type = "ActorMixer" ' %
                                                        os.path.join(
                                                            config.wwise_sfx_path, sys_name)})
            if rnd_parent_id:

                # 创建的rnd属性
                args = {
                    # 选择父级
                    "parent": rnd_parent_id,
                    # 创建类型机名称
                    "type": "RandomSequenceContainer",
                    "name": rnd_name,
                    "notes": event_descrip,
                    "@RandomOrSequence": 1,
                    "@NormalOrShuffle": 1,
                    "@RandomAvoidRepeatingCount": 1
                }
                rnd_container_object = client.call("ak.wwise.core.object.create", args)
                # 查找新创建的容器的路径
                _, _, rnd_path = find_obj(
                    {'waql': ' "%s"  ' % rnd_container_object['id']})
                if rnd_path:
                    oi_h.print_warning(rnd_name + "：RandContainer创建")
                else:
                    oi_h.print_error(rnd_name + "：RandContainer未能成功创建")

                # 声音容器创建及导入
                import_sound_sfx_and_media(media_name, rnd_path)

        return rnd_path, rnd_name


    """设置新的占位Sound资源超越父级为灰色"""


    def set_sound_color_default(obj_id):
        args_property = {
            "object": obj_id,
            "property": "OverrideColor",
            "value": True,
        }
        client.call("ak.wwise.core.object.setProperty", args_property)

        args_property = {
            "object": obj_id,
            "property": "Color",
            "value": 0,
        }
        client.call("ak.wwise.core.object.setProperty", args_property)


    """VO：复制Chinese下的文件到其他文件夹里"""


    def copy_reference_VO_to_other(media_name):
        wwise_vo_media_path_cn = config.wwise_vo_media_path + "\\Chinese" + "\\" + media_name + ".wav"
        wwise_vo_media_path_en = config.wwise_vo_media_path + "\\English" + "\\" + media_name + ".wav"
        wwise_vo_media_path_jp = config.wwise_vo_media_path + "\\Japanese" + "\\" + media_name + ".wav"
        wwise_vo_media_path_kr = config.wwise_vo_media_path + "\\Korean" + "\\" + media_name + ".wav"
        if os.path.isfile(wwise_vo_media_path_cn):
            # print(wwise_vo_media_path_cn)
            if not os.path.isfile(wwise_vo_media_path_en):
                source_path = shutil.copy2(wwise_vo_media_path_cn,
                                           wwise_vo_media_path_en)
                # print(wwise_vo_media_path_en)
            if not os.path.isfile(wwise_vo_media_path_jp):
                source_path = shutil.copy2(wwise_vo_media_path_cn,
                                           wwise_vo_media_path_jp)
                # print(wwise_vo_media_path_jp)
            if not os.path.isfile(wwise_vo_media_path_kr):
                source_path = shutil.copy2(wwise_vo_media_path_cn,
                                           wwise_vo_media_path_kr)
                # print(wwise_vo_media_path_kr)


    """媒体资源导入"""


    def import_media_in_rnd(source_path, media_name, rnd_path):
        # 导入语音
        if sys_name == "VO":
            args_import = {
                "importOperation": "useExisting",
                "imports": [
                    {
                        "audioFile": source_path,
                        "objectPath": rnd_path + "\\<Sound Voice>" + media_name + "\\<AudioFileSource>" + "CN_" + media_name,
                        "importLanguage": "Chinese"
                    },
                    {
                        "audioFile": source_path,
                        "objectPath": rnd_path + "\\<Sound Voice>" + media_name + "\\<AudioFileSource>" + "EN_" + media_name,
                        "importLanguage": "English"
                    },
                    {
                        "audioFile": source_path,
                        "objectPath": rnd_path + "\\<Sound Voice>" + media_name + "\\<AudioFileSource>" + "JP_" + media_name,
                        "importLanguage": "Japanese"
                    },
                    {
                        "audioFile": source_path,
                        "objectPath": rnd_path + "\\<Sound Voice>" + media_name + "\\<AudioFileSource>" + "KR_" + media_name,
                        "importLanguage": "Korean"
                    }
                ]
            }
        # 导入音效
        else:
            args_import = {
                # createNew
                # useExisting：会增加一个新媒体文件但旧的不会删除
                # replaceExisting:会销毁Sound，上面做的设置都无了
                "importOperation": "replaceExisting",
                "default": {
                    "importLanguage": "SFX"
                },
                "imports": [
                    {
                        "audioFile": source_path,
                        "objectPath": rnd_path + '\\<Sound SFX>' + media_name,
                        "originalsSubFolder": sys_name
                        #                                                         名为Test 0的顺序容器            名为My SFX 0 的音效
                        # "objectPath": "\\Actor-Mixer Hierarchy\\Default Work Unit\\<Sequence Container>Test 0\\<Sound SFX>My SFX 0"
                    }
                ]
            }

        # 定义返回结果参数，让其只返回 Windows 平台下的信息，信息中包含 GUID 和新创建的对象名
        opts = {
            "platform": "Windows",
            "return": [
                "id", "name"
            ]
        }
        client.call("ak.wwise.core.audio.import", args_import, options=opts)

        if sys_name == "VO":
            copy_reference_VO_to_other(media_name)


    """创建事件的参数"""


    def get_event_args(event_parent_path, event_name, event_action, event_target, notes):
        args_new_event = {
            # 上半部分属性中分别为 Event 创建后存放的路径、类型、名称、遇到名字冲突时的处理方法
            "parent": event_parent_path,
            "type": "Event",
            "name": event_name,
            "onNameConflict": "merge",
            "children": [
                {
                    # 为其创建播放行为，名字留空，使用 @ActionType 定义其播放行为为 Play，@Target 为被播放的声音对象
                    "name": "",
                    "type": "Action",
                    "@ActionType": event_action,
                    "@Target": event_target
                }
            ],
            "notes": notes
        }
        return args_new_event


    """事件生成"""


    def create_event(rnd_name, rnd_path):
        parent_id = find_obj_parent(rnd_name, {
            'waql': ' "%s" select descendants where type = "WorkUnit" ' % os.path.join(config.wwise_event_path,
                                                                                       sys_name)})
        event_name = "AKE_" + "Play_" + rnd_name
        event_stop_name = "AKE_" + "Stop_" + rnd_name
        # 查找事件是否已存在
        event_list, _, _ = find_obj(
            {'waql': ' "%s" select descendants where type = "Event" ' %
                     config.wwise_event_path})

        # 判断其random容器的父级是不是actor-mixer，如果不是就不生成event
        rnd_parent_list, _, _ = find_obj(
            {'waql': ' "%s" select parent ' %
                     rnd_path})
        if rnd_parent_list:
            if rnd_parent_list[0]['type'] == "ActorMixer":
                # pprint(event_list)
                flag = 0
                for event_dict in event_list:
                    if rnd_name in event_dict['name']:
                        # Play事件的notes重新设置
                        if event_dict['name'] == event_name:
                            flag = 1
                            set_obj_notes(event_dict['id'], event_descrip)
                            # print(event_dict['name'])
                        elif event_dict['name'] == event_stop_name:
                            flag = 2
                            set_obj_notes(event_dict['id'], event_descrip + "停止")
                            # print(event_dict['name'])
                    else:
                        # Play:没有重名但描述一致，判定为改名
                        if event_dict['notes'] == event_descrip:
                            # pprint(rnd_name + "：RandomContainer已存在，将不再导入")
                            flag = 3
                            change_name_by_wwise_content(event_dict['id'], event_name, event_dict['name'], "Event")
                        # Stop：没有重名但描述一致，判定为改名
                        elif event_dict['notes'] == event_descrip + "停止":
                            # pprint(rnd_name + "：RandomContainer已存在，将不再导入")
                            flag = 4
                            change_name_by_wwise_content(event_dict['id'], event_stop_name, event_dict['name'], "Event")

                if flag == 0:
                    event_args = get_event_args(parent_id, event_name, 1, rnd_path, event_descrip)
                    client.call("ak.wwise.core.object.create", event_args)
                    # {'children': [{'id': '{26A1FB06-3908-4F55-AB8A-E67264100F18}', 'name': ''}],
                    #  'id': '{F04A7B4F-9266-4DAB-8DD2-CAF6BFD6D78A}',
                    #  'name': 'AKE_Play_Mon_Boss_Body_B01_Atk5'}
                    oi_h.print_warning(event_name + "事件创建")
                    if "_LP" in rnd_name:
                        event_name = "AKE_" + "Stop_" + rnd_name
                        event_args = get_event_args(parent_id, event_name, 2, rnd_path, event_descrip + "停止")
                        client.call("ak.wwise.core.object.create", event_args)
                        pprint(event_name + "事件创建")


    def create_wwise_content(media_name, sys_name):

        # 随机容器创建
        rnd_path, rnd_name = create_rnd_container(media_name, sys_name)

        # 事件自动生成
        create_event(rnd_name, rnd_path)


    def delete_wwise_content():
        # 随机容器引用的资产删除
        media_list, media_id, _ = find_obj(
            {'waql': ' "%s" select children where type = "Sound" ' % rnd_container['id']})
        if media_list:
            for media_dict in media_list:
                if os.path.isfile(media_dict['originalWavFilePath']):
                    os.remove(media_dict['originalWavFilePath'])
                    oi_h.print_warning(media_dict['originalWavFilePath'] + "已删除")

        # 随机容器删除
        args = {
            "object": "%s" % rnd_container['id']
        }
        client.call("ak.wwise.core.object.delete", args)
        oi_h.print_warning(rnd_container['name'] + ":RandomContainer已删除")

        # 事件删除
        for event_dict in event_list:
            if rnd_container['name'] in event_dict['name']:
                args = {
                    "object": "%s" % event_dict['id']
                }
                client.call("ak.wwise.core.object.delete", args)
                oi_h.print_warning(event_dict['name'] + ":Event已删除")


    def delete_or_modify_wwise_content():
        for i in file_name_list:
            if ".xlsx" in i:
                # 拼接xlsx的路径
                file_path_xlsx = os.path.join(excel_path, i)
                # 获取xlsx的workbook
                wb = openpyxl.load_workbook(file_path_xlsx)
                # 获取xlsx的所有sheet
                sheet_names = wb.sheetnames
                # 加载所有工作表
                for sheet_name in sheet_names:
                    sheet = wb[sheet_name]
                    # 获取工作表第一行数据
                    for cell in list(sheet.rows)[0]:
                        if 'Sample Name' in str(cell.value):
                            # 获取音效名下的内容
                            for cell_sound in list(sheet.columns)[cell.column - 1]:
                                # 空格和中文不检测
                                if cell_sound.value != None:
                                    if check_is_chinese(cell_sound.value) == False:

                                        """事件描述"""
                                        event_descrip = get_event_descrip()

                                        # 如果音效名与rnd名相同
                                        if rnd_container['name'] == cell_sound.value:
                                            # 如果状态标为取消的删除
                                            for status in list(sheet.rows)[cell_sound.row - 1]:
                                                if "cancel" == status.value:
                                                    delete_wwise_content()
                                            return
        # 其他为在表格中搜索不到的名字
        delete_wwise_content()

        # 删除语音不对应的rnd sound资源


    def delete_wwise_content_vo_sound():
        # Sound容器删除
        args = {
            "object": "%s" % sound_container['id']
        }
        client.call("ak.wwise.core.object.delete", args)
        oi_h.print_warning(rnd_container['name'] + ":SoundContainer已删除")

        # Sound容器引用的资产删除
        if os.path.isfile(sound_container['originalWavFilePath']):
            os.remove(sound_container['originalWavFilePath'])
            oi_h.print_warning(sound_container['originalWavFilePath'] + "已删除")

        # 删除语音不对应的rnd sound资源


    def delete_or_modify_wwise_content_vo_sound():
        for i in file_name_list:
            if ".xlsx" in i:
                # 拼接xlsx的路径
                file_path_xlsx = os.path.join(excel_path, i)
                # 获取xlsx的workbook
                wb = openpyxl.load_workbook(file_path_xlsx)
                # 获取xlsx的所有sheet
                sheet_names = wb.sheetnames
                # 加载所有工作表
                for sheet_name in sheet_names:
                    sheet = wb[sheet_name]
                    # 获取工作表第一行数据
                    for cell in list(sheet.rows)[0]:
                        if 'Sample Name' in str(cell.value):
                            # 获取音效名下的内容
                            for cell_sound in list(sheet.columns)[cell.column - 1]:
                                # 空格和中文不检测
                                if cell_sound.value != None:
                                    if check_is_chinese(cell_sound.value) == False:

                                        """事件描述"""
                                        event_descrip = get_event_descrip()

                                        # 如果音效名与rnd名相同
                                        if sound_container['name'] == cell_sound.value:
                                            # 如果状态标为取消的删除
                                            for status in list(sheet.rows)[cell_sound.row - 1]:
                                                if "cancel" == status.value:
                                                    delete_wwise_content_vo_sound()
                                            return
        # 其他为在表格中搜索不到的名字
        delete_wwise_content_vo_sound()


    """*****************主程序处理******************"""
    # 撤销开始
    client.call("ak.wwise.core.undo.beginGroup")
    # 记录所有资源名称
    sound_name_list = []

    # 事件描述字典：用于记录描述键和行作为值
    event_descrip_dict = {}

    # 提取规则：只提取xlsx文件
    for i in file_name_list:
        if ".xlsx" in i:
            # 拼接xlsx的路径
            file_path_xlsx = os.path.join(excel_path, i)
            # 获取xlsx的workbook
            wb = openpyxl.load_workbook(file_path_xlsx)
            # 获取xlsx的所有sheet
            sheet_names = wb.sheetnames
            # 加载所有工作表
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                word_list_len = sheet.max_column

                require_module_column, second_module_column, require_name_column, status_column, sample_name_column = get_descrip_and_status_column()
                if sample_name_column:
                    for cell_sound in list(sheet.columns)[sample_name_column - 1]:
                        if status_column:
                            if sheet.cell(row=cell_sound.row,
                                          column=status_column).value in config.status_list:
                                if (cell_sound.value) and (
                                        not check_is_chinese(cell_sound.value)):
                                    # pprint(cell_sound.value)
                                    """检查表格中是否有内容重复项"""
                                    if cell_sound.value in sound_name_list:
                                        oi_h.print_error(cell_sound.value + "：表格中有重复项音效名，请检查")
                                    else:
                                        sound_name_list.append(cell_sound.value)
                                        """测试名称"""
                                        name = cell_sound.value
                                        """事件描述"""
                                        event_descrip = get_event_descrip()
                                        if (event_descrip in event_descrip_dict) and (
                                                not check_is_random(name)):
                                            oi_h.print_error(event_descrip + "：表格中有重复项描述，请检查")

                                        else:
                                            event_descrip_dict[event_descrip] = cell_sound.value
                                            name_list = name.split("_")
                                            sys_name = name_list[0]
                                            # print(sys_name)
                                            create_wwise_content(cell_sound.value, sys_name)

    """同步表中删除的内容"""
    # 查找所有Rnd和Event
    rnd_container_list, rnd_id, _ = find_obj(
        {'waql': ' "%s" select descendants where type = "RandomSequenceContainer" ' % config.wwise_sfx_path})
    # [{'id': '{C1BFDDC1-CA6F-45BC-8B43-15AE866AA20A}',
    #   'name': 'Char_Skill_C01_Atk1',
    #   'notes': '',
    #   'path': '\\Actor-Mixer '
    #           'Hierarchy\\v1\\Char\\Char\\Char_Skill\\Char_Skill\\Char_Skill_C01\\Char_Skill_C01\\Char_Skill_C01_Atk1'}
    event_list, event_id, _ = find_obj(
        {'waql': ' "%s" select descendants where type = "Event" ' % config.wwise_event_path})

    # 查找Amb_Global的内容，不删除
    obj_sub_list, _, _ = find_obj(
        {
            'waql': ' "{8B09C109-84EA-448C-BDF9-C7E371E6375C}" select descendants where type = "RandomSequenceContainer" '})
    extract_key = lambda d: d['id']
    amb_2d_list = list(map(extract_key, obj_sub_list))
    # pprint(amb_2d_list)

    for rnd_container in rnd_container_list:
        if rnd_container['id'] not in amb_2d_list:
            # 先禁用删除
            # delete_or_modify_wwise_content()
            pass

    # 对于语音，需要查找Sound容器里的随机资源
    # 查找所有Sound
    sound_container_list, _, _ = find_obj(
        {'waql': ' "%s" select descendants where type = "Sound" ' % config.wwise_vo_game_path})

    for sound_container in sound_container_list:
        # 先禁用删除
        # delete_or_modify_wwise_content_vo_sound()
        pass

    # 撤销结束
    client.call("ak.wwise.core.undo.endGroup", displayName="rnd创建撤销")

    oi_h.delete_type_files(os.path.join(root_path, "New_Media"), '.wav')

print("*****************媒体占位资源生成完毕******************")
os.system("pause")
