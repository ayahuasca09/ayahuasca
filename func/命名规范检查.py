import module.cloudfeishu.cloudfeishu_h as cloudfeishu_h
import config
import os
import module.excel.excel_h as excel_h
from pprint import pprint
import re
from openpyxl.styles import PatternFill

is_pass = True

"""根目录获取"""
# 获取当前脚本的文件名及路径
script_name_with_extension = __file__.split('/')[-1]
# 去掉扩展名
script_name = script_name_with_extension.rsplit('.', 1)[0]
# 替换为files
root_path = script_name.replace("func", "files")
# print(f"当前脚本的名字是: {root_path}")

"""文件获取"""
# 规范检查表
excel_check_path = os.path.join(root_path, config.excel_check_name)
# print(wb)

"""在线表获取"""
# 规范检查表
check_name_token = config.check_name_token
cloudfeishu_h.download_cloud_sheet(check_name_token, excel_check_path)
sheet_config, wb = excel_h.excel_get_sheet(excel_check_path, "配置说明")

"""初始化配置说明"""
# 初始化空列表
first_column_data = []

# 遍历工作表的第一列,索引为表格行-1
for row in sheet_config.iter_rows(min_row=1, max_col=1, max_row=sheet_config.max_row):
    for cell in row:
        first_column_data.append(cell.value)
        # 字体颜色测验
        # font_color = cell.font.color
        # if cell.row == 3:
        #     print(cell.value)
        #     if font_color is not None:
        #         # font_color 是一个 Color 对象
        #         rgb = font_color.rgb
        #         if rgb is not None:
        #             print(f"The font color of cell A1 is: {rgb}")
        #         else:
        #             print("The font color of cell A1 is a theme color or indexed color.")
        #     else:
        #         print("The font color of cell A1 is not set.")

        # 填充色测验
        # if cell.row == 4:
        #     # 获取单元格的背景填充色
        #     fill = cell.fill
        #     if fill.fill_type == 'solid':
        #         bg_color = fill.start_color.index
        #         print(f'Cell {cell.coordinate} has background color: {bg_color}')
        #     else:
        #         print(f'Cell {cell.coordinate} has no solid fill')

"""********************功能函数***********************"""
"""检测正则表达式的pattern是否正确"""


def is_valid_regex(pattern):
    try:
        re.compile(pattern)
        return True, None
    except re.error as e:
        print(f"正则表达式不合规。错误信息: {str(e)}")
        return False


"""分隔符的大小写和长度检查"""


def check_by_length_and_word(name, word_list_len):
    # pprint(cell_sound.value)
    word_list = name.split("_")
    if len(word_list) <= word_list_len:
        is_title = True
        for word in word_list:
            """判定_的每个都不能超过10个"""
            check_by_str_length(word, config.one_word_len, name)
            """判定_的每个开头必须大写"""
            if len(word) > 0:
                # 只需要检查第一个字符，因为istitle会导致检查字符串只能首字母大写
                if word[0].istitle() == False and word[0].isdigit() == False:
                    is_title = False
                    break
        if is_title == False:
            print_error(name + "：通过”_“分隔的每个单词开头都需要大写")
    else:
        print_error(name + "：通过”_“分隔的单词个数不能超过" + str(config.word_list_len))

    if word_list:
        return word_list[0]
    else:
        return None


"""分隔符的大小写和长度检查返回bool"""


def check_by_length_and_word_bool(name, word_list_len):
    word_list = name.split("_")
    if len(word_list) <= word_list_len:
        for word in word_list:
            """判定_的每个都不能超过10个"""
            if not check_by_str_length(word, config.one_word_len, name):
                return False
            """判定_的每个开头必须大写"""
            if len(word) > 0:
                # 只需要检查第一个字符，因为istitle会导致检查字符串只能首字母大写
                if word[0].istitle() == False and word[0].isdigit() == False:
                    is_title = False
                    print_error(name + "：通过”_“分隔的每个单词开头都需要大写")
                    return False
    else:
        print_error(name + "：通过”_“分隔的单词个数不能超过" + str(config.word_list_len))
        return False
    return True


"""字符串长度检查"""


def check_by_str_length(str1, length, name):
    if len(str1) > length:
        if "_" in str1:
            print_error(name + "：描述后缀名过长，限制字符数为" + str(length))
        else:
            print_error(name + "：通过_切分的某个单词过长，每个单词长度不应超过10个字符，需要再拆分")
        return False
    return True


"""基础规范检查"""


def check_basic(media_name, audio_unit_list, event_unit_list, audio_mixer_list, word_list_len):
    global is_pass
    # 检查是否为空
    if media_name:
        # 检查是否中文
        if not check_is_chinese(media_name):
            check_by_length_and_word(media_name, word_list_len)
            check_by_wb(media_name, audio_unit_list, event_unit_list, audio_mixer_list)
        else:
            is_pass = False
    else:
        is_pass = False

    return is_pass


"""正则匹配"""


def check_by_re(pattern, name, media_name):
    if pattern:
        # pprint(media_name)
        # pprint(name + ":" + pattern)
        global is_pass
        if pattern == "GenTable":
            is_pass = True
            # print(media_name)
            # print(check_by_gentable(name, media_name))

            return check_by_gentable(name, media_name)

        is_pass = False
        # 将双反斜杠替换为单反斜杠
        pattern = pattern.replace('\\\\', '\\')
        if is_valid_regex(pattern):
            result = re.search(pattern, name)
            if result:
                is_pass = True
                return is_pass
    # else:
    #     pprint(media_name + "：" + name + "的正则表达式" + pattern + "为空")


"""log捕获"""


def print_warning(warn_info):
    print("[warning]" + warn_info)


"""报错捕获"""


def print_error(error_info):
    global is_pass
    is_pass = False
    print("[error]" + error_info)
    return False


"""检测字符串是否含有中文"""


def check_is_chinese(string):
    for ch in string:
        if u'\u4e00' <= ch <= u'\u9fff':
            return True
    return False


"""存储需要创建的unit结构"""


def get_audio_unit_list():
    # 用于存储unit值
    audio_unit_list = []

    sheet_names = wb.sheetnames
    # 加载所有工作表
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        """获取工作表中的系统名"""
        # 获取工作表的颜色
        if sheet.sheet_properties.tabColor:
            # upper：字符串转大写
            if sheet.sheet_properties.tabColor.rgb.upper() == first_column_data[1]:
                # print(sheet_name)
                audio_unit_list.append(sheet_name)
                """获取工作列中的系统名"""
                # 从第二行开始遍历每行
                for row in sheet.iter_rows(min_row=2):
                    row_unit_name = sheet_name
                    # 遍历每列的单元格
                    for cell in row:
                        if cell.column == 1:
                            # 检查若为中文则跳过
                            if check_is_chinese(cell.value):
                                break
                        # print(cell.value)
                        # 获取单元格的背景填充色
                        fill = cell.fill
                        if fill.fill_type == 'solid':
                            if fill.start_color.index == first_column_data[1]:
                                row_unit_name += "_" + cell.value
                                audio_unit_list.append(row_unit_name)
                                # print(row_unit_name)

    return audio_unit_list


# pprint(get_audio_unit_list())

"""Audio Mixer检测"""


def check_audio_mixer(begin_name, end_name, cell, audio_mixer_list):
    if end_name:
        # 表格内容获取
        if cell:
            fill = cell.fill
            if fill.fill_type == 'solid':
                begin_name += "_" + end_name
                if fill.start_color.index == first_column_data[1]:
                    if begin_name not in audio_mixer_list:
                        audio_mixer_list.append(begin_name)

                elif fill.start_color.index == first_column_data[3]:
                    if begin_name not in audio_mixer_list:
                        audio_mixer_list.append(begin_name)

    return begin_name


"""Audio Unit检测"""


# begin_name：资源表的前一个名字
# sheet：规范检查表的当前sheet
# end_name：资源表的后一个名字
# cell：规范检查表单元格是否有标注，例如填色或颜色
# audio_unit_list：资源表中用来收集unit的名字

def check_audio_unit(begin_name, end_name, cell, audio_unit_list):
    if end_name and begin_name:
        # 表格内容获取
        if cell:
            fill = cell.fill
            if fill.fill_type == 'solid':
                begin_name += "_" + end_name
                if fill.start_color.index == first_column_data[1]:
                    # print(unit_name)
                    if begin_name not in audio_unit_list:
                        audio_unit_list.append(begin_name)

    return begin_name


"""Event Unit检测"""


def check_event_unit(begin_name, end_name, cell, event_unit_list, sheet):
    if begin_name and end_name:
        if cell is not None and cell.font is not None and cell.font.color is not None:
            # 表格内容获取
            font_color = cell.font.color.rgb
            if font_color == first_column_data[2]:
                # 它右边的单元格如果也会灰色， 则不加
                flag = 0
                # cell_right = sheet.cell(row=cell.row, column=cell.column + 1)
                # if cell_right is not None and cell_right.font is not None and cell_right.font.color is not None:
                #     right_font_color = cell_right.font.color.rgb
                #     if right_font_color == first_column_data[2]:
                #         flag = 1
                # 它左边的单元格如果也会灰色， 则不加
                if cell.column > 1:
                    cell_left = sheet.cell(row=cell.row, column=cell.column - 1)
                    if cell_left is not None and cell_left.font is not None and cell_left.font.color is not None:
                        left_font_color = cell_left.font.color.rgb
                        if left_font_color == first_column_data[2]:
                            flag = 1

                if flag == 0:
                    begin_name += "_" + end_name

                if begin_name not in event_unit_list:
                    event_unit_list.append(begin_name)

    return begin_name


"""通用词汇表检查"""


def check_by_gentable(media_name_check_word, media_name):
    global is_pass
    sheet = wb['GenTable']
    return_word = ""
    # 获取最大行和最大列
    max_row = sheet.max_row
    max_column = sheet.max_column

    # 检查第一列中有没有，有则直接通过
    for cell in sheet['A']:
        if media_name_check_word == cell.value:
            return True

    # 查找除第一列之外还有没有别的列含有
    flag = 0

    # 从第二行第二列开始遍历
    for row in range(2, max_row + 1):
        for col in range(2, max_column + 1):
            # 获取单元格的值
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value:
                if not check_is_chinese(cell_value):
                    if cell_value == media_name_check_word:
                        flag = 1
                        if sheet.cell(row=row, column=1).value == ".*":
                            return True
                        else:
                            print_error(media_name + "：" + media_name_check_word + "需要替换为" + sheet.cell(row=row,
                                                                                                             column=1).value + "(" + sheet.cell(
                                row=row - 1,
                                column=1).value + ")")
                            # print(cell_value)
                            break
    if flag == 0:
        return True
        # print_error(media_name_check_word + "：需要添加到通用词汇表中")


"""命名规范表检查"""


def check_by_wb(media_name, audio_unit_list, event_unit_list, audio_mixer_list):
    global is_pass
    sheet_names = wb.sheetnames

    # 解析media_name的数据
    media_name_list = media_name.split('_')
    # print(media_name_list)

    sys_name_flag = 0
    begin_name = ""
    # 加载所有工作表
    for sheet_name in sheet_names:
        """media_name_list[0]"""
        if media_name_list[0] == sheet_name:
            sys_name_flag = 1

            sheet = wb[sheet_name]
            if sheet.sheet_properties.tabColor:
                # upper：字符串转大写
                if sheet.sheet_properties.tabColor.rgb.upper() == first_column_data[1]:
                    if media_name_list[0] not in audio_unit_list:
                        audio_unit_list.append(media_name_list[0])
                    if media_name_list[0] not in event_unit_list:
                        event_unit_list.append(media_name_list[0])
                    if media_name_list[0] not in audio_mixer_list:
                        audio_mixer_list.append(media_name_list[0])

            # 获取列数
            max_column = sheet.max_column
            # print(max_column)
            """第一列索引获取"""
            # 遍历列
            check_cell = None
            check_cell_col = 0
            is_merge = False
            merge_min_row = 0
            merge_max_row = 0
            event_begin_name = ""
            audio_begin_name = ""
            mixer_begin_name = ""
            for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
                for cell in row:
                    if cell.value:
                        if check_is_chinese(cell.value):
                            break
                        elif check_by_re(cell.value, media_name_list[1], media_name):
                            _, is_merge, merge_min_row, merge_max_row = excel_h.check_mergecell_row(cell, sheet)
                            # print(cell.value + str(merge_min_row) + str(merge_max_row))
                            check_cell = cell
                            check_cell_col = cell.column
                            break

            if check_cell:
                event_begin_name = check_event_unit(media_name_list[0], media_name_list[1], check_cell,
                                                    event_unit_list, sheet)
                audio_begin_name = check_audio_unit(media_name_list[0], media_name_list[1], check_cell,
                                                    audio_unit_list)
                mixer_begin_name = check_audio_mixer(media_name_list[0], media_name_list[1], check_cell,
                                                     audio_mixer_list)

                # pprint(media_name)
                # pprint(check_cell.value)
                """其他列数据检查"""
                # 遍历选中规则的行
                # 从第2列开始遍历该行的每一列
                # print(media_name)
                # 检查是否为合并单元格
                if is_merge:
                    check_cell = None
                    # print(check_cell.value)
                    for row in range(merge_min_row, merge_max_row + 1):
                        cell_value = sheet.cell(row=row, column=check_cell_col + 1).value
                        # print(cell_value)
                        if check_by_re(cell_value, media_name_list[2], media_name):
                            check_cell = sheet.cell(row=row, column=check_cell_col + 1)
                            # print(cell_value)
                            break

                    if check_cell:
                        event_begin_name = check_event_unit(event_begin_name, media_name_list[2], check_cell,
                                                            event_unit_list, sheet)
                        audio_begin_name = check_audio_unit(audio_begin_name, media_name_list[2], check_cell,
                                                            audio_unit_list)
                        mixer_begin_name = check_audio_mixer(mixer_begin_name, media_name_list[2], check_cell,
                                                             audio_mixer_list)
                        for cell in sheet[check_cell.row]:
                            if cell.column != 1 and cell.value and (cell.column != 2):
                                # print(cell.value)
                                if len(media_name_list) > cell.column:
                                    # print(cell.value)
                                    # print(media_name_list[cell.column])
                                    if check_by_re(cell.value, media_name_list[cell.column], media_name):
                                        event_begin_name = check_event_unit(event_begin_name,
                                                                            media_name_list[cell.column],
                                                                            cell,
                                                                            event_unit_list, sheet)
                                        audio_begin_name = check_audio_unit(audio_begin_name,
                                                                            media_name_list[cell.column],
                                                                            cell,
                                                                            audio_unit_list)
                                        mixer_begin_name = check_audio_mixer(mixer_begin_name,
                                                                             media_name_list[cell.column],
                                                                             cell,
                                                                             audio_mixer_list)
                                    else:
                                        print_error(
                                            media_name + "：" + media_name_list[cell.column] + "未通过" + cell.value)

                else:
                    for cell in sheet[check_cell.row]:
                        if cell.column != 1 and cell.value:
                            # print(cell.value)
                            if len(media_name_list) > cell.column:
                                # print(cell.value)
                                # print(media_name_list[cell.column])
                                if check_by_re(cell.value, media_name_list[cell.column], media_name):
                                    event_begin_name = check_event_unit(event_begin_name,
                                                                        media_name_list[cell.column],
                                                                        cell,
                                                                        event_unit_list, sheet)
                                    audio_begin_name = check_audio_unit(audio_begin_name,
                                                                        media_name_list[cell.column],
                                                                        cell,
                                                                        audio_unit_list)
                                    mixer_begin_name = check_audio_mixer(mixer_begin_name,
                                                                         media_name_list[cell.column],
                                                                         cell,
                                                                         audio_mixer_list)
                                else:
                                    print_error(media_name + "：" + media_name_list[cell.column] + "未通过" + cell.value)




            else:
                print_error(media_name + "：" + media_name_list[1] + "未找到，请检查命名是否正确")

            break

    if sys_name_flag == 0:
        print_error(media_name + "：" + media_name_list[0] + "未找到，请检查命名是否正确")
    return is_pass


# 检查命名规范测试
# audio_unit_list = []
# event_unit_list = []
# audio_mixer_list = []
# name_list = ['Char_Skill_C01_Counter', 'Amb_A01_MgSt_Crowd_Street_Day_LP', 'Imp_Greatsword_Chop_Flesh_Hvy',
#              'Char_Skill_C01_Execution2_Stand',
#              'Char_Skill_C01_Strafe_Exit', "Mon_Mob_Skill_MN01_Da", "Mon_Mob_Mov_MN01_Da", "Mon_Mob_Mov_MN_Da"]
# for name in name_list:
#     if check_by_wb(name, audio_unit_list, event_unit_list, audio_mixer_list):
#         # print(name)
#         pass

# unit列表长度排序
# audio_unit_list.sort(key=len)
# event_unit_list.sort(key=len)
#
# pprint("audio_unit_list：")
# pprint(audio_unit_list)
# print()
#
# pprint("event_unit_list：")
# pprint(event_unit_list)
# print()
#
# pprint("audio_mixer_list：")
# pprint(audio_mixer_list)
