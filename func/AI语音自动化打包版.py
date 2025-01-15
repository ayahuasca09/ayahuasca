# xml读取库
import shutil
import os
import sys
from os.path import abspath, dirname
import openpyxl
import re
# xml写入库
from xml.dom.minidom import Document, parse
from pprint import pprint

# 自定义库
import comlib.excel_h as excel_h
import comlib.config as config
import comlib.oi_h as oi_h
import comlib.csv_h as csv_h
import comlib.exe_h as exe_h
import comlib.cloudfeishu_h as cloudfeishu_h

"""****************数据获取******************"""

# 获取文件所在目录
py_path = ""
if hasattr(sys, 'frozen'):
    py_path = dirname(sys.executable)
elif __file__:
    py_path = dirname(abspath(__file__))

# 飞书在线文档下载
media_sheet_token_dict = config.es_sheet_token_dict
c_vo_sheet_name = 'C类（操控演员表演，接管镜头）剧情语音'
d_vo_sheet_name = 'D类（操控演员表演，不接管镜头）剧情语音'
c_vo_sheet_token = media_sheet_token_dict[c_vo_sheet_name]
d_vo_sheet_token = media_sheet_token_dict[d_vo_sheet_name]
print(c_vo_sheet_name + "：音频资源表更新")
cloudfeishu_h.download_cloud_sheet(c_vo_sheet_token,
                                   os.path.join(py_path, "Excel", c_vo_sheet_name) + '.xlsx')
print(d_vo_sheet_name + "：音频资源表更新")
cloudfeishu_h.download_cloud_sheet(d_vo_sheet_token,
                                   os.path.join(py_path, "Excel", d_vo_sheet_name) + '.xlsx')

# 获取wsources文件
external_input_path = os.path.join(py_path, config.external_input_path)

# 获取.csv文件
csv_mediainfo_path = os.path.join(py_path, config.csv_mediainfo_path)
csv_wwisecookie_path = os.path.join(py_path, config.csv_wwisecookie_path)
csv_DT_AudioPlotInfo_path = os.path.join(py_path, config.csv_DT_AudioPlotInfo_path)
# csv_DT_AudioPlotSoundInfo_path = os.path.join(py_path, config.csv_DT_AudioPlotSoundInfo_path)

# 获取.xlsx文件
excel_mediainfo_path = os.path.join(py_path, config.excel_mediainfo_path)
excel_wwisecookie_path = os.path.join(py_path, config.excel_wwisecookie_path)
excel_DT_AudioPlotInfo_path = os.path.join(py_path, config.excel_DT_AudioPlotInfo_path)
# excel_DT_AudioPlotSoundInfo_path = os.path.join(py_path, config.excel_DT_AudioPlotSoundInfo_path)

# 获取相应的excel表信息
sheet_mediainfo, wb_mediainfo = excel_h.excel_get_sheet(excel_mediainfo_path, 'Sheet1')
sheet_wwisecookie, wb_wwisecookie = excel_h.excel_get_sheet(excel_wwisecookie_path, 'Sheet1')
sheet_DT_AudioPlotInfo, wb_DT_AudioPlotInfo = excel_h.excel_get_sheet(excel_DT_AudioPlotInfo_path, 'Sheet1')
# sheet_DT_AudioPlotSoundInfo, wb_DT_AudioPlotSoundInfo = excel_h.excel_get_sheet(excel_DT_AudioPlotSoundInfo_path,
#                                                                                 'Sheet1')

# 获取媒体资源文件列表
wav_path = os.path.join(py_path, "New_Media")

# excel表标题列
excel_es_title_list = config.excel_es_title_list

# 获取当前目录下要遍历的表路径列表
file_name_list = excel_h.excel_get_path_list(os.path.join(py_path, "Excel"))

# 标题列索引获取
mediainfo_title_dict = excel_h.excel_get_all_sheet_title_column(sheet_mediainfo)
wwisecookie_title_dict = excel_h.excel_get_all_sheet_title_column(sheet_wwisecookie)
DT_AudioPlotInfo_title_dict = excel_h.excel_get_all_sheet_title_column(sheet_DT_AudioPlotInfo)
# DT_AudioPlotSoundInfo_dict = excel_h.excel_get_all_sheet_title_column(sheet_DT_AudioPlotSoundInfo)

# 复制列
copy_wwisecookie_list = [wwisecookie_title_dict['MediaInfoId'], wwisecookie_title_dict['MediaInfoId'],
                         wwisecookie_title_dict['ExternalSourceName'], wwisecookie_title_dict['MediaInfoId']]
# copy_DT_AudioPlotInfo_list = [DT_AudioPlotSoundInfo_dict['Row'], DT_AudioPlotSoundInfo_dict['PlotID'],
#                               DT_AudioPlotSoundInfo_dict['Extern Source Name'],
#                               DT_AudioPlotSoundInfo_dict['Extern Source ID']]
#
# copy_dict = {
#     DT_AudioPlotSoundInfo_dict['Row']: wwisecookie_title_dict['MediaInfoId'],
#     DT_AudioPlotSoundInfo_dict['PlotID']: wwisecookie_title_dict['MediaInfoId'],
#     DT_AudioPlotSoundInfo_dict['Extern Source Name']: wwisecookie_title_dict['ExternalSourceName'],
#     DT_AudioPlotSoundInfo_dict['Extern Source ID']: wwisecookie_title_dict['MediaInfoId']
# }
#
# copy_DT_AudioPlotSoundInfo_list = copy_DT_AudioPlotInfo_list

_, wem_list = oi_h.get_type_file_name_and_path('.wem', config.external_output_path)

# 获取写死的cookie数据表
es_vo_data_dict = config.es_vo_data_dict

"""**************写xml**************"""
# 创建一个对象
doc = Document()
# 创建根节点
root = doc.createElement("ExternalSourcesList")
root.setAttribute("SchemaVersion", "1")

# 如果保存在Wwise工程下，此Root为Wwise工程的相对路径，则不需要输入绝对路径，仅需输入.wav的名字
# 否则取文件的绝对路径
# root.setAttribute("Root", "UngeneratedExternalSources")
# 添加载Document对象上
doc.appendChild(root)

"""*****************功能检测区******************"""
"""语音ID生成"""


def create_es_id(name):
    # 以'_'分隔字符串，组成列表
    parts = name.split('_')

    # 确保列表有足够的元素
    if len(parts) < 5:
        raise ValueError(name + "：输入字符串格式不正确")

    # 根据第3个元素确定数字
    third_element = parts[2]
    if third_element == 'C':
        number = '1'
    elif third_element == 'D':
        number = '2'
    else:
        raise ValueError(name + "以_切分思维第三个元素非C或D，请检查")

    # 拼接数字及第4和第5个元素
    id_str = number + parts[3] + parts[4]
    es_id = int(id_str)

    return es_id


"""xml创建元素"""


def xml_create_element(media_path):
    source = doc.createElement("Source")
    source.setAttribute("Path", media_path)
    if "CG" in media_path:
        source.setAttribute("Conversion", "CG")
    elif "VO" in media_path:
        source.setAttribute("Conversion", "VO")
    root.appendChild(source)


"""写入excel_DT_AudioPlotInfo_path表"""


def write_dt_excel(vo_id, cell_sound, dt_sheet, cell_es_type):
    if cell_es_type in config.es_event_id_dict:
        event_id = config.es_event_id_dict[cell_es_type]
        # 插入为空的行
        insert_row = dt_sheet.max_row + 1
        value_dict = {
            "Row": vo_id,
            'PlotID': vo_id,
            'Extern Source Name': cell_sound.value + ".wem",
            'Extern Source ID': vo_id,
            'Event ID': event_id,

        }
        for cell in list(dt_sheet.rows)[0]:
            if cell.value in value_dict:
                dt_sheet.cell(row=insert_row, column=cell.column).value = value_dict[cell.value]
    else:
        oi_h.print_error(cell_sound.value + ":ES Type请映射请补充在config")


"""写入media_info excel表"""


def write_media_info_excel(vo_id, cell_sound):
    # 插入为空的行
    insert_row = sheet_mediainfo.max_row + 1
    value_dict = {
        "Name": vo_id,
        'ExternalSourceMediaInfoId': vo_id,
        'MediaName': cell_sound.value + ".wem",
        'CodecID': 4,
        'bIsStreamed': 'TRUE',
        'bUseDeviceMemory': "FALSE",
        'MemoryAlignment': 0,
        'PrefetchSize': 0
    }
    for cell in list(sheet_mediainfo.rows)[0]:
        if cell.value in value_dict:
            sheet_mediainfo.cell(row=insert_row, column=cell.column).value = value_dict[cell.value]


"""写入wwise_cookie excel表"""


def write_wwise_cookie_excel(vo_id, cell_sound, external_cookiedata):
    # 插入为空的行
    insert_row = sheet_wwisecookie.max_row + 1
    value_dict = {
        "Name": vo_id,
        'ExternalSourceCookie': int(external_cookiedata[1]),
        'ExternalSourceName': str(external_cookiedata[0]),
        'MediaInfoId': vo_id,
        'MediaName': cell_sound.value + ".wem"
    }
    for cell in list(sheet_wwisecookie.rows)[0]:
        if cell.value in value_dict:
            sheet_wwisecookie.cell(row=insert_row, column=cell.column).value = value_dict[cell.value]


"""删除相应的.wem文件"""


def delete_cancel_wem(media_name):
    # [{'VO_C01_15_World.wem': 'S:\\Ver_1.0.0\\Project\\Content\\Audio\\GeneratedExternalSources\\Android\\VO_C01_15_World.wem'},
    #  {'VO_C01_33_Battle.wem': 'S:\\Ver_1.0.0\\Project\\Content\\Audio\\GeneratedExternalSources\\Android\\VO_C01_33_Battle.wem'},
    #  {'VO_C02_06_Menu.wem': 'S:\\Ver_1.0.0\\Project\\Content\\Audio\\GeneratedExternalSources\\Android\\VO_C02_06_Menu.wem'},
    #  {'VO_C02_33_Battle.wem': 'S:\\Ver_1.0.0\\Project\\Content\\Audio\\GeneratedExternalSources\\Android\\VO_C02_33_Battle.wem'},]
    for wem_dict in wem_list:
        for wem_name in wem_dict:
            # print(wem_name)
            if media_name in wem_name:
                if os.path.isfile(wem_dict[wem_name]):
                    os.remove(wem_dict[wem_name])
                    oi_h.print_warning("wem文件删除：" + str(wem_dict[wem_name]))


"""移除标记为cancel的所有内容"""


def delete_cancel_content(table_name_list):
    # pprint(table_name_list)
    # 记录要删除的行的索引
    mediainfo_del_list = []
    plot_del_list = []
    plot_sound_del_list = []
    # 删除excel表的相应内容

    for row in range(2, sheet_mediainfo.max_row + 1):
        cell = sheet_mediainfo.cell(row=row, column=mediainfo_title_dict['MediaName'])
        # pprint(cell.value)
        flag = 0
        if cell.value:
            for table_name in table_name_list:
                # 在media_info表中找到该ID，则标记
                if table_name in cell.value:
                    flag = 1
                    break
            # 未找到ID的则删除
            if flag == 0:
                # sheet_mediainfo.delete_rows(cell.row)
                # wb_mediainfo.save(excel_mediainfo_path)
                # sheet_wwisecookie.delete_rows(cell.row)
                # wb_wwisecookie.save(excel_wwisecookie_path)
                if cell.row not in mediainfo_del_list:
                    mediainfo_del_list.append(cell.row)
                oi_h.print_warning(cell.value + ":在mediainfo和wwisecookie中删除相关信息")
                delete_cancel_wem(cell.value)

                # 剧情语音表查找并删除
                for row in range(2, sheet_DT_AudioPlotInfo.max_row + 1):
                    plot_delete_cell = sheet_DT_AudioPlotInfo.cell(row=row, column=config.plot_media_name_column)
                    cell_value = plot_delete_cell.value
                    if cell_value == cell.value:
                        if plot_delete_cell.row not in plot_del_list:
                            plot_del_list.append(plot_delete_cell.row)
                        # sheet_DT_AudioPlotInfo.delete_rows(plot_delete_cell.row)
                        # wb_DT_AudioPlotSoundInfo.save(excel_DT_AudioPlotSoundInfo_path)
                        oi_h.print_warning(cell.value + ":在DT_AudioPlotInfo中删除相关信息")
                        break

                # 剧情音效表查找并删除
                # for row in range(2, sheet_DT_AudioPlotSoundInfo.max_row + 1):
                #     plot_sound_delete_cell = sheet_DT_AudioPlotSoundInfo.cell(row=row,
                #                                                               column=config.plot_media_name_column)
                #     cell_value = plot_sound_delete_cell.value
                #     if cell_value == cell.value:
                #         if plot_sound_delete_cell.row not in plot_del_list:
                #             plot_sound_del_list.append(plot_sound_delete_cell.row)
                #         # sheet_DT_AudioPlotSoundInfo.delete_rows(plot_sound_delete_cell.row)
                #         # wb_DT_AudioPlotInfo.save(excel_DT_AudioPlotInfo_path)
                #         oi_h.print_warning(cell.value + ":在DT_AudioPlotSoundInfo中删除相关信息")
                #         break

    # 列表多行删除，为避免行数出问题，需倒序删除
    # 从最后一行开始删除
    for row in sorted(mediainfo_del_list, reverse=True):
        sheet_mediainfo.delete_rows(row)
        sheet_wwisecookie.delete_rows(row)
    for row in sorted(plot_del_list, reverse=True):
        sheet_DT_AudioPlotInfo.delete_rows(row)
    # for row in sorted(plot_sound_del_list, reverse=True):
    #     sheet_DT_AudioPlotSoundInfo.delete_rows(row)
    # 保存excel表的信息
    wb_mediainfo.save(excel_mediainfo_path)
    wb_wwisecookie.save(excel_wwisecookie_path)
    # wb_DT_AudioPlotSoundInfo.save(excel_DT_AudioPlotSoundInfo_path)
    wb_DT_AudioPlotInfo.save(excel_DT_AudioPlotInfo_path)


"""检查是否以CG_External_或VO_External_开头"""


# 返回值为True或False

def check_prefix(string):
    pattern = r'^(CG_External_|VO_External_)'
    return re.match(pattern, string) is not None


"""资源命名规范检查"""


def check_name():
    is_pass = True
    for i in file_name_list:
        if (".xlsx" in i) and ("MediaInfoTable" not in i) and ("ExternalSourceDefaultMedia" not in i):
            # 拼接xlsx的路径
            file_path_xlsx = os.path.join(os.path.join(py_path, "Excel"), i)
            # 获取xlsx的workbook
            wb = openpyxl.load_workbook(file_path_xlsx)
            # 获取xlsx的所有sheet
            sheet_names = wb.sheetnames
            # 加载所有工作表
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                if not excel_h.is_sheet_empty(sheet):
                    title_colunmn_dict = excel_h.excel_get_sheet_title_column(sheet, excel_es_title_list)
                    if title_colunmn_dict:
                        # 获取音效名下的内容
                        for cell_sound in list(sheet.columns)[
                            excel_h.sheet_title_column("文件名", title_colunmn_dict) - 1]:
                            if cell_sound.value and (cell_sound.value != "文件名"):
                                """资源命名规范检查"""
                                check_list = [

                                ]
                                _, _, is_pass = oi_h.check_name_all(cell_sound.value)
                                if is_pass:
                                    # 检查是否以CG_External_或VO_External_开头
                                    if not check_prefix(cell_sound.value):
                                        is_pass = oi_h.print_error(
                                            cell_sound.value + "：请检查名称是否以CG_External_或VO_External_开头")
                                    if is_pass and sheet.cell(row=cell_sound.row,
                                                              column=excel_h.sheet_title_column("State",
                                                                                                title_colunmn_dict)).value != 'cancel':
                                        if cell_sound.value not in table_name_list:
                                            table_name_list.append(cell_sound.value)
                                        else:
                                            is_pass = oi_h.print_error(
                                                cell_sound.value + "：在表中有重复，请检查")

    # 文件清理
    delete_cancel_content(table_name_list)
    # pprint(table_name_list)

    return is_pass


"""自动化生成ES数据"""


def auto_gen_es_file(file_wav_dict):
    # 查找要导入的媒体文件里有没有对应的
    for file_wav_name in file_wav_dict:
        flag = 0
        # 读excel表
        # 提取规则：只提取xlsx文件
        for i in file_name_list:

            if (".xlsx" in i) and ("MediaInfoTable" not in i) and ("ExternalSourceDefaultMedia" not in i):
                # 拼接xlsx的路径
                file_path_xlsx = os.path.join(os.path.join(py_path, "Excel"), i)
                # 获取xlsx的workbook
                wb = openpyxl.load_workbook(file_path_xlsx)
                # 获取xlsx的所有sheet
                sheet_names = wb.sheetnames
                # 加载所有工作表
                for sheet_name in sheet_names:
                    sheet = wb[sheet_name]
                    if not excel_h.is_sheet_empty(sheet):
                        title_colunmn_dict = excel_h.excel_get_sheet_title_column(sheet, excel_es_title_list)
                        if title_colunmn_dict:
                            # 获取音效名下的内容
                            for cell_sound in list(sheet.columns)[
                                excel_h.sheet_title_column("文件名", title_colunmn_dict) - 1]:
                                if cell_sound.value and (cell_sound.value != "文件名"):

                                    """查找文件"""
                                    if (cell_sound.value in file_wav_name) and (
                                            cell_sound.value not in excel_name_list):
                                        excel_name_list.append(cell_sound.value)
                                        flag = 1
                                        if sheet.cell(row=cell_sound.row, column=excel_h.sheet_title_column("State",
                                                                                                            title_colunmn_dict)).value == 'AI':
                                            # print(file_wav_dict[file_wav_name])
                                            # 在表格中找到此音效才将其写入xml
                                            xml_create_element(file_wav_dict[file_wav_name])

                                            # 在状态处自动标记完成
                                            # sheet.cell(row=cell_sound.row,
                                            #            column=excel_h.sheet_title_column("State",
                                            #                                              title_colunmn_dict)).value = 'done'
                                            is_wwise_have = 0

                                            # 查找字典里有无对应的ES容器
                                            for external_sound in es_vo_data_dict:
                                                # for external_sound_dict in external_sound_list:
                                                cell_es_type = sheet.cell(
                                                    row=cell_sound.row,
                                                    column=excel_h.sheet_title_column("External_Type",
                                                                                      title_colunmn_dict)).value
                                                if external_sound == cell_es_type:
                                                    # 获取已有的媒体资源列表
                                                    have_media_list = excel_h.get_colunmn_one_list(
                                                        config.plot_media_name_column, sheet_mediainfo)
                                                    wem_name = cell_sound.value + ".wem"
                                                    if wem_name not in have_media_list:
                                                        # print(cell_es_type)
                                                        # print(external_sound_dict['shortId'])
                                                        # 查找mediainfo表的id有没有，没有则添加，有则修改内容
                                                        vo_id = create_es_id(cell_sound.value)
                                                        # 写入media_info excel表
                                                        write_media_info_excel(vo_id, cell_sound)
                                                        # # 写入wwise_cookie excel表
                                                        write_wwise_cookie_excel(vo_id, cell_sound,
                                                                                 es_vo_data_dict[external_sound])
                                                        write_dt_excel(vo_id, cell_sound, sheet_DT_AudioPlotInfo,
                                                                       cell_es_type)
                                                        oi_h.print_warning(file_wav_name + "：已导入")
                                                        is_wwise_have = 1
                                                        break
                                                    else:
                                                        oi_h.print_warning(file_wav_name + "：已替换")
                                                        is_wwise_have = 2
                                                        break

                                            if is_wwise_have == 0:
                                                oi_h.print_error(
                                                    file_wav_name + "：在Wwise中未建立相应容器，请检查填写是否正确或Wwise中容器是否存在")

                wb.save(file_path_xlsx)

        if flag == 0:
            oi_h.print_error(file_wav_name + "：在语音需求表中不存在，请检查名称是否正确或在表格中补充该名字")


"""*******************主程序*******************"""
excel_name_list = []
table_name_list = []

# 命名规范检查
if check_name():
    # 语音ES生成
    for language in config.language_list:
        wav_language_path = os.path.join(wav_path, language)
        file_wav_language_dict, _ = oi_h.get_type_file_name_and_path('.wav', wav_language_path)
        # print(file_wav_language_dict)
        auto_gen_es_file(file_wav_language_dict)
        # # xml文件写入
        # pprint(file_wav_dict)
        with open(config.es_xml_path, 'w+') as f:
            # 按照格式写入
            f.write(doc.toprettyxml())
            f.close()
        # pprint(doc.toprettyxml())
        # 复制xml为wsources
        shutil.copy2(os.path.join(py_path, config.es_xml_path),
                     external_input_path)

        # gen_external(language)
        exe_h.gen_ai_language(external_input_path)

        # 初始化xml
        doc = Document()
        root = doc.createElement("ExternalSourcesList")
        root.setAttribute("SchemaVersion", "1")
        doc.appendChild(root)

        # 将删除的xml内容写入wsources
        shutil.copy2(os.path.join(py_path, config.es_xml_path),
                     external_input_path)

    # 删除xml的内容
    doc = parse(config.es_xml_path)
    parent_node = doc.getElementsByTagName('ExternalSourcesList')[0]
    while parent_node.hasChildNodes():
        parent_node.removeChild(parent_node.firstChild)

    # 保存excel表的信息
    wb_mediainfo.save(excel_mediainfo_path)
    wb_wwisecookie.save(excel_wwisecookie_path)
    # wb_DT_AudioPlotSoundInfo.save(excel_DT_AudioPlotSoundInfo_path)
    wb_DT_AudioPlotInfo.save(excel_DT_AudioPlotInfo_path)

    # 将excel转为csv
    csv_h.excel_to_csv(excel_mediainfo_path, csv_mediainfo_path)
    csv_h.excel_to_csv(excel_wwisecookie_path, csv_wwisecookie_path)
    csv_h.excel_to_csv(excel_DT_AudioPlotInfo_path, csv_DT_AudioPlotInfo_path)
    # csv_h.excel_to_csv(excel_DT_AudioPlotSoundInfo_path, csv_DT_AudioPlotSoundInfo_path)

    # have_media_list = excel_h.get_colunmn_one_list(3, sheet_mediainfo)
    # pprint(have_media_list)

# else:
#     print(1)

# 清除复制的媒体资源
shutil.rmtree(wav_path)
os.mkdir(wav_path)
os.mkdir(os.path.join(wav_path, "Chinese"))
os.mkdir(os.path.join(wav_path, "English"))
os.mkdir(os.path.join(wav_path, "Japanese"))
os.mkdir(os.path.join(wav_path, "Korean"))
os.mkdir(os.path.join(wav_path, "SFX"))

# dt表导入UE
ue_csv_dt_es_path = os.path.join(py_path, "ue_csv_dt_es.py")
exe_h.run_unreal_editor_with_script(ue_csv_dt_es_path)
os.system("pause")
