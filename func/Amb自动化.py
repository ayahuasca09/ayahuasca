import openpyxl
import json
import sys
from os.path import abspath, dirname
import os
from pprint import pprint
import re
from waapi import WaapiClient
import shutil
from openpyxl.cell import MergedCell
import module.cloudfeishu.cloudfeishu_h as cloudfeishu_h
import module.oi.oi_h as oi_h
import config
import module.excel.excel_h as excel_h
import module.waapi.waapi_h as waapi_h
from pathlib import Path
import 命名规范检查

"""根目录获取"""
# 获取当前脚本的文件名及路径
script_name_with_extension = __file__.split('/')[-1]
# 去掉扩展名
script_name = script_name_with_extension.rsplit('.', 1)[0]
# 替换为files
root_path = script_name.replace("func", "files")
# print(f"当前脚本的名字是: {root_path}")c

file_name_list = excel_h.excel_get_path_list(root_path)
# pprint(file_name_list)

amb_unit_path = r''

"""文件获取"""
# 规范检查表
excel_amb_path = os.path.join(root_path, config.excel_amb_config_path)
# print(wb)

"""在线表获取"""
# 规范检查表
# check_name_token = config.amb_sheet_token
# cloudfeishu_h.download_cloud_sheet(check_name_token, excel_amb_path)

"""amb媒体资源名称及文件夹映射"""
media_info_dict, _ = oi_h.get_type_file_name_and_path('.wav', os.path.join(root_path, 'New_Media'))

"""功能数据初始化"""
# switch创建列表
switch_create_list = []
# unit创建列表
unit_create_list = []
# rnd创建列表
rnd_create_list = []

"""*****************功能检测区******************"""

with WaapiClient() as client:
    """*****************Wwise功能区******************"""
    """设置obj的notes"""


    def set_obj_notes(obj_id, notes_value):
        args = {
            'object': obj_id,
            'value': notes_value
        }
        client.call("ak.wwise.core.object.setNotes", args)
        # print_warning(obj_name + "描述更改为：" + notes_value)


    """设置对象属性"""


    def set_obj_property(obj_id, property, value):
        args = {
            "object": obj_id,
            "property": property,
            "value": value
        }
        client.call("ak.wwise.core.object.setProperty", args)
        # print(name_2 + ":" + str(trigger_rate))


    """设置子集switch"""


    def set_switch_child(obj_child, switch_child):
        args = {
            'child': obj_child,
            'stateOrSwitch': switch_child
        }
        client.call("ak.wwise.core.switchContainer.addAssignment", args)


    # 设置对象引用
    def set_obj_refer(obj_id, refer, value):
        args = {
            "object": obj_id,
            "reference": refer,
            "value": value
        }
        client.call("ak.wwise.core.object.setReference", args)


    """obj创建"""


    def create_wwise_obj(type, parent, name, notes):
        unit_object = None
        if name != "Amb":
            pprint(type)
            if type == "RandomSequenceContainer":
                pprint(name)
                args = waapi_h.args_rnd_create(parent,
                                               type, name, notes)
            else:
                args = waapi_h.args_object_create(parent,
                                                  type, name, notes)
            unit_object = client.call("ak.wwise.core.object.create",
                                      args)
            oi_h.print_warning(name + "：" + type + "已创建")

        return unit_object


    """获取Wwise中的Unit列表"""


    def get_wwise_type_list(root_path, type):
        # 获取Wwise所有的Audio Unit
        wwise_bus_name_all_list = client.call("ak.wwise.core.object.get",
                                              waapi_h.waql_by_type(type, root_path),
                                              options=config.options)['return']
        wwise_bus_name_name_dict = {item['name']: item['id'] for item in wwise_bus_name_all_list}
        # pprint(wwise_bus_name_name_dict)
        # {'Amb': '{4BB49221-4DC1-4A05-A57C-3D42FA570675}',
        #  'Amb_A01': '{415AC456-D390-43D5-A09E-B5D30B253D76}',
        #  'CG': '{BBB332B4-C230-4D60-9ABA-91A36D968C11}'}
        return wwise_bus_name_name_dict


    """获取Switch对象是否已指派"""


    def get_switch_is_assign(switch_id):
        switch_is_assign_list = client.call("ak.wwise.core.switchContainer.getAssignments",
                                            {"id": switch_id})['return']
        state_or_switch_list = [item['stateOrSwitch'] for item in switch_is_assign_list]
        # pprint(state_or_switch_values)
        return state_or_switch_list


    """获取Wwise中子级的Unit列表"""


    def get_wwise_children_list(root_path):
        # 获取Wwise所有的Audio Unit
        wwise_bus_name_all_list = client.call("ak.wwise.core.object.get",
                                              waapi_h.waql_find_children(root_path),
                                              options=config.options)['return']
        wwise_bus_name_name_dict = {item['name']: item['id'] for item in wwise_bus_name_all_list}
        return wwise_bus_name_name_dict


    """amb创建"""


    def create_wwise_amb(rnd_list, switch_list):
        # switch创建
        for amb_name in switch_list:

            switch_have_dict = get_wwise_type_list(config.wwise_amb_global_path, "SwitchContainer")

            if amb_name not in switch_have_dict:
                # 找到的父级为unit
                unit_parent = oi_h.find_longest_prefix_key(amb_name, switch_have_dict)
                if not unit_parent:
                    switch_object = create_wwise_obj("SwitchContainer",
                                                     config.wwise_amb_global_path, amb_name,
                                                     "")
                else:
                    switch_have_dict = get_wwise_type_list(config.wwise_amb_global_path, "SwitchContainer")
                    switch_object = create_wwise_obj("SwitchContainer",
                                                     switch_have_dict[unit_parent], amb_name,
                                                     "")
        # rnd创建
        for amb_name in rnd_list:
            switch_have_dict = get_wwise_type_list(config.wwise_amb_global_path, "SwitchContainer")
            rnd_have_dict = get_wwise_type_list(config.wwise_amb_global_path, "RandomSequenceContainer")
            if amb_name not in rnd_have_dict:
                unit_parent = oi_h.find_longest_prefix_key(amb_name, switch_have_dict)
                if not unit_parent:
                    rnd_object = create_wwise_obj("RandomSequenceContainer",
                                                  config.wwise_amb_global_path, amb_name,
                                                  "")
                else:
                    rnd_object = create_wwise_obj("RandomSequenceContainer",
                                                  switch_have_dict[unit_parent], amb_name,
                                                  "")


    """获取需要创建的amb列表"""


    def get_create_amb_name_list():
        # 遍历每一行
        for row in amb_config_sheet.iter_rows():
            amb_name = None
            # 遍历每一列
            for cell in row:
                cell_value, is_merge = excel_h.check_is_mergecell(cell, amb_config_sheet)
                if cell_value:
                    is_unit = False
                    # print(cell_value)
                    if not amb_name:
                        amb_name = cell_value
                        is_unit = True
                    else:
                        amb_name += "_" + cell_value
                        fill = cell.fill
                        if fill.fill_type == 'solid':
                            if fill.start_color.index == "FFD9F3FD":
                                is_unit = True
                    is_switch = False

                    if is_unit:
                        if amb_name not in unit_create_list:
                            unit_create_list.append(amb_name)
                    # 是否为switch容器检查
                    if is_merge:
                        is_switch = True
                    else:
                        if amb_config_sheet.cell(row=cell.row,
                                                 column=cell.column + 1).value:
                            is_switch = True
                    if is_switch:
                        if amb_name not in switch_create_list:
                            switch_create_list.append(amb_name)
                    else:
                        if amb_name not in rnd_create_list:
                            rnd_create_list.append(amb_name)

            # print()


    """*****************主程序处理******************"""
    # 撤销开始
    # client.call("ak.wwise.core.undo.beginGroup")

    """环境音结构创建"""

    # 获取xlsx的workbook
    wb = openpyxl.load_workbook(excel_amb_path)
    amb_config_sheet = wb[config.sheet_amb_config]
    amb_media_sheet = wb[config.sheet_amb_media]
    get_create_amb_name_list()
    switch_create_list.sort(key=len)
    rnd_create_list.sort(key=len)
    unit_create_list.sort(key=len)
    # pprint(switch_create_list)
    # pprint(rnd_create_list)
    # pprint(unit_create_list)
    create_wwise_amb(rnd_create_list, switch_create_list)

    """******************对象删除清理********************"""
    """对象删除"""


    def delete_obj(obj_id, obj_name, obj_type):
        args = {
            "object": "%s" % obj_id
        }
        client.call("ak.wwise.core.object.delete", args)
        oi_h.print_warning(obj_name + "(" + obj_type + ")" + ":已删除")


    """删除状态"""


    def delete_check(wwise_obj_dict, excel_list, obj_type):
        for wwise_obj in wwise_obj_dict:
            # pprint(wwise_obj_dict['name'])
            if wwise_obj not in excel_list:
                # State/StateGroup删除
                delete_obj(wwise_obj_dict[wwise_obj], wwise_obj, obj_type)
                # StateGroupUnit删除


    switch_dict = get_wwise_type_list(config.wwise_amb_global_path, "SwitchContainer")
    rnd_dict = get_wwise_type_list(config.wwise_amb_global_path, "RandomSequenceContainer")

    # 查找state/switch，再跟表格比对有没有，没有就删除资源及事件引用
    delete_check(switch_dict, switch_create_list, "SwitchContainer")
    delete_check(rnd_dict, rnd_create_list, "RandomSequenceContainer")

    """******************Switch/State指派********************"""
    # 获取switch容器
    switch_dict = get_wwise_type_list(config.wwise_amb_global_path, "SwitchContainer")
    # 获取stategroup容器
    switchgroup_dict = get_wwise_type_list(config.wwise_state_path, "StateGroup")
    # pprint(switchgroup_dict)

    # 遍历switch容器获取其子级
    for switch_name in switch_dict:
        """设置属性为Continuous"""
        set_obj_property(switch_dict[switch_name], "SwitchBehavior", 1)
        # statrgroup和switch的映射需要手动添加
        if switch_name in config.amb_state_dict:
            # 获取已经指派的对象列表
            have_state_assign_list = get_switch_is_assign(switch_dict[switch_name])
            """StateGroup指派"""
            stategroup_id = switchgroup_dict[config.amb_state_dict[switch_name]]
            set_obj_refer(switch_dict[switch_name], "SwitchGroupOrStateGroup",
                          stategroup_id)
            # set_obj_refer(switch_dict[switch_name], "DefaultSwitchOrState", "{BF22F66E-BA24-44FB-884C-8C591B785423}")

            """子级State映射"""
            switch_children_dict = get_wwise_children_list(switch_dict[switch_name])
            # pprint(switch_children_dict)
            state_children_dict = get_wwise_children_list(stategroup_id)
            for state_children in state_children_dict:
                flag = 0
                for switch_children in switch_children_dict:
                    if state_children in switch_children:
                        if state_children_dict[state_children] not in have_state_assign_list:
                            set_switch_child(switch_children_dict[switch_children], state_children_dict[state_children])
                            oi_h.print_warning(switch_children + "：指派State为" + state_children)
                        flag = 1
                        break
                # if flag == 0:
                #     """状态值中没有的容器默认设为None"""
                #     for switch_children in switch_children_dict:
                #         if "None" in switch_children:
                #             if state_children_dict[state_children] not in have_state_assign_list:
                #                 set_switch_child(switch_children_dict[switch_children],
                #                                  state_children_dict[state_children])

        # 撤销结束
        # client.call("ak.wwise.core.undo.endGroup", displayName="rnd创建撤销")

        """******************音频资源导入********************"""
        """修改Wwise内容的名字"""


        def change_name_by_wwise_content(obj_id, name, old_name, obj_type):
            args = {
                "objects": [
                    {

                        "object": obj_id,
                        "name": name,
                    }
                ]
            }
            client.call("ak.wwise.core.object.set", args)
            oi_h.print_warning(old_name + "(" + obj_type + ")改名为：" + name)


        """获取表格中的事件描述和状态所在的列"""


        def get_descrip_and_status_column():
            pos_column = None
            require_name_column = None
            status_column = None
            if list(amb_config_sheet.rows)[0]:
                for cell in list(amb_media_sheet.rows)[0]:
                    if cell.value:
                        if '资源名称' == str(cell.value):
                            require_name_column = cell.column
                            # print(require_name_column)
                        elif '资源描述' == str(cell.value):
                            status_column = cell.column
                            # print(status_column)

            return require_name_column, status_column


        """查找对象"""


        def find_obj(args):
            options = {
                'return': ['name', 'id', 'path', 'notes', 'originalWavFilePath', 'type', 'parent', "Tempo"]

            }
            obj_sub_list = client.call("ak.wwise.core.object.get", args, options=options)['return']
            if not obj_sub_list:
                obj_sub_id = ""
                obj_sub_path = ""
            else:
                obj_sub_id = obj_sub_list[0]['id']
                obj_sub_path = obj_sub_list[0]['path']
            return obj_sub_list, obj_sub_id, obj_sub_path


        """通用内容创建"""


        def create_obj_content(obj_parent_path, obj_type,
                               obj_name, obj_desc, obj_id):
            flag = 0
            obj_id = ""
            obj_path = ""
            obj_list = client.call("ak.wwise.core.object.get",
                                   waapi_h.waql_by_type(obj_type, obj_parent_path),
                                   options=config.options)['return']
            # obj已存在
            for obj_dict in obj_list:
                if obj_dict['name'] == obj_name:
                    flag = 1
                    # 设置notes
                    set_obj_notes(obj_dict['id'], obj_desc)
                    obj_id = obj_dict['id']
                    obj_path = obj_dict['path']
                    break
                # 改名了但描述一致
                elif obj_dict['notes'] == obj_desc:
                    flag = 2
                    # 改名
                    change_name_by_wwise_content(obj_dict['id'], obj_name, obj_dict['name'],
                                                 obj_type)
                    obj_id = obj_dict['id']
                    obj_path = obj_dict['path']
                    break
            # obj不存在，需要创建
            # print(flag)
            # 不存在则创建
            if flag == 0:
                # 创建的rnd属性
                args = {
                    # 选择父级
                    "parent": obj_parent_path,
                    # 创建类型名称
                    "type": obj_type,
                    "name": obj_name,
                    "notes": obj_desc,
                }
                obj_object = client.call("ak.wwise.core.object.create", args)
                # print(obj_object)
                obj_id = obj_object['id']
                _, _, obj_path = find_obj(
                    {'waql': ' "%s" ' % obj_id})
                oi_h.print_warning(obj_object['name'] + ":" + obj_type + "已创建")

            return obj_id, obj_path


        """音乐资源导入"""


        def import_media(amb_name, rnd_path):
            flag = 0
            # 音乐资源导入
            for media_info_name in media_info_dict:
                # check_name = media_info_name.replace(".wav", "")
                if amb_name in media_info_name:
                    media_path = media_info_dict[media_info_name]
                    # import_media_in_track(media_path, amb_track_path)
                    client.call("ak.wwise.core.audio.import",
                                waapi_h.args_sfx_create(media_path, rnd_path, amb_name, "Amb"))
                    oi_h.print_warning((amb_name + ":媒体资源已导入"))
                    flag = 1
                    break


        """创建新的amb"""


        def create_amb_content(amb_name, amb_desc):
            flag = 0
            for rnd_name in rnd_dict:
                if rnd_name in amb_name:
                    amb_sound_id, _ = create_obj_content(rnd_dict[rnd_name], "Sound",
                                                         amb_name,
                                                         amb_desc, "")
                    _, _, obj_path = find_obj(
                        {'waql': ' "%s" ' % rnd_dict[rnd_name]})
                    import_media(amb_name, obj_path)
                    flag = 1
                    break
            if flag == 0:
                oi_h.print_error(amb_name + "：无相应前缀的rnd父级，请检查命名是否有误或先创建结构")


        amb_name_list = []
        value_desc_list = []

        value, value_desc = get_descrip_and_status_column()
        if value:
            for cell in list(amb_media_sheet.columns)[value - 1]:
                if (cell.value) and (not 命名规范检查.check_is_chinese(cell.value)):
                    """名称查重"""
                    if cell.value not in amb_name_list:
                        amb_name_list.append(cell.value)
                        # pprint(cell.value)
                        """资源描述"""
                        value_desc_value = amb_media_sheet.cell(row=cell.row,
                                                                column=value_desc).value
                        """资源描述检查"""
                        if value_desc_value:
                            if value_desc_value not in value_desc_list:
                                value_desc_list.append(value_desc_value)

                                """音频资源创建"""
                                create_amb_content(cell.value,
                                                   value_desc_value)

                            else:
                                oi_h.print_error(value_desc_value + "：表格中有重复项描述，请检查")
                        else:
                            oi_h.print_error(cell.value + "：描述不能为空，请补充")
                    else:
                        oi_h.print_error(cell.value + "：表格中有重复项描述，请检查")

    """同步表中删除的内容"""
    sound_have_dict = get_wwise_type_list(config.wwise_amb_global_path, "Sound")

    # 查找state/switch，再跟表格比对有没有，没有就删除资源及事件引用
    delete_check(sound_have_dict, amb_name_list, "Sound")

    # 清除复制的媒体资源
    shutil.rmtree(os.path.join(root_path, 'New_Media'))
    os.mkdir(os.path.join(root_path, 'New_Media'))
