import openpyxl
import os
from . import oi_h
from openpyxl.cell import MergedCell

"""检查表格是否为空"""


def is_sheet_empty(sheet):
    # 遍历工作表中的所有行和列，检查是否有非空单元格
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                return False
    return True


"""合并单元格检查"""


def check_is_mergecell(cell, sheet):
    is_merge = False
    if isinstance(cell, MergedCell):  # 判断该单元格是否为合并单元格
        for merged_range in sheet.merged_cells.ranges:  # 循环查找该单元格所属的合并区域
            if cell.coordinate in merged_range:
                # 获取合并区域左上角的单元格作为该单元格的值返回
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
        is_merge = True
    return cell.value, is_merge


"""合并单元格检查及获取合并行区间"""


def check_mergecell_row(cell, sheet):
    is_merge = False
    merge_min_row = 0
    merge_max_row = 0
    for merged_range in sheet.merged_cells.ranges:  # 循环查找该单元格所属的合并区域
        if cell.coordinate in merged_range:
            # 获取合并区域左上角的单元格作为该单元格的值返回
            cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            merge_min_row = merged_range.min_row
            merge_max_row = merged_range.max_row
            is_merge = True
            break

    return cell.value, is_merge, merge_min_row, merge_max_row


"""打开工作簿并获取工作表"""


def excel_get_sheet(path, sheetname):
    # 打开工作簿
    wb = openpyxl.load_workbook(path)
    sheet = None
    # print(wb)
    if sheetname:
        # 获取工作表
        sheet = wb[sheetname]

    # print(sheet)
    return sheet, wb


"""获取表头数据所在列"""


def excel_get_sheet_title_column(sheet, title_list):
    # 标题所在列的字典映射
    title_colunmn_dict = {}
    # 遍历第一行
    for cell in list(sheet.rows)[0]:
        for title_name in title_list:
            if title_name in str(cell.value):
                title_colunmn_dict[title_name] = cell.column
    return title_colunmn_dict


"""获取表头所有数据所在列"""


def excel_get_all_sheet_title_column(sheet):
    # 标题所在列的字典映射
    title_colunmn_dict = {}
    # 遍历第一行
    for cell in list(sheet.rows)[0]:
        title_colunmn_dict[cell.value] = cell.column
    return title_colunmn_dict


"""获取当前目录下要遍历的表路径列表"""


def excel_get_path_list(py_path):
    file_names = []
    for i in os.walk(py_path):
        file_names.append(i)
    # pprint("输出文件夹下的文件名：")
    if file_names:
        return file_names[0][2]





"""获取标题所在列"""


def sheet_title_column(title_name, title_colunmn_dict):
    file_name = ""
    title_colunmn = ""
    if title_name in title_colunmn_dict.keys():
        # 读取测试
        title_colunmn = title_colunmn_dict[title_name]
    else:
        oi_h.print_error(title_name + ":标题列不存在，请检查表格中是否有该标题的列表")

    return title_colunmn


"""检测字符串是否含有中文"""


def check_is_chinese(string):
    if isinstance(string, int):
        return False
    else:
        for ch in string:
            if u'\u4e00' <= ch <= u'\u9fff':
                return True

    return False


"""遍历第n列内容并生成列表,列表生成的是单列"""


def get_colunmn_one_list(column_index, sheet):
    # 初始化列表
    column_data = []

    # 遍历第 8 列，从第二行开始
    for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True):
        # row 是一个元组，只有一个元素
        if row[0]:
            if not check_is_chinese(row[0]):
                if row not in column_data:
                    column_data.append(row[0])

    return column_data


"""遍历第n列内容并生成列表,列表生成的是元组"""


def get_colunmn_multi_list(column_min, column_max, sheet):
    # 初始化列表
    column_data = []

    # 遍历第 8 列，从第二行开始
    for row in sheet.iter_rows(min_row=2, min_col=column_min, max_col=column_max, values_only=True):
        # row 是一个元组，只有一个元素
        if row not in column_data:
            column_data.append(row)

    return column_data


# 测试
# sheet, wb = excel_get_sheet(r"F:\pppppy\SP\module\waapi\waapi_Auto_Import_ExternalSource\MediaInfoTable.xlsx", "Sheet1")
# column_data = get_colunmn_one_list(1, sheet)
# print(column_data)

"""创建表格ID"""


# id_list：获取当前id表中的所有id
# id_type_dict:获取当前类型的id区间段列表
# id_type_name：要获取的区间段类型
# id_sheet：id表的sheet
def create_id(id_type_dict, id_type_name, id_sheet):
    # 获取目前已有的ID列表
    id_list = get_colunmn_one_list(1, id_sheet)
    for id_type in id_type_dict:
        if id_type in id_type_name:
            # print(id_type)
            id_min = id_type_dict[id_type]['min']
            id_max = id_type_dict[id_type]['max']
            for i in range(id_min, id_max):
                if i not in id_list:
                    return i


# 测试
# sheet_mediainfo, wb_mediainfo = excel_get_sheet(
#     r"F:\pppppy\SP\module\waapi\waapi_Auto_Import_ExternalSource\MediaInfoTable.xlsx", 'Sheet1')
# id_l = create_id(config.es_id_config,
#                  r"F:\pppppy\SP\module\waapi\waapi_Auto_Import_ExternalSource\B类（LevelSequence）剧情语音.xlsx",
#                  sheet_mediainfo)
# print(id_l)


"""将表1的多列复制到表二的多列"""


# source_columns_list：源需要复制的列的索引列表
# target_columns_list：目标需要复制的列的索引列表

def copy_excel1_to_excel2_n_column(sheet1, sheet2, wb2, source_columns_list, target_columns_list, excel2_path):
    # 定义源列和目标列的对应关系

    # 遍历每一行并复制指定列
    for row in sheet1.iter_rows(min_row=1, max_row=sheet1.max_row):
        for src_col, tgt_col in zip(source_columns_list, target_columns_list):
            sheet2.cell(row=row[0].row, column=tgt_col).value = row[src_col - 1].value

    # 保存目标 Excel 文件
    wb2.save(excel2_path)
