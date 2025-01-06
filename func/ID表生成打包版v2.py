from waapi import WaapiClient
import os
import re
import pandas as pd
import tkinter as tk
from tkinter import messagebox
import openpyxl

# 新增特性：优化id生成算法

# 自定义库
import comlib.excel_h as excel_h
import comlib.config as config
import comlib.oi_h as oi_h
import comlib.waapi_h as waapi_h
import comlib.exe_h as exe_h

from os.path import abspath, dirname

import sys


def get_py_path():
    py_path = ""
    if hasattr(sys, 'frozen'):
        py_path = dirname(sys.executable)
    elif __file__:
        py_path = dirname(abspath(__file__))
    return py_path


# 打包版
root_path = get_py_path()

"""excel表路径"""
excel_path = os.path.join(root_path, config.excel_dt_audio_path)
# 获取excel ID表路径
sheet, wb = excel_h.excel_get_sheet(excel_path, config.dt_audio_sheet_name)

"""csv表路径"""
csv_path = os.path.join(root_path, config.csv_dt_audio_path)

"""UE工程路径"""
ue_audio_path = config.ue_event_path

"""Wwise路径"""
event_root = config.wwise_event_path


"""正则表达式匹配以某变量开头"""


def re_match_start_string(start_string, s):
    pattern = rf'^{re.escape(start_string)}'
    return bool(re.match(pattern, s))


"""ue路径转换"""


def convert_ue_event_path(old_path):
    new_path = ""
    # 获取从\Audio开始的内容
    result = re.search('(?<=\\\\Content).*', old_path).group()
    # \Audio\WwiseAudio\Events\v1\VO\VO_Game\VO_Game_C01\VO_Game_Battle_C01\AKE_Play_VO_Game_Battle_C01_01.uasset
    # /替换\
    result = result.replace("\\", "/")
    # /Audio/WwiseAudio/Events/v1/VO/VO_Game/VO_Game_C01/VO_Game_Battle_C01/AKE_Play_VO_Game_Battle_C01_01.uasset
    # 去除.uasset
    result = result.replace(".uasset", "")
    # /Audio/WwiseAudio/Events/v1/VO/VO_Game/VO_Game_C01/VO_Game_Battle_C01/AKE_Play_VO_Game_Battle_C01_01
    # 获取后缀
    result_list = result.split("/")
    # 拼接字符串
    if result_list[-1]:
        new_path = '/Script/AkAudio.AkAudioEvent\'/Game' + result + "." + result_list[-1] + '\''
        # print(result)
    return new_path


# 设置事件路径
def set_event_path(insert_row):
    flag = 0
    # Event路径赋值
    for asset_path in ue_audio_path_list:
        if event_dict['name'] in asset_path:
            sheet.cell(row=insert_row, column=config.audioevent_index).value = asset_path
            flag = 1
            break
    if flag == 0:
        sheet.cell(row=insert_row, column=config.audioevent_index).value = ""


# 获取是否循环并赋值
def get_is_loop():
    is_loop = False
    # 查找事件引用的对象
    obj_list = client.call("ak.wwise.core.object.get",
                           waapi_h.waql_by_type("Action", event_dict['id']),
                           options=config.options)['return']
    obj_sub_list, _, _ = waapi_h.find_obj(obj_list)

    if obj_sub_list:
        for obj_dict in obj_sub_list:
            if is_loop:
                break
            else:
                # 为play的事件才继续查
                if obj_dict['ActionType'] == 1:
                    obj_list = client.call("ak.wwise.core.object.get",
                                           waapi_h.waql_by_type("Sound", obj_dict['Target']['id']),
                                           options=config.options)['return']
                    sound_list, _, _ = waapi_h.find_obj(obj_list)
                    # pprint(sound_list)
                    if sound_list:
                        for sound_dict in sound_list:
                            if sound_dict['IsLoopingEnabled']:
                                is_loop = True
                                break

                    # 音乐循环查询
                    obj_list = client.call("ak.wwise.core.object.get",
                                           waapi_h.waql_by_type("MusicPlaylistContainer", obj_dict['Target']['id']),
                                           options=config.options)['return']
                    sound_list, _, _ = waapi_h.find_obj(obj_list)
                    # pprint(sound_list)
                    if sound_list:
                        for sound_dict in sound_list:
                            if sound_dict['id'] in loop_musicplaylist_container_id_list:
                                is_loop = True
                                # print(event_dict['name'])
                                break

    return is_loop


# 表的值设置
def set_value(event_dict):
    temp = excel_h.create_id(config.event_id_config, event_dict['name'], sheet)
    flag = 0
    for cell in list(sheet.columns)[2]:
        if event_dict['name'] == cell.value:
            flag = 1
            # 事件描述赋值
            sheet.cell(row=cell.row, column=config.desc_index).value = event_dict['notes']
            set_event_path(cell.row)
            # 获取是否循环并赋值
            sheet.cell(row=cell.row, column=config.isloop_index).value = get_is_loop()
            # 持有赋值
            sheet.cell(row=cell.row, column=config.HoldEventType_index).value = state_event_bank(event_dict)
            break
        elif (event_dict['notes'] == sheet.cell(cell.row, column=config.desc_index).value) and event_dict['notes'] \
                and sheet.cell(cell.row, column=config.desc_index).value:
            flag = 2
            # 事件名称赋值
            sheet.cell(row=cell.row, column=config.audioname_index).value = event_dict['name']
            set_event_path(cell.row)
            # 获取是否循环并赋值
            sheet.cell(row=cell.row, column=config.isloop_index).value = get_is_loop()
            # 持有赋值
            sheet.cell(row=cell.row, column=config.HoldEventType_index).value = state_event_bank(event_dict)
        # 是否循环赋值
    if flag == 0:
        # 插入为空的行
        insert_row = sheet.max_row + 1
        # pprint(insert_row)
        # rowname赋值
        sheet.cell(row=insert_row, column=config.rowname_index).value = temp
        # id编号赋值
        sheet.cell(row=insert_row, column=config.audioid_index).value = temp
        # 事件名称赋值
        sheet.cell(row=insert_row, column=config.audioname_index).value = event_dict['name']
        # 事件描述赋值
        sheet.cell(row=insert_row, column=config.desc_index).value = event_dict['notes']
        set_event_path(insert_row)
        # 获取是否循环并赋值
        sheet.cell(row=insert_row, column=config.isloop_index).value = get_is_loop()
        # 持有赋值
        sheet.cell(row=insert_row, column=config.HoldEventType_index).value = state_event_bank(event_dict)


"""复制到dcc的Audio表下"""


def copy_to_dcc_audio():
    # 获取xlsx的所有sheet的list
    sheet_names = wb_dcc.sheetnames
    # print(sheet_names)
    # 加载所有工作表
    # sheet = wb_dcc[sheet_name]

    # 定义目标表的映射
    sheet_mapping = {}
    for sheet_name in sheet_names:
        sheet_mapping[sheet_name] = wb_dcc[sheet_name]

    # """dcc audio表默认值填充"""
    # 加载所有工作表
    for sheet_name in sheet_names:
        sheet_dcc = wb_dcc[sheet_name]
        # 遍历每一行（跳过标题行）
        for row in sheet_dcc.iter_rows(min_row=2):  # 假设第一行是标题行
            # 该行内容不能为空
            if not all(cell.value is None for cell in row):
                # 获取第7列、第8列和第9列的单元格
                cell_col7 = row[config.FadeDuration_index - 1]  # 第7列
                cell_col8 = row[config.FadeCurveNum_index - 1]  # 第8列
                cell_col9 = row[config.ObjectType_index - 1]  # 第9列

                # 填充第7列的空值为1000
                if cell_col7.value is None:
                    cell_col7.value = 1000

                # 填充第8列的空值为0
                if cell_col8.value is None:
                    cell_col8.value = 8

                # 填充第9列的空值为0
                if cell_col9.value is None:
                    cell_col9.value = 0

    # """先从dcc复制手动配置项到audio表"""
    # 创建一个字典来存储目标表（表2）第一列的值及其对应行的第7至9列数据
    excel2_data = {}

    # 加载所有工作表
    for sheet_name in sheet_names:
        sheet_dcc = wb_dcc[sheet_name]
        # 遍历dcc的第一列，存储第7至9列的数据
        for row in sheet_dcc.iter_rows(min_row=2, max_row=sheet_dcc.max_row, max_col=config.ObjectType_index,
                                       values_only=True):
            key = row[0]  # 第一列的值作为键
            values = row[6:9]  # 获取第7到第9列的数据
            excel2_data[key] = values
        # pprint(excel2_data)

        # 遍历audio的第一列，若找到相等的值，则复制数据
        # 遍历 Excel1 的第一列，复制数据
        for row in sheet.iter_rows(min_row=2, max_col=1):
            key = row[0].value
            # print(key)
            if key in excel2_data:
                # 如果在 Excel2 中找到了匹配的键，复制数据
                for i, value in enumerate(excel2_data[key], start=config.FadeDuration_index):
                    sheet.cell(row=row[0].row, column=i, value=value)

    # 保存原工作簿
    wb.save(excel_path)

    """再从audio表获取数据复制到dcc"""

    # 清空目标工作表的内容
    # # 加载所有工作表
    for sheet_name in sheet_names:
        sheet_dcc = wb_dcc[sheet_name]
        sheet_dcc.delete_rows(2, sheet_dcc.max_row)

        # 获取最大行数
        # max_row = sheet_dcc.max_row
        # print(max_row)

    # 复制所有到dcc
    # 遍历audio_sheet的所有行
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始，假设第一行是标题
        first_column_value = row[2]  # 获取第3列，即Event名称的值
        if first_column_value:
            first_column_value = first_column_value.replace("AKE_", "")
            if "Set" not in first_column_value:
                first_column_value = first_column_value.replace("Play_", "")
                first_column_value = first_column_value.replace("Stop_", "")

            # 检查名称中是否包含目标字符串
            for keyword, target_sheet in sheet_mapping.items():
                if re_match_start_string(keyword, first_column_value):
                    target_sheet_max_row = target_sheet.max_row + 1
                    for col_index, value in enumerate(row, start=1):
                        target_sheet.cell(row=target_sheet_max_row, column=col_index, value=value)
                    break

    # 保存目标工作簿
    wb_dcc.save(config.excel_dcc_dt_audio_page_path)


"""从Wwise读Event_Info"""
with WaapiClient() as client:
    """通过获取事件引用的对象的时长进行标记是否为长音效"""


    def is_long_sound(event_dict):
        wwise_event_list = client.call("ak.wwise.core.object.get",
                                       waapi_h.waql_by_type("Action", event_dict['id'],
                                                            ),
                                       options=config.options)['return']

        if wwise_event_list:
            for obj_dict in wwise_event_list:
                if obj_dict['ActionType'] == 1:
                    sound_list = client.call("ak.wwise.core.object.get",
                                             waapi_h.waql_by_type("Sound", obj_dict['Target']['id'],
                                                                  ),
                                             options=config.options)['return']
                    if sound_list:

                        for sound_dict in sound_list:
                            if 'maxDurationSource' in sound_dict:
                                if 'trimmedDuration' in sound_dict['maxDurationSource']:
                                    if sound_dict['maxDurationSource']['trimmedDuration'] > config.long_sound_time:
                                        # pprint(event_dict)
                                        return 2
        return 0


    """Event持有值判定"""


    def state_event_bank(event_dict):
        # 通过获取事件的颜色：3进行标记1，即状态标记
        # if event_dict['Color'] == 3:
        #     return 1
        if "Set_State" in event_dict['name']:
            return 1
        else:
            return is_long_sound(event_dict)


    """判定音乐是否循环"""


    def find_music_lp_list():
        # PS：MusicPlaylistItem的根节点只有owner，
        # MusicPlaylistItem的其他子节点只有parent，parent为根节点
        obj_list = client.call("ak.wwise.core.object.get",
                               waapi_h.waql_from_type('MusicPlaylistItem'),
                               options=config.options)['return']
        musicplaylistitem_list, _, _ = waapi_h.find_obj(obj_list)
        # pprint(refer_list)

        # 使用一个列表保存为loop的musicplaylist container id
        loop_musicplaylist_container_id_list = []

        # 查找为loop的、非根节点的子节点
        for musicplaylistitem_dict in musicplaylistitem_list:
            if musicplaylistitem_dict['LoopCount'] == 0:
                # 获取非根子节点
                if (musicplaylistitem_dict['PlaylistItemType'] == 1):
                    # pprint(musicplaylistitem_dict)
                    # print()
                    # 获取子节点的parent，即根节点
                    root_id = musicplaylistitem_dict['parent']['id']
                    # print(root_id)
                    # 获取根节点的owner，即musicplaylist容器的id
                    obj_list = client.call("ak.wwise.core.object.get",
                                           waapi_h.waql_from_id(root_id),
                                           options=config.options)['return']
                    root_list, _, _ = waapi_h.find_obj(obj_list)

                # 获取根节点
                else:
                    # 获取根节点的owner，即musicplaylist容器的id
                    obj_list = client.call("ak.wwise.core.object.get",
                                           waapi_h.waql_from_id(musicplaylistitem_dict['id']),
                                           options=config.options)['return']
                    root_list, _, _ = waapi_h.find_obj(obj_list)
                    # pprint(root_list)
                for root_dict in root_list:
                    # 添加非重复元素，set为去重，然后再转为list
                    # 获取的是为loop的非根子节点的父级
                    if 'owner' in root_dict:
                        owner_id = root_dict['owner']['id']
                        loop_musicplaylist_container_id_list = list(
                            set(loop_musicplaylist_container_id_list + [owner_id]))

        return loop_musicplaylist_container_id_list


    """*************主程序*************"""

    # 获取为循环的音乐列表
    loop_musicplaylist_container_id_list = find_music_lp_list()
    # pprint(loop_musicplaylist_container_id_list)

    # 存放UE的Event路径
    ue_audio_path_list = []
    # 获取ue工程下的audio并转换路径
    ue_audio_dict, _ = oi_h.get_type_file_name_and_path('.uasset', ue_audio_path)
    for ue_audio in ue_audio_dict:
        if convert_ue_event_path(ue_audio_dict[ue_audio]):
            ue_audio_path_list.append(convert_ue_event_path(ue_audio_dict[ue_audio]))
    obj_list = client.call("ak.wwise.core.object.get",
                           waapi_h.waql_by_type("Event", event_root),
                           options=config.options)['return']
    event_list, event_id, _ = waapi_h.find_obj(obj_list)
    for event_dict in event_list:
        # 事件需要勾选才生成info信息
        if event_dict['Inclusion']:
            set_value(event_dict)

    # 同步删除Wwise中没有的事件
    # 创建Event Name列表
    extract_key = lambda d: d['name']
    event_name_list = list(map(extract_key, event_list))
    # pprint(event_name_list)
    # 获取AudioName列的内容
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == 'AudioName':
            # 返回该列的值（除了标题行）
            for cell in col[1:]:
                if cell.value not in event_name_list:
                    sheet.delete_rows(cell.row, 1)
                    oi_h.print_warning(cell.value + "：在Wwise Event中未被找到，该ID及数据已删除")

    wb.save(excel_path)

    # 将表内容同步到dcc下的Audio表中
    wb_dcc = openpyxl.load_workbook(config.excel_dcc_dt_audio_page_path)

    copy_to_dcc_audio()
    copy_to_dcc_audio()

    # excel转csv
    df = pd.read_excel(excel_path)
    # 指定CSV文件名和编码格式
    encoding = 'utf-8'
    df.to_csv(csv_path, encoding=encoding, index=False)

    # 应用程序弹窗
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo("Info", "ID表已更新，记得提交DCC下的Audio表")

import comlib.DCC表的引用路径生成打包版

# dt表导入UE
ue_csv_dt_audio_path = os.path.join(root_path, "ue_csv_dt_audio.py")
exe_h.run_unreal_editor_with_script(ue_csv_dt_audio_path)

os.system("pause")
