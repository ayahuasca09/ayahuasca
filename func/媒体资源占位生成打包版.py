import openpyxl
import os
from waapi import WaapiClient
from os.path import abspath, dirname
import sys
from pprint import pprint

import comlib.cloudfeishu_h as cloudfeishu_h
import comlib.oi_h as oi_h
import comlib.config as config
import comlib.waapi_h as waapi_h
import 命名规范检查打包版

is_pass_check_name = True

"""根目录获取"""


def get_py_path():
    py_path = ""
    if hasattr(sys, 'frozen'):
        py_path = dirname(sys.executable)
    elif __file__:
        py_path = dirname(abspath(__file__))
    return py_path


# 打包版
root_path = get_py_path()
# 通过_间隔分割的单词个数
word_list_len = config.word_list_len

# 飞书在线文档下载
media_excel_token_dict = config.media_sheet_token_dict
cloudfeishu_h.input_show(media_excel_token_dict, root_path)
excel_path = os.path.join(root_path, "Excel")
file_name_list = oi_h.find_all_files_by_type(excel_path, '.xlsx')
# print(file_name_list)
# print(1111)

"""收集unit的路径"""
audio_unit_list = []
event_unit_list = []
audio_mixer_list = []

"""*****************功能检测区******************"""
"""event结构找父级测试"""


# python检查 `list1` 中每个以“_”拆分后的字符串组成的列表的所有元素是否全部存在于 `list2` 中某个以“_”拆分后的字符串组成的列表中,
# 若存在，则继续判断选出“list1”中最长的字符串，与“list2”的字符串生成一字典
# 将list2改为dict2的键
def find_event_parent_match(dict1, string):
    # 将字符串拆分为列表
    string_list = string.split('_')

    # 过滤符合条件的键
    valid_keys = [
        key for key in dict1.keys()
        if all(part in string_list for part in key.split('_'))
    ]

    if valid_keys:
        # 找出最长的键
        longest_key = max(valid_keys, key=len)
        return longest_key
    else:
        return None


"""获取event_unit_list中缺失的unit"""


# 遍历 `list1`：对于 `list1` 中的每个元素，将其用下划线分隔，并转换为集合 `parts1`。
# 遍历 `list2`：对于 `list2` 中的每个元素，同样进行分隔并转换为集合 `parts2`。
# 检查子集关系：使用 `issubset` 方法检查 `parts1` 是否是 `parts2` 的子集。
# 更新结果列表：如果 `parts1` 是 `parts2` 的子集，则将 `list2` 中的该元素添加到 `result` 列表中。
# 输出结果：最终输出更新后的列表 `result`。
def get_missing_event_unit(list1, list2):
    for element1 in list1[:]:  # 使用 list1[:] 复制列表以避免修改时的迭代问题
        parts1 = set(element1.split('_'))
        world_list1 = element1.split('_')
        flag = 0
        # 至少要有两层_的才运算
        if len(world_list1) > 1:
            # 若为怪物，至少要三层_才运算
            if ("Mon_Mob" in element1) or ("Mon_Boss" in element1) or ("VO_Game" in element1):
                if len(world_list1) > 2:
                    flag = 1
            else:
                flag = 1
        if flag == 1:
            for element2 in list2:
                parts2 = set(element2.split('_'))
                if parts1.issubset(parts2):  # 检查是否有共同元素
                    if element2 not in list1:
                        list1.append(element2)
                    # break  # 如果已经匹配到一个，跳出内层循环


"""获取最长共同前缀"""


def longest_common_prefix(s1, s2):
    """Helper function to find the longest common prefix of two strings."""
    min_length = min(len(s1), len(s2))
    for i in range(min_length):
        if s1[i] != s2[i]:
            return s1[:i]
    return s1[:min_length]


"""在字典中查找最长共同前缀的值，且字典值长度需小于target"""


def find_longest_prefix_key(target, dictionary):
    max_prefix_length = 0
    best_key = None

    for key in dictionary.keys():
        if key in target:
            if len(key) < len(target):
                prefix = longest_common_prefix(target, key)
                if len(prefix) > max_prefix_length:
                    max_prefix_length = len(prefix)
                    best_key = key

    return best_key


"""获取表格中的事件描述和状态所在的列"""


def get_descrip_and_status_column():
    require_module_column = None
    second_module_column = None
    require_name_column = None
    status_column = None
    sample_name_column = None
    if list(sheet.rows):
        if list(sheet.rows)[0]:
            for cell in list(sheet.rows)[0]:
                if cell.value:
                    if 'Require Module' in str(cell.value):
                        require_module_column = cell.column
                        # print(require_module_column)
                    elif 'Second Module' in str(cell.value):
                        second_module_column = cell.column
                        # print(second_module_column)
                    elif 'Require Name' in str(cell.value):
                        require_name_column = cell.column
                        # print(require_name_column)
                    elif 'Status' in str(cell.value):
                        status_column = cell.column
                        # print(status_column)
                    elif 'Sample Name' in str(cell.value):
                        sample_name_column = cell.column
                        # print(status_column)

    return require_module_column, second_module_column, require_name_column, status_column, sample_name_column


with WaapiClient() as client:
    """*****************Wwise功能区******************"""
    """获取Wwise中的Unit列表"""


    def get_wwise_type_list(root_path, type):
        # 获取Wwise所有的Audio Unit
        wwise_audio_unit_all_list = client.call("ak.wwise.core.object.get",
                                                waapi_h.waql_by_type(type, root_path),
                                                options=config.options)['return']
        wwise_audio_unit_name_dict = {item['name']: item['id'] for item in wwise_audio_unit_all_list}
        # pprint(wwise_audio_unit_name_dict)
        # {'Amb': '{4BB49221-4DC1-4A05-A57C-3D42FA570675}',
        #  'Amb_A01': '{415AC456-D390-43D5-A09E-B5D30B253D76}',
        #  'CG': '{BBB332B4-C230-4D60-9ABA-91A36D968C11}'}
        return wwise_audio_unit_name_dict


    """obj创建"""


    def create_wwise_obj(type, parent, name, notes):
        unit_object = client.call("ak.wwise.core.object.create",
                                  waapi_h.args_object_create(parent,
                                                             type, name, notes))
        oi_h.print_warning(name + "：" + type + "已创建")
        return unit_object


    """eventunit结构创建"""


    def create_wwise_event_unit(event_unit_list):
        unit_object = {}
        # 查找Event中是否有该unit（第一层级）
        for audio_unit in event_unit_list:
            wwise_audio_unit_name_dict = get_wwise_type_list(config.wwise_event_path, "WorkUnit")
            # 若Wwise中无相应Unit则需要创建
            if audio_unit not in wwise_audio_unit_name_dict:
                event_unit_parent = find_event_parent_match(wwise_audio_unit_name_dict, audio_unit)
                if event_unit_parent:
                    # print(audio_unit + ":" + event_unit_parent)
                    # WorkUnit创建
                    unit_object = create_wwise_obj("WorkUnit",
                                                   wwise_audio_unit_name_dict[event_unit_parent], audio_unit, "")
                # 若未找到父级则在根目录创建
                else:
                    unit_object = create_wwise_obj("WorkUnit",
                                                   config.wwise_event_path, audio_unit, "")


    """audiounit结构创建"""


    # 有2种方式
    # 1：audio下，unit和actormixer都需要创建
    # 2：audio下，只需要创建actormixer
    def create_wwise_audio_unit(unit_list, flag):
        # 查找Wwise中是否有该unit
        for audio_unit in unit_list:
            wwise_audio_unit_name_dict = get_wwise_type_list(config.wwise_sfx_path, "ActorMixer")
            # 若Wwise中无相应Unit则需要创建
            if audio_unit not in wwise_audio_unit_name_dict:
                # 查找该unit需要放的父级是谁
                audio_unit_parent = find_longest_prefix_key(audio_unit, wwise_audio_unit_name_dict)
                # 根目录创建
                if not audio_unit_parent:
                    # print(audio_unit)
                    audio_unit_parent_obj = create_wwise_obj("WorkUnit",
                                                             config.wwise_sfx_path, audio_unit, "")
                    audio_unit_parent = audio_unit_parent_obj['id']
                else:
                    # 2：audio下，只需要创建actormixer
                    if flag == 2:
                        # ActorMixer创建
                        unit_object = create_wwise_obj("ActorMixer",
                                                       wwise_audio_unit_name_dict[audio_unit_parent], audio_unit, "")
                    elif flag == 1:
                        # WorkUnit创建
                        unit_object = create_wwise_obj("WorkUnit",
                                                       wwise_audio_unit_name_dict[audio_unit_parent], audio_unit, "")
                        # 1：audio下，unit和actormixer都需要创建
                        if unit_object:
                            # ActorMixer创建
                            unit_object = create_wwise_obj("ActorMixer",
                                                           unit_object['id'], audio_unit, "")


    """*****************主程序处理******************"""

    # 获取Wwise所有的Event Unit
    wwise_event_unit_all_list = client.call("ak.wwise.core.object.get",
                                            waapi_h.waql_by_type("WorkUnit", config.wwise_event_path),
                                            options=config.options)['return']
    wwise_event_unit_name_dict = {item['name']: item['id'] for item in wwise_event_unit_all_list}
    # pprint(wwise_event_unit_name_dict)

    # 记录所有资源名称
    sound_name_list = []

    # 提取规则：只提取xlsx文件
    for i in file_name_list:
        if ".xlsx" in i:
            # 拼接xlsx的路径
            file_path_xlsx = os.path.join(excel_path, i)
            # 获取xlsx的workbook
            wb = openpyxl.load_workbook(file_path_xlsx)
            # 获取xlsx的所有sheet
            sheet_names = wb.sheetnames
            # pprint(sheet_names)
            # pprint(1111)
            # 加载所有工作表
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                require_module_column, second_module_column, require_name_column, status_column, sample_name_column = get_descrip_and_status_column()
                if sample_name_column:
                    for cell_sound in list(sheet.columns)[sample_name_column - 1]:
                        if status_column:
                            if sheet.cell(row=cell_sound.row,
                                          column=status_column).value in config.status_list:
                                # pprint(cell_sound.value)
                                # 通过命名规范检查
                                if 命名规范检查打包版.check_basic(cell_sound.value, audio_unit_list,
                                                                  event_unit_list,
                                                                  audio_mixer_list, word_list_len):
                                    # 通过命名规范检查
                                    pass
                                # 命名规范只要一次不通过就算失败
                                else:
                                    is_pass_check_name = False

    """*****************work unit创建******************"""
    # 对列表进行排序
    audio_unit_list.sort(key=len)
    event_unit_list.sort(key=len)
    audio_mixer_list.sort(key=len)

    # 获取event_unit_list中缺失的unit
    get_missing_event_unit(event_unit_list, audio_unit_list)

    print("*************Audio的Unit和ActorMixer创建****************")
    # # 1：audio下，unit和actormixer都需要创建
    create_wwise_audio_unit(audio_unit_list, 1)
    print()
    print()
    print("*************Event的Unit创建****************")
    # # event下，unit创建
    create_wwise_event_unit(event_unit_list)
    print()
    print()
    print("*************ActorMixer创建****************")
    # 2：audio下，只需要创建actormixer
    create_wwise_audio_unit(audio_mixer_list, 2)
    print()
    print()

    pprint("audio_unit_list：")
    pprint(audio_unit_list)
    print()

    pprint("event_unit_list：")
    pprint(event_unit_list)
    print()

    pprint("audio_mixer_list：")
    pprint(audio_mixer_list)

if is_pass_check_name:
    print("--------------------------")
    print("资源命名检查通过，请输入以下数字选项继续：")
    print("1：生成占位资源")
    print("0：结束")
    temp = input("请输入数字:")
    if temp == 1 or temp == "1":
        import 媒体资源从表导入打包版
os.system("pause")
