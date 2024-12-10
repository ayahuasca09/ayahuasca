import openpyxl
import json
import sys
from os.path import abspath, dirname
import os
from pprint import pprint
import re
from waapi import WaapiClient
import shutil
from openpyxl.cell import MergedCell
import module.cloudfeishu.cloudfeishu_h as cloudfeishu_h
import module.oi.oi_h as oi_h
import config
import module.excel.excel_h as excel_h
import module.waapi.waapi_h as waapi_h
from pathlib import Path

"""根目录获取"""
# 获取当前脚本的文件名及路径
script_name_with_extension = __file__.split('/')[-1]
# 去掉扩展名
script_name = script_name_with_extension.rsplit('.', 1)[0]
# 替换为files
root_path = script_name.replace("func", "files")
# print(f"当前脚本的名字是: {root_path}")c

file_name_list = excel_h.excel_get_path_list(root_path)
# pprint(file_name_list)

"""在线表获取"""
# 规范检查表
# for excel_media_token in config.mus_sheet_token_list:
#     _, excel_name = cloudfeishu_h.get_excel_token(excel_media_token)
#     # pprint(excel_name)
#     for file_name in file_name_list:
#         if excel_name in file_name:
#             print(excel_name)
#             cloudfeishu_h.download_cloud_sheet(excel_media_token, os.path.join(root_path, file_name))
#             break

"""功能数据初始化"""
# switch创建列表
switch_create_list = []
# unit创建列表
unit_create_list = []
# playlist创建列表
playlist_create_list = []

"""*****************功能检测区******************"""

with WaapiClient() as client:
    """*****************Wwise功能区******************"""
    """obj创建"""


    def create_wwise_obj(type, parent, name, notes):
        unit_object = None
        if name != "Mus":
            unit_object = client.call("ak.wwise.core.object.create",
                                      waapi_h.args_object_create(parent,
                                                                 type, name, notes))
            oi_h.print_warning(name + "：" + type + "已创建")

        return unit_object


    """获取Wwise中的Unit列表"""


    def get_wwise_type_list(root_path, type):
        # 获取Wwise所有的Audio Unit
        wwise_bus_name_all_list = client.call("ak.wwise.core.object.get",
                                              waapi_h.waql_by_type(type, root_path),
                                              options=config.options)['return']
        wwise_bus_name_name_dict = {item['name']: item['id'] for item in wwise_bus_name_all_list}
        # pprint(wwise_bus_name_name_dict)
        # {'Amb': '{4BB49221-4DC1-4A05-A57C-3D42FA570675}',
        #  'Amb_A01': '{415AC456-D390-43D5-A09E-B5D30B253D76}',
        #  'CG': '{BBB332B4-C230-4D60-9ABA-91A36D968C11}'}
        return wwise_bus_name_name_dict


    """mus创建"""


    def create_wwise_mus(unit_list, playlist_list, switch_list):

        # unit创建
        for mus_name in unit_list:
            # 查找已有unit列表
            unit_have_dict = get_wwise_type_list(config.wwise_mus_state_path, "WorkUnit")
            playlist_have_dict = get_wwise_type_list(config.wwise_mus_state_path, "MusicPlaylistContainer")
            switch_have_dict = get_wwise_type_list(config.wwise_mus_state_path, "MusicSwitchContainer")
            # pprint(unit_have_dict)
            # Wwise里没有unit的的
            if mus_name not in unit_have_dict:
                # 查找该unit需要放的父级是谁
                unit_parent = oi_h.find_longest_prefix_key(mus_name, unit_have_dict)
                if not unit_parent:
                    # WorkUnit创建
                    unit_object = create_wwise_obj("WorkUnit",
                                                   config.wwise_mus_global_path, mus_name,
                                                   "")
                else:
                    # t5apprint(playlist_list)
                    unit_have_dict = get_wwise_type_list(config.wwise_mus_state_path, "WorkUnit")
                    unit_parent = oi_h.find_longest_prefix_key(mus_name, unit_have_dict)
                    if mus_name in switch_list:
                        if mus_name not in switch_have_dict:
                            # Switch父级创建
                            switch_object = create_wwise_obj("MusicSwitchContainer",
                                                             unit_have_dict[unit_parent], mus_name,
                                                             "")
                            # WorkUnit创建
                            unit_object = create_wwise_obj("WorkUnit",
                                                           switch_object['id'], mus_name,
                                                           "")
                        else:
                            unit_object = create_wwise_obj("WorkUnit",
                                                           switch_have_dict[mus_name], mus_name,
                                                           "")

                    elif mus_name in playlist_list:
                        if mus_name not in playlist_have_dict:
                            # WorkUnit创建
                            unit_object = create_wwise_obj("WorkUnit",
                                                           unit_have_dict[unit_parent], mus_name,
                                                           "")

                            # Switch父级创建
                            playlist_object = create_wwise_obj("MusicPlaylistContainer",
                                                               unit_object['id'], mus_name,
                                                               "")

        # switch创建
        for mus_name in switch_list:
            unit_have_dict = get_wwise_type_list(config.wwise_mus_state_path, "WorkUnit")
            switch_have_dict = get_wwise_type_list(config.wwise_mus_state_path, "MusicSwitchContainer")

            if mus_name not in switch_have_dict:
                # 找到的父级为unit
                unit_parent_1 = oi_h.find_longest_prefix_key(mus_name, unit_have_dict)
                unit_parent_2 = oi_h.find_longest_prefix_key(mus_name, switch_have_dict)
                parent_id = None
                if unit_parent_1:
                    if unit_parent_2:
                        if len(unit_parent_2) > len(unit_parent_1):
                            parent_id = switch_have_dict[unit_parent_2]
                        else:
                            parent_id = unit_have_dict[unit_parent_1]
                    else:
                        parent_id = unit_have_dict[unit_parent_1]
                elif unit_parent_2:
                    parent_id = switch_have_dict[unit_parent_2]

                if parent_id:
                    switch_object = create_wwise_obj("MusicSwitchContainer",
                                                     parent_id, mus_name,
                                                     "")
        # playlist创建
        for mus_name in playlist_list:
            unit_have_dict = get_wwise_type_list(config.wwise_mus_state_path, "WorkUnit")
            switch_have_dict = get_wwise_type_list(config.wwise_mus_state_path, "MusicSwitchContainer")
            playlist_have_dict = get_wwise_type_list(config.wwise_mus_state_path, "MusicPlaylistContainer")
            if mus_name not in playlist_have_dict:
                # 找到的父级为unit
                unit_parent_1 = oi_h.find_longest_prefix_key(mus_name, unit_have_dict)
                unit_parent_2 = oi_h.find_longest_prefix_key(mus_name, switch_have_dict)
                parent_id = None
                if unit_parent_1:
                    if unit_parent_2:
                        if len(unit_parent_2) > len(unit_parent_1):
                            parent_id = switch_have_dict[unit_parent_2]
                        else:
                            parent_id = unit_have_dict[unit_parent_1]
                    else:
                        parent_id = unit_have_dict[unit_parent_1]
                elif unit_parent_2:
                    parent_id = switch_have_dict[unit_parent_2]

                if parent_id:
                    playlist_object = create_wwise_obj("MusicPlaylistContainer",
                                                       parent_id, mus_name,
                                                       "")


    """获取需要创建的mus列表"""


    def get_create_mus_name_list():
        # 遍历每一行
        for row in mus_config_sheet.iter_rows():
            mus_name = None
            # 遍历每一列
            for cell in row:
                cell_value, is_merge = excel_h.check_is_mergecell(cell, mus_config_sheet)
                if cell_value:
                    is_unit = False
                    # print(cell_value)
                    if not mus_name:
                        mus_name = cell_value
                        is_unit = True
                    else:
                        mus_name += "_" + cell_value
                        fill = cell.fill
                        if fill.fill_type == 'solid':
                            if fill.start_color.index == "FFD9F3FD":
                                is_unit = True
                    is_switch = False

                    if is_unit:
                        if mus_name not in unit_create_list:
                            unit_create_list.append(mus_name)
                    # 是否为switch容器检查
                    if is_merge:
                        is_switch = True
                    else:
                        if mus_config_sheet.cell(row=cell.row,
                                                 column=cell.column + 1).value:
                            is_switch = True
                    if is_switch:
                        if mus_name not in switch_create_list:
                            switch_create_list.append(mus_name)
                    else:
                        if mus_name not in playlist_create_list:
                            playlist_create_list.append(mus_name)

            # print()


    """*****************主程序处理******************"""
    """音乐结构创建"""

    # 拼接xlsx的路径
    file_path_xlsx = os.path.join(root_path, config.excel_mus_config_path)
    # 获取xlsx的workbook
    wb = openpyxl.load_workbook(file_path_xlsx)
    mus_config_sheet = wb['Mus']
    get_create_mus_name_list()
    switch_create_list.sort(key=len)
    playlist_create_list.sort(key=len)
    unit_create_list.sort(key=len)
    # pprint(switch_create_list)
    # pprint(playlist_create_list)
    # pprint(unit_create_list)
    create_wwise_mus(unit_create_list, playlist_create_list, switch_create_list)
