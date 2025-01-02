import os
import openpyxl
from waapi import WaapiClient

import sys

import comlib.cloudfeishu_h as cloudfeishu_h
import comlib.excel_h as excel_h
import comlib.config as config
import comlib.exe_h as exe_h
from os.path import abspath, dirname


def get_py_path():
    py_path = ""
    if hasattr(sys, 'frozen'):
        py_path = dirname(sys.executable)
    elif __file__:
        py_path = dirname(abspath(__file__))
    return py_path


# 打包版
root_path = get_py_path()

"""文件获取"""
# State表
excel_state_path = os.path.join(root_path, config.excel_state_path)
# print(wb)

# 规范检查表
state_token = config.state_sheet_token
cloudfeishu_h.download_cloud_sheet(state_token, excel_state_path)

"""状态类通用后缀列表"""
list_state_suffix = config.list_state_suffix

"""事件描述查重列表"""
event_desc_list = []

"""创建内容的根目录"""
state_root_path = '{ADF05AAB-970F-46B3-A5BE-471221EDBD17}'
switch_root_path = '{E1F6B98B-D124-426E-B031-B2BDC5089C31}'
trigger_root_path = '{94BBA60A-29EF-4328-83E8-075770BF2687}'

set_state_event_path = '{4466FD27-21D5-44BB-A0F1-AF801D870945}'
set_switch_event_path = '{F5985FEE-CE20-4C2F-9B3B-46A1EB8AB612}'
set_event_path = '{80A74274-275D-4554-AE98-EE0112489C5C}'
set_trigger_event_path = '{EDD34B59-3938-4C22-A63D-838FD8DCD57E}'

# Event的WorkUnit路径
event_workunit_path = config.wwise_event_workunit_path

"""状态名称列表"""
state_name_list = []
state_group_name_list = []

"""*****************功能检测区******************"""

"""检查是否使用通用词汇"""


def check_by_com_suffix(suffix_name):
    if suffix_name in list_state_suffix:
        return
    else:
        print_error(cell_sound.value + "后缀错误或未添加至通用后缀列表")


"""报错捕获"""


def print_error(error_info):
    global is_pass
    is_pass = False
    print("[error]" + error_info)


"""检测字符串是否含有中文"""


def check_is_chinese(string):
    for ch in string:
        if u'\u4e00' <= ch <= u'\u9fff':
            return True
    return False


"""字符串长度检查"""


def check_by_str_length(str1, length):
    if len(str1) > length:
        if "_" in str1:
            print_error(cell_sound.value + "：描述后缀名过长，限制字符数为" + str(length))
        else:
            print_error(cell_sound.value + "：通过_切分的某个单词过长，每个单词长度不应超过10个字符，需要再拆分")


"""获取表格中的事件描述和状态所在的列"""


def get_descrip_and_status_column():
    require_module_column = None
    second_module_column = None
    require_name_column = None
    status_column = None
    is_bank_column = None
    if list(sheet.rows)[0]:
        for cell in list(sheet.rows)[0]:
            if cell.value:
                if 'Group名称' == str(cell.value):
                    require_module_column = cell.column
                    # print(require_module_column)
                elif 'Group描述' == str(cell.value):
                    second_module_column = cell.column
                    # print(second_module_column)
                elif '值' == str(cell.value):
                    require_name_column = cell.column
                    # print(require_name_column)
                elif '值描述' == str(cell.value):
                    status_column = cell.column
                    # print(status_column)
                elif '是否创建Bank' == str(cell.value):
                    is_bank_column = cell.column
                    # print(status_column)

    return require_module_column, second_module_column, require_name_column, status_column, is_bank_column


"""分隔符的大小写和长度检查"""


def check_by_length_and_word(name):
    # pprint(cell_sound.value)
    word_list = name.split("_")
    is_title = True
    for word in word_list:
        """判定_的每个都不能超过10个"""
        check_by_str_length(word, 10)
        """判定_的每个开头必须大写"""
        if len(word) > 0:
            # 只需要检查第一个字符，因为istitle会导致检查字符串只能首字母大写
            if word[0].istitle() == False and word[0].isdigit() == False:
                is_title = False
                break
    if is_title == False:
        print_error(name + "：通过”_“分隔的每个单词开头都需要大写")
    if word_list:
        return word_list[-1]
    else:
        return None


with WaapiClient() as client:
    """*****************Wwise功能区******************"""
    """设置对象属性"""


    def set_obj_property(obj_id, property, value):
        args = {
            "object": obj_id,
            "property": property,
            "value": value
        }
        client.call("ak.wwise.core.object.setProperty", args)
        # print(name_2 + ":" + str(trigger_rate))


    """设置对象引用"""


    def set_obj_refer(obj_id, refer, value):
        args = {
            "object": obj_id,
            "reference": refer,
            "value": value
        }
        client.call("ak.wwise.core.object.setReference", args)


    """警告捕获"""


    def print_warning(warn_info):
        print("[warning]" + warn_info)


    """查找对象"""


    def find_obj(args):
        options = {
            'return': ['name', 'id', 'path', 'notes', 'originalWavFilePath', 'type', 'parent']

        }
        obj_sub_list = client.call("ak.wwise.core.object.get", args, options=options)['return']
        if not obj_sub_list:
            obj_sub_id = ""
            obj_sub_path = ""
        else:
            obj_sub_id = obj_sub_list[0]['id']
            obj_sub_path = obj_sub_list[0]['path']
        return obj_sub_list, obj_sub_id, obj_sub_path


    """设置obj的notes"""


    def set_obj_notes(obj_id, notes_value):
        args = {
            'object': obj_id,
            'value': notes_value
        }
        client.call("ak.wwise.core.object.setNotes", args)
        # print_warning(obj_name + "描述更改为：" + notes_value)


    """修改Wwise内容的名字"""


    def change_name_by_wwise_content(obj_id, name, old_name, obj_type):
        args = {
            "objects": [
                {

                    "object": obj_id,
                    "name": name,
                }
            ]
        }
        client.call("ak.wwise.core.object.set", args)
        print_warning(old_name + "(" + obj_type + ")改名为：" + name)


    """通用内容创建"""


    def create_obj_content(state_group_parent_path, state_group_type,
                           state_group_name, state_group_desc, state_id):
        flag = 0
        state_group_id = ""
        # 查找此state_group是否存在
        state_group_list, _, _ = find_obj(
            {'waql': ' "%s" select descendants where type = "%s"' % (
                state_group_parent_path, state_group_type)})
        # pprint(state_group_list)
        # state_group已存在
        for state_group_dict in state_group_list:
            if state_group_dict['name'] == state_group_name:
                flag = 1
                # 设置notes
                set_obj_notes(state_group_dict['id'], state_group_desc)
                state_group_id = state_group_dict['id']
                break
            # 改名了但描述一致
            elif state_group_dict['notes'] == state_group_desc:
                flag = 2
                # 改名
                change_name_by_wwise_content(state_group_dict['id'], state_group_name, state_group_dict['name'],
                                             state_group_type)
                state_group_id = state_group_dict['id']
                break
        # state_group不存在，需要创建
        # 不存在则创建
        if flag == 0:
            if state_group_type == "Event":
                action_type = None
                if sheet_name == "State":
                    action_type = 22
                elif sheet_name == "Switch":
                    action_type = 23
                elif sheet_name == "Trigger":
                    action_type = 35
                args = {
                    # 选择父级
                    "parent": state_group_parent_path,
                    # 创建类型名称
                    "type": state_group_type,
                    "name": state_group_name,
                    "notes": state_group_desc,
                    "children": [
                        {
                            "name": "",
                            "type": "Action",
                            "@ActionType": action_type,
                            "@Target": state_id
                        }]
                }
            else:
                # 创建的rnd属性
                args = {
                    # 选择父级
                    "parent": state_group_parent_path,
                    # 创建类型名称
                    "type": state_group_type,
                    "name": state_group_name,
                    "notes": state_group_desc,
                }
            state_group_object = client.call("ak.wwise.core.object.create", args)
            # print(state_group_object)
            state_group_id = state_group_object['id']
            print_warning(state_group_object['name'] + ":" + state_group_type + "已创建")
        return state_group_id


    """创建新的state/switch"""


    def create_state_content(state_group_name, state_group_desc, state_name, state_desc, is_bank_value):
        # 查找该state_group_name是否已存在
        state_group_parent_path = ''
        state_group_type = ''
        event_path = ''
        if sheet_name == "State":
            state_group_parent_path = state_root_path
            state_group_type = "StateGroup"
            event_path = set_state_event_path
        elif sheet_name == "Switch":
            state_group_parent_path = switch_root_path
            state_group_type = "SwitchGroup"
            event_path = set_switch_event_path

        """state group创建"""
        if state_group_parent_path and state_group_type:
            state_group_id = create_obj_content(state_group_parent_path, state_group_type, state_group_name,
                                                state_group_desc, "")
            """state 创建"""
            state_id = create_obj_content(state_group_id, sheet_name, state_name,
                                          state_desc, "")
            if is_bank_value == 1:
                set_obj_property(state_id, "OverrideColor", True)
                set_obj_property(state_id, "Color", 3)
            else:
                set_obj_property(state_id, "OverrideColor", False)

            """event unit创建"""
            event_parent_id = create_obj_content(event_path, "WorkUnit", state_group_name,
                                                 state_group_desc, "")

            """event 创建"""
            # event命名
            event_state_name = state_group_name.replace(suffix_name, state_name)
            event_name = 'AKE_Set_' + sheet_name + "_" + event_state_name
            event_id = create_obj_content(event_parent_id, "Event", event_name,
                                          state_desc, state_id)
            if is_bank_value == 1:
                set_obj_property(event_id, "OverrideColor", True)
                set_obj_property(event_id, "Color", 3)
            else:
                set_obj_property(event_id, "OverrideColor", False)


    """*****************主程序处理******************"""
    # 撤销开始
    client.call("ak.wwise.core.undo.beginGroup")

    # 记录所有资源名称
    sound_name_list = []

    wb = openpyxl.load_workbook(excel_state_path)
    # 获取xlsx的所有sheet
    sheet_names = wb.sheetnames
    # 加载所有工作表
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        group_name, group_desc, value, value_desc, is_bank_column = get_descrip_and_status_column()
        # 获取工作表第一行数据
        for cell in list(sheet.rows)[0]:
            if '值' == str(cell.value):
                # 获取音效名下的内容
                for cell_sound in list(sheet.columns)[cell.column - 1]:
                    # 空格和中文不检测
                    if cell_sound.value:
                        if not check_is_chinese(cell_sound.value):

                            """❤❤❤❤数据获取❤❤❤❤"""
                            """检测表格内容"""
                            is_pass = True

                            """检查表格中是否有内容重复项"""
                            if cell_sound.value in sound_name_list:
                                print_error(cell_sound.value + "：表格中有重复项Group名称 ，请检查")
                            else:
                                sound_name_list.append(cell_sound.value)

                                """❤❤❤state内容检查❤❤❤"""
                                """state/Trigger名称"""
                                value = cell_sound.value
                                state_name_list.append(value)
                                check_by_length_and_word(value)

                                """state/Trigger描述"""
                                value_desc_value = sheet.cell(row=cell_sound.row,
                                                              column=value_desc).value
                                if value_desc_value:
                                    if group_name:
                                        """❤❤❤state group内容检查❤❤❤"""
                                        # 检测是否为合并单元格，是则读取合并单元格的内容
                                        """state group名称"""
                                        group_value, _ = excel_h.check_is_mergecell(
                                            sheet.cell(row=cell_sound.row, column=group_name), sheet)
                                        state_group_name_list = list(set(state_group_name_list + [group_value]))
                                        if group_value:
                                            # pprint(group_value)
                                            # 分隔符的大小写和长度检查，取后缀名
                                            suffix_name = check_by_length_and_word(group_value)
                                            # print(suffix_name)

                                            # 通用后缀检查
                                            check_by_com_suffix(suffix_name)

                                            """❤❤❤描述项检查❤❤❤"""
                                            """state group描述"""
                                            group_desc_value, _ = excel_h.check_is_mergecell(
                                                sheet.cell(row=cell_sound.row, column=group_desc), sheet)
                                            if group_desc_value:
                                                # print(value_desc_value)
                                                # 拼接Group和值的描述
                                                """state描述"""
                                                state_desc_value = group_desc_value + "_" + value_desc_value
                                                # print(state_desc_value)
                                                # state描述查重
                                                if state_desc_value not in event_desc_list:
                                                    event_desc_list.append(state_desc_value)
                                                else:
                                                    print_error(state_desc_value + "：表格中有重复项描述，请检查")

                                                """是否创建Bank读取"""
                                                is_bank_value, _ = excel_h.check_is_mergecell(
                                                    sheet.cell(row=cell_sound.row, column=is_bank_column), sheet)



                                            else:
                                                print_error(value + ":没有状态描述，请补充！")
                                        else:
                                            print_error(group_value + ":没有状态描述，请补充！")

                                        # # 生成Wwise内容
                                        if is_pass:
                                            create_state_content(group_value, group_desc_value, value,
                                                                 state_desc_value, is_bank_value)
                                else:
                                    print_error(value + ":没有值描述，请补充")
    print()
    """******************对象删除清理********************"""
    """对象删除"""


    def delete_obj(obj_id, obj_name, obj_type):
        args = {
            "object": "%s" % obj_id
        }
        client.call("ak.wwise.core.object.delete", args)
        print_warning(obj_name + "(" + obj_type + ")" + ":已删除")


    """对象列表查找"""


    def find_obj_list(obj_path, obj_type):
        obj_list, _, _ = find_obj(
            {'waql': ' "%s" select descendants where type = "%s" ' % (obj_path, obj_type)})
        return obj_list


    """删除状态引用的事件"""


    def delete_state_refer_event(state_id):
        args = {
            'waql': '"%s" select referencesTo' % state_id
        }
        refer_list, _, _ = find_obj(args)
        # print(refer_list)
        if refer_list:
            # pprint(refer_list)
            for refer_dict in refer_list:
                if 'parent' in refer_dict:
                    if 'AKE_Set_' in refer_dict['parent']['name']:
                        # print(refer_dict['parent'])
                        delete_obj(refer_dict['parent']['id'], refer_dict['parent']['name'], "Event")


    """删除状态"""


    def delete_state(wwise_obj_list, excel_list, obj_type):
        for wwise_obj_dict in wwise_obj_list:
            # pprint(wwise_obj_dict['name'])
            if wwise_obj_dict['name'] != "None":
                if wwise_obj_dict['name'] not in excel_list:
                    # print(wwise_obj_dict['name'])
                    # State引用的事件删除
                    delete_state_refer_event(wwise_obj_dict['id'])
                    # State/StateGroup删除
                    delete_obj(wwise_obj_dict['id'], wwise_obj_dict['name'], obj_type)
                    # StateGroupUnit删除
                    """WorkUnit在文件夹中删除"""
                    if "Group" in obj_type:
                        group_event_workunit_path = os.path.join(event_workunit_path, wwise_obj_dict['name'] + ".wwu")
                        os.remove(group_event_workunit_path)
                        print(group_event_workunit_path)


    """同步表中删除的内容"""
    # Wwise中的内容列表获取
    state_list = find_obj_list(state_root_path, "State")
    state_group_list = find_obj_list(state_root_path, "StateGroup")
    switch_list = find_obj_list(switch_root_path, "Switch")
    switch_group_list = find_obj_list(switch_root_path, "SwitchGroup")
    event_list = find_obj_list(set_event_path, "Event")
    trigger_list = find_obj_list(trigger_root_path, "Trigger")

    # 查找state/switch，再跟表格比对有没有，没有就删除资源及事件引用
    delete_state(state_list, state_name_list, "State")
    delete_state(switch_list, state_name_list, "Switch")

    delete_state(state_group_list, state_group_name_list, "StateGroup")
    delete_state(switch_group_list, state_group_name_list, "SwitchGroup")

    delete_state(trigger_list, state_name_list, "Trigger")

    # 撤销结束
    client.call("ak.wwise.core.undo.endGroup", displayName="rnd创建撤销")

# gen soundbank
exe_h.gen_soundbank()
# ue reconcile
exe_h.run_ue_reconcile()
# id表生成
auto_sound_path = config.auto_sound_path
gen_dt_id_path = os.path.join(auto_sound_path, 'ID表生成', 'ID表生成.exe')
exe_h.run_exe(gen_dt_id_path)

os.system("pause")
