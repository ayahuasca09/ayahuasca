import openpyxl
import os
from os.path import abspath, dirname
import sys
import re

import comlib.cloudfeishu_h as cloudfeishu_h
import comlib.oi_h as oi_h
import comlib.config as config

is_pass_check_name = True

"""根目录获取"""


def get_py_path():
    py_path = ""
    if hasattr(sys, 'frozen'):
        py_path = dirname(sys.executable)
    elif __file__:
        py_path = dirname(abspath(__file__))
    return py_path


# 打包版
root_path = get_py_path()
# 通过_间隔分割的单词个数
word_list_len = config.word_list_len

# 飞书在线文档下载
media_excel_token_dict = config.media_sheet_token_dict
cloudfeishu_h.input_show(media_excel_token_dict, root_path)
excel_path = os.path.join(root_path, "Excel")

# 获取媒体资源文件列表
wav_path = os.path.join(root_path, "New_Media")
_, file_wav_list = oi_h.get_type_file_name_and_path('.wav', wav_path)

# excel表的名称和资源路径
excel_name_and_path_dict, _ = oi_h.get_type_file_name_and_path('.xlsx', excel_path)

"""功能函数"""


def get_descrip_and_status_column():
    sample_name_column = None
    old_name_column = None

    if list(sheet.rows):
        if list(sheet.rows)[0]:
            for cell in list(sheet.rows)[0]:
                if cell.value:
                    if 'Old Name' in str(cell.value):
                        old_name_column = cell.column
                        # print(require_module_column)
                    elif 'Sample Name' in str(cell.value):
                        sample_name_column = cell.column
                        # print(status_column)

    return old_name_column, sample_name_column


"""*****************主程序处理******************"""

# 读取媒体资源文件
for file_wav_dict in file_wav_list:
    for wav_name in file_wav_dict:
        # 去除wav的名字
        no_wav_name = re.sub(".wav", "", wav_name)
        # 提取规则：只提取xlsx文件
        for excel_name_and_path in excel_name_and_path_dict:
            # 获取xlsx的workbook
            wb = openpyxl.load_workbook(excel_name_and_path_dict[excel_name_and_path])
            # 获取xlsx的所有sheet
            sheet_names = wb.sheetnames
            # 加载所有工作表
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                word_list_len = sheet.max_column

                old_name_column, sample_name_column = get_descrip_and_status_column()
                flag = 0
                if old_name_column and sample_name_column:
                    for cell_sound in list(sheet.columns)[old_name_column - 1]:
                        # 如果找到和文件名相等的旧名字
                        if no_wav_name == cell_sound.value:
                            # 获取新名字列
                            new_name = sheet.cell(row=cell_sound.row,
                                                  column=sample_name_column).value
                            old_path = file_wav_dict[wav_name]
                            new_path = old_path.replace(no_wav_name, new_name)
                            if new_name:
                                os.rename(old_path, new_path)
                                oi_h.print_warning(no_wav_name + '：已改名为' + new_name)

                            else:
                                oi_h.print_error(no_wav_name + '：在' + sheet_name + '表中的音效名为空，请先命名')


                else:
                    oi_h.print_error(sheet_name + '：没有音效名列或旧名字列，请添加')

os.system("pause")
