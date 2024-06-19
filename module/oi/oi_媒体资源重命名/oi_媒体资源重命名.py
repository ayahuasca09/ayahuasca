import openpyxl
import sys
from os.path import abspath, dirname
import os
from module.oi.oi_h import get_type_file_name_and_path
from pprint import pprint
import shutil
from waapi import WaapiClient

# 文件所在目录
py_path = ""
if hasattr(sys, 'frozen'):
    py_path = dirname(sys.executable)
elif __file__:
    py_path = dirname(abspath(__file__))

# wav文件目录
wav_path = r"S:\chen.gong_DCC_Audio\Audio\SilverPalace_WwiseProject\Originals\SFX"

# 容器根目录
waapi_path = "\\Actor-Mixer Hierarchy\\v1"

"""获取.xlsx文件"""

file_names = []
for i in os.walk(py_path):
    file_names.append(i)
file_name_list = file_names[0][2]
# pprint(file_name_list)

"""获取.wav文件"""
file_wav_dict = get_type_file_name_and_path('.wav', wav_path)
# pprint(file_wav_dict)

"""*****************功能检测区******************"""

"""检测字符串是否含有中文"""


def check_is_chinese(string):
    for ch in string:
        if u'\u4e00' <= ch <= u'\u9fff':
            return True
    return False


with WaapiClient() as client:
    """****************Wwise功能区****************"""
    """查找对象"""


    def find_obj(args):
        options = {
            'return': ['name']

        }
        obj_sub_list = client.call("ak.wwise.core.object.get", args, options=options)['return']
        obj_name_list = []
        for obj_sub_dict in obj_sub_list:
            obj_name_list.append(obj_sub_dict['name'])
        return obj_name_list


    rnd_container_list = find_obj(
        {'waql': ' "%s" select descendants where type = "RandomSequenceContainer" ' % waapi_path})
    # pprint(rnd_container_list)

    """*****************主程序处理******************"""

    # 提取规则：只提取xlsx文件
    for i in file_name_list:
        if ".xlsx" in i:
            # 拼接xlsx的路径
            file_path_xlsx = os.path.join(py_path, i)
            # 获取xlsx的workbook
            wb = openpyxl.load_workbook(file_path_xlsx)
            # 获取xlsx的所有sheet
            sheet_names = wb.sheetnames
            # 加载所有工作表
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                # 获取工作表第一行数据
                for cell in list(sheet.rows)[0]:
                    if 'Sample Name' in str(cell.value):
                        # 获取音效名下的内容
                        for cell_sound in list(sheet.columns)[cell.column - 1]:
                            # 空格和中文不检测
                            if cell_sound.value != None:
                                if check_is_chinese(cell_sound.value) == False:
                                    """❤❤❤❤数据获取❤❤❤❤"""
                                    """测试名称"""
                                    new_name = cell_sound.value
                                    old_name = sheet.cell(row=cell_sound.row, column=cell_sound.column - 1).value
                                    # pprint(old_name)
                                    if new_name in rnd_container_list:
                                        # 表中有名字，但没有对应的wav文件
                                        if old_name and (old_name + '.wav' not in file_wav_dict):
                                            pprint(old_name + ":目前没有表中对应的wav文件")
                                        else:
                                            # 有名字也有文件
                                            if old_name:
                                                pprint(old_name)
                                            # 表里没有旧名字，应该没资源
                                            else:
                                                pprint(new_name)



