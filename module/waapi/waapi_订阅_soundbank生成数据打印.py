from waapi import WaapiClient, CannotConnectToWaapiException
from pprint import pprint
import openpyxl
from module.excel.excel_h import excel_get_sheet as excel_get_sheet

excel = r'S:\chen.gong_DCC_Audio\Audio\Tool\生成的wem文件查询.xlsx'
# 打开excel表
sheet, wb = excel_get_sheet(excel, "Sheet1")

# 存放.wem的list
wem_id_list = []

# 先清除excel数据
for cell in list(sheet.columns)[0]:
    # 第一排和空的不读
    if cell.row != 1:
        sheet.delete_rows(cell.row)

# Python 的异常处理 try…except…else… 语句
try:
    client = WaapiClient()

except CannotConnectToWaapiException:
    print("Could not connect to Waapi: Is Wwise running and Wwise Authoring API enabled?")

else:

    """生成soundbank时订阅生成信息发送"""


    def soundbank_generated(*args, **kwargs):

        bankinfo_list = kwargs.get("bankInfo")
        print("*****生成的单个SoundBank*****")
        # pprint(bankinfo_list)

        for bankinfo_dict in bankinfo_list:
            print("*****生成的单个BankInfo*****")
            # pprint(bankinfo_dict)

            if 'Media' in bankinfo_dict:
                for media_dict in bankinfo_dict['Media']:
                    wem_id = media_dict['Id']
                    wem_name = media_dict['ShortName']
                    if wem_id not in wem_id_list:
                        # 加入一个ID表进行后面重复的比对
                        wem_id_list.append(wem_id)
                        # pprint(wem_id)
                        # pprint(wem_name)
                        insert_row = sheet.max_row + 1
                        # 插入为空的行
                        sheet.cell(row=insert_row, column=1).value = int(wem_id)
                        # 事件名称赋值
                        sheet.cell(row=insert_row, column=2).value = wem_name

            print("*****生成的单个BankInfo结束*****")

        # 数据解析
        # bankinfo_wem = kwargs.get("bankInfo", {}).get("Media")

        print("*****生成的单个SoundBank结束*****\n")

        # 执行完成后断开 WAMP 连接，当然，要是想一直监控信息也可以不断开
        client.disconnect()


    # 订阅所需主题，传入回调函数
    client.subscribe("ak.wwise.core.soundbank.generated", soundbank_generated, infoFile=True)

    """soundbank生成"""
    args = {
        "soundbanks": [
            {"name": "Stop_All"}
        ],
        "writeToDisk": True,
        # "clearAudioFileCache": True
    }

    gen_log = client.call("ak.wwise.core.soundbank.generate", args)

wb.save(excel)
