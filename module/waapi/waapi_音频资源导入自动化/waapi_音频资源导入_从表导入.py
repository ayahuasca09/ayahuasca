import openpyxl
import json
import sys
from os.path import abspath, dirname
import os
from pprint import pprint
import re
from waapi import WaapiClient
import shutil
from openpyxl.cell import MergedCell

# 文件所在目录
py_path = ""
if hasattr(sys, 'frozen'):
    py_path = dirname(sys.executable)
elif __file__:
    py_path = dirname(abspath(__file__))

"""获取json文件路径"""
json_name = '导入规则.json'
json_path = os.path.join(py_path, json_name)

# 打开json文件并转为python字典
with open(json_path, 'r', encoding='UTF-8') as jsfile:
    js_dict = json.load(jsfile)

# 获取Wwise的配置
wwise_dict = js_dict["Wwise"]
wwise_proj_path = wwise_dict['WwiseProj_Root']

"""获取.xlsx文件"""

file_names = []
for i in os.walk(py_path):
    file_names.append(i)
# pprint("输出文件夹下的文件名：")
file_name_list = file_names[0][2]

# 一级系统名称
system_name = ""

# 状态列表
status_list = ['占位', '临时资源', 'to do', 'in progress', 'done']
# 要删除的状态列表
del_status_list = ['占位', '临时资源', 'to do', 'in progress', 'done', 'cancel']

# 事件描述字典：用于记录描述键和行作为值
event_descrip_dict = {}

# 是否导入语音
wwise_vo_media_path = wwise_proj_path + "\\Originals\\Voices"

"""*****************功能检测区******************"""
"""遍历文件目录获取文件名称和路径"""


def get_type_file_name_and_path(file_type, dir_path):
    file_dict = {}
    file_list = []
    # 遍历文件夹下的所有子文件
    # 绝对路径，子文件夹，文件名
    for root, dirs, files in os.walk(dir_path):
        # {'name': ['1111.akd', 'Creature Growls 4.akd', 'Sonic Salute_005_2_20.akd'],
        #  'path': 'S:\\chen.gong_DCC_Audio\\Audio\\SilverPalace_WwiseProject\\Originals\\SFX'}
        for file in files:
            if file_type in file:
                new_dict = {}
                file_dict[file] = os.path.join(
                    root, file)
                new_dict[file] = os.path.join(
                    root, file)
                file_list.append(new_dict)

    return file_dict, file_list


"""报错捕获"""


def print_error(error_info):
    global is_pass
    is_pass = False
    print("[error]" + error_info)


def print_warning(warn_info):
    print("[warning]" + warn_info)


"""检测字符串是否含有中文"""


def check_is_chinese(string):
    for ch in string:
        if u'\u4e00' <= ch <= u'\u9fff':
            return True
    return False


"""通过正则表达式检查"""


def check_by_re(pattern, name):
    pattern = str(pattern)
    name = str(name)
    result = re.search(pattern, name)
    new_name = ""
    if result == None:
        print_error(cell_sound.value + "：请检查" + pattern + "是否拼写错误或未添加到列表中")
    else:
        new_name = name.replace(result.group(), "")
    return new_name


"""字符串长度检查"""


def check_by_str_length(str1, length):
    if len(str1) > length:
        if "_" in str1:
            print_error(cell_sound.value + "：描述后缀名过长，限制字符数为" + str(length))
        else:
            print_error(cell_sound.value + "：通过_切分的某个单词过长，每个单词长度不应超过10个字符，需要再拆分")


"""模块类型输出报错"""


def print_error_by_module(module_dict):
    module_type = ""
    for module in module_dict:
        # 如果不是最后一位，需要加|
        if module != list(module_dict.keys())[-1]:
            module_type += module + "|"
        else:
            module_type += module
    print_error(cell_sound.value + "：" + module_type + "有误，请检查是否添加模块名称或是否拼写有误")


"""随机容器后缀检查"""


def check_is_random():
    is_random = re.search(r"(_R\d{2,4})$", name)
    if is_random:
        return True
    else:
        return False


"""Amb类型检查"""


def check_by_Amb(name):
    global is_pass
    is_pass = False


"""CG类型检查"""


def check_by_CG(name):
    global is_pass
    is_pass = False


"""Imp类型检查"""


def check_by_Imp(name):
    if name:
        # sub
        name = check_by_re(js_dict[system_name]['sub'] + "_", name)
        if name:
            # property
            name = check_by_re(js_dict[system_name]['property'] + "_", name)
            if name:
                # desc1
                name = check_by_re(js_dict[system_name]['desc1'] + "_", name)
                if name:
                    # desc2
                    check_by_re(js_dict[system_name]['desc2'], name)
                else:
                    print_error(cell_sound.value + "：" + js_dict[system_name][
                        'desc2'] + "有误，请检查是否添加模块名称或是否拼写有误")
            else:
                print_error(cell_sound.value + "：" + js_dict[system_name][
                    'desc1'] + "有误，请检查是否添加模块名称或是否拼写有误")
        else:
            print_error(
                cell_sound.value + "：" + js_dict[system_name]['property'] + "有误，请检查是否添加模块名称或是否拼写有误")
    else:
        print_error(
            cell_sound.value + "：" + js_dict[system_name]['sub'] + "有误，请检查是否添加模块名称或是否拼写有误")


"""Mus类型检查"""


def check_by_Mus(name):
    global is_pass
    is_pass = False


"""VO类型检查"""


def check_by_VO(name):
    module_dict = js_dict[system_name]['module']
    flag = 0
    for module in module_dict:
        # 模块层查询
        if module + "_" in name:
            name = name.replace(module + "_", "")
            if name:
                # 类型名查询
                name = check_by_re(module_dict[module]['property'] + "_", name)
                if name:
                    # 角色名查询
                    name = check_by_re(js_dict[system_name]['name'] + "_", name)
                    if name:
                        # 数字编号查询
                        name = check_by_re(r"\d*", name)
            flag = 1
            break
    if flag == 0:
        print_error_by_module(module_dict)


"""Char类型检查"""


def check_by_Char(name):
    module_dict = js_dict[system_name]['module']
    flag = 0
    is_mov = False
    for module in module_dict:
        # 模块层查询
        if module + "_" in name:
            name = name.replace(module + "_", "")
            # 角色名层查询
            name = check_by_re(js_dict[system_name]['name'] + "_", name)
            if name:
                if "Mov_" in name:
                    is_mov = True
                # property
                name = check_by_re(module_dict[module]['property'] + "_*", name)
                # 长度限制查询
                if name:
                    check_by_str_length(name, js_dict[system_name]['length'])
                # Mov层加查询
                if is_mov:
                    name = check_by_re(module_dict[module]['property2'] + "_*", name)
                    if name:
                        name = check_by_re(module_dict[module]['property3'] + "_*", name)

            flag = 1
            break

    if flag == 0:
        print_error_by_module(module_dict)


"""Mon类型检查"""


def check_by_Mon(name):
    type_dict = js_dict[system_name]
    flag = 0
    for type in type_dict:
        if type in name:
            name = name.replace(type + "_", "")
            flag = 1
            module_dict = js_dict[system_name][type]['module']
            flag = 0
            for module in module_dict:
                # 模块层查询
                if module + "_" in name:
                    name = name.replace(module + "_", "")
                    # Boss名层查询
                    name = check_by_re(js_dict[system_name][type]['name'] + "_", name)
                    if name:
                        # 技能层查询
                        name = check_by_re(module_dict[module]['property'] + "_*", name)
                        # 长度限制查询
                        if name:
                            check_by_str_length(name, js_dict[system_name][type]['length'])

                    flag = 1
                    break

            if flag == 0:
                print_error_by_module(module_dict)
            break
    if flag == 0:
        print_error_by_module(type_dict)


def check_by_Sys(name):
    module_dict = js_dict[system_name]['module']
    flag = 0
    for module in module_dict:
        # 模块层查询
        if module + "_" in name:
            name = name.replace(module + "_", "")
            # Show类查询
            if module == "Show":
                if name:
                    # 角色名称
                    name = check_by_re(module_dict[module]['property'] + "_", name)
                    if name:
                        # 皮肤名称
                        name = check_by_re(module_dict[module]['property2'] + "_", name)
                        if name:
                            # 动作名称
                            name = check_by_re(module_dict[module]['property3'] + "_*", name)
                            if name:
                                # 长度限制查询
                                check_by_str_length(name, module_dict[module]['length'])
            flag = 1
            break
    if flag == 0:
        print_error_by_module(module_dict)


"""检查LP是否在末尾"""


def check_LP_in_last():
    if "LP" in name:
        # print("LP")
        # .*代表所有字符 _LP$代表以_LP结尾
        pattern = r".*_LP$"
        result = re.match(pattern, name)
        if result:
            new_name = re.sub(r"_LP$", "", name)
            return new_name
        else:
            print_error(cell_sound.value + "：_LP应放在最末尾")
    else:
        return name


"""检查是否使用通用词汇"""


def check_by_com_word(name):
    com_word_dict = js_dict['Gen_Word']
    for key in com_word_dict:
        for value in com_word_dict[key]:
            if value in name:
                print_error(cell_sound.value + value + "应改为通用词" + key)
                if "Medium" in name:
                    print_error(cell_sound.value + "PS：Mid表体积/重量，Med表距离")


"""分隔符的大小写和长度检查"""


def check_by_length_and_word():
    # pprint(cell_sound.value)
    word_list = name.split("_")
    is_title = True
    for word in word_list:
        """判定_的每个都不能超过10个"""
        check_by_str_length(word, 10)
        """判定_的每个开头必须大写"""
        if len(word) > 0:
            # 只需要检查第一个字符，因为istitle会导致检查字符串只能首字母大写
            if word[0].istitle() == False and word[0].isdigit() == False:
                is_title = False
                break
    if is_title == False:
        print_error(cell_sound.value + "：通过”_“分隔的每个单词开头都需要大写")
    if word_list:
        return word_list[0]
    else:
        return None


"""获取一级系统名走不同的检测方式"""


def check_first_system_name(name):
    if system_name and (system_name in js_dict):
        func_name = "check_by_" + system_name
        eval(func_name)(name.replace(system_name + "_", ""))
    else:
        print_error_by_module(js_dict)


"""获取表格中的事件描述和状态所在的列"""


def get_descrip_and_status_column():
    require_module_column = None
    second_module_column = None
    require_name_column = None
    status_column = None
    if list(sheet.rows)[0]:
        for cell in list(sheet.rows)[0]:
            if cell.value:
                if 'Require Module' in str(cell.value):
                    require_module_column = cell.column
                    # print(require_module_column)
                elif 'Second Module' in str(cell.value):
                    second_module_column = cell.column
                    # print(second_module_column)
                elif 'Require Name' in str(cell.value):
                    require_name_column = cell.column
                    # print(require_name_column)
                elif 'Status' in str(cell.value):
                    status_column = cell.column
                    # print(status_column)

    return require_module_column, second_module_column, require_name_column, status_column


"""检查是否为合并单元格"""


def check_is_mergecell(cell):
    if isinstance(cell, MergedCell):  # 判断该单元格是否为合并单元格
        for merged_range in sheet.merged_cells.ranges:  # 循环查找该单元格所属的合并区域
            if cell.coordinate in merged_range:
                # 获取合并区域左上角的单元格作为该单元格的值返回
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    return cell.value


"""拼接事件描述"""


def get_event_descrip():
    event_descrip = ""
    if require_module_column:
        require_module_value = check_is_mergecell(sheet.cell(row=cell_sound.row, column=require_module_column))
        if require_module_value:
            event_descrip = require_module_value + "_"
    if second_module_column:
        second_module_value = check_is_mergecell(sheet.cell(row=cell_sound.row, column=second_module_column))
        if event_descrip:
            if second_module_value:
                event_descrip = event_descrip + second_module_value + "_"
        else:
            if second_module_value:
                event_descrip = second_module_value + "_"

    if require_name_column:
        require_name_value = check_is_mergecell(sheet.cell(row=cell_sound.row, column=require_name_column))
        if event_descrip:
            if require_name_value:
                event_descrip = event_descrip + require_name_value + "_"
        else:
            if require_name_value:
                event_descrip = require_name_value + "_"
    if not event_descrip:
        if cell_sound.value:
            print_error(cell_sound.value + "：没有事件描述，请补充！")

    # 格式规范
    else:
        event_descrip = re.sub(r'^_', "", event_descrip)
        event_descrip = re.sub(r'_*$', "", event_descrip)
        event_descrip = re.sub(r'\s+', "", event_descrip)
    return event_descrip


with WaapiClient() as client:
    """*****************Wwise功能区******************"""

    """查找对象"""


    def find_obj(args):
        options = {
            'return': ['name', 'id', 'path', 'notes', 'originalWavFilePath', 'type']

        }
        obj_sub_list = client.call("ak.wwise.core.object.get", args, options=options)['return']
        if not obj_sub_list:
            obj_sub_id = ""
            obj_sub_path = ""
        else:
            obj_sub_id = obj_sub_list[0]['id']
            obj_sub_path = obj_sub_list[0]['path']
        return obj_sub_list, obj_sub_id, obj_sub_path


    """获取wwise对象的最长字符串父级"""


    def find_obj_parent(obj_name, waql):
        # 所有Unit获取
        parent_list, parent_id, _ = find_obj(waql)

        # 存储符合的对象的父级，最后找出最长最符合的
        parent_len = 0
        obj_parent_id = {}
        if parent_list != None:
            for parent_dict in parent_list:
                if parent_dict['name'] in obj_name:
                    if len(parent_dict['name']) > parent_len:
                        parent_len = len(parent_dict['name'])
                        obj_parent_id = parent_dict['id']
        else:
            print_error(cell_sound.value + "相应的Unit目录结构未创建！需要创建")

        return obj_parent_id


    """媒体资源创建及导入"""


    def import_media(media_name, rnd_path):
        # 媒体资源复制
        # 复制到的目录
        copy_catalog = os.path.join(py_path, "New_Media")
        source_path = shutil.copy2(os.path.join(py_path, "Media_Temp.wav"),
                                   os.path.join(copy_catalog, media_name + ".wav"))

        # 在容器中创建媒体资源
        import_media_in_rnd(source_path, media_name, rnd_path)


    """修改Wwise内容的名字"""


    def change_name_by_wwise_content(obj_id, name, old_name, obj_type):
        args = {
            "objects": [
                {

                    "object": obj_id,
                    "name": name,
                }
            ]
        }
        client.call("ak.wwise.core.object.set", args)
        print_warning(old_name + "(" + obj_type + ")改名为：" + name)


    """设置obj的notes"""


    def set_obj_notes(obj_id, notes_value):
        args = {
            'object': obj_id,
            'value': notes_value
        }
        client.call("ak.wwise.core.object.setNotes", args)
        # print_warning(obj_name + "描述更改为：" + notes_value)


    """导入媒体资源"""


    def import_sound_sfx_and_media(media_name, rnd_path):
        # 媒体资源导入
        import_media(media_name, rnd_path)
        # Sound颜色设置
        _, sound_container_id, _ = find_obj(
            {'waql': 'from type Sound where name = "%s"' % media_name})
        if sound_container_id:
            set_sound_color_default(sound_container_id)

        print_warning(media_name + "：Sound创建及Media导入")


    """创建新的随机容器"""


    def create_rnd_container(media_name, system_name):
        # 查找该rnd是否已存在
        flag = 0
        # 去除随机容器数字
        rnd_name = re.sub(r"(_R\d{2,4})$", "", media_name)

        rnd_path = ""
        rnd_container_list, rnd_id, _ = find_obj(
            {'waql': ' "%s" select descendants where type = "RandomSequenceContainer" ' % os.path.join(
                wwise_dict['Root'], system_name)})
        # 找到重名的
        for rnd_container_dict in rnd_container_list:
            if rnd_container_dict['name'] == rnd_name:
                flag = 1
                rnd_path = rnd_container_dict['path']
                set_obj_notes(rnd_container_dict['id'], event_descrip)
                break
            # 没有重名但描述一致，判定为改名
            elif rnd_container_dict['notes'] == event_descrip:
                # pprint(rnd_name + "：RandomContainer已存在，将不再导入")
                flag = 2
                rnd_path = rnd_container_dict['path'].replace(rnd_container_dict['name'], rnd_name)
                change_name_by_wwise_content(rnd_container_dict['id'], rnd_name, rnd_container_dict['name'],
                                             "RandContainer")
                # Sound也要改名
                sound_container_list, sound_container_id, _ = find_obj(
                    {'waql': ' "%s" select children  ' % rnd_container_dict['id']})
                for sound_cotainer_dict in sound_container_list:
                    wav_tail = re.search(r"(_R\d{2,4})$", sound_cotainer_dict['name'])
                    sound_name = rnd_name
                    if wav_tail:
                        sound_name = rnd_name + wav_tail.group()
                    change_name_by_wwise_content(sound_cotainer_dict['id'], sound_name, sound_cotainer_dict['name'],
                                                 "Sound")
                    # Media路径更改
                    new_media_path = sound_cotainer_dict['originalWavFilePath'].replace(sound_cotainer_dict['name'],
                                                                                        sound_name)
                    # Media重命名
                    os.rename(sound_cotainer_dict['originalWavFilePath'],
                              new_media_path)
                    # 在容器中重新导入media
                    import_media_in_rnd(new_media_path, sound_name, rnd_path)
                    print_warning(sound_cotainer_dict['originalWavFilePath'] + "(Media)改名为：" + new_media_path)

                    # 语音的话需要删除容器内的引用
                    if system_name == "VO":
                        sound_language_list, _, _ = find_obj(
                            {'waql': ' "%s" select children  ' %
                                     sound_cotainer_dict['id']})
                        if sound_language_list:
                            for sound_language_dict in sound_language_list:
                                if sound_cotainer_dict['name'] in sound_language_dict['name']:
                                    args = {
                                        "object": "%s" % sound_language_dict['id']
                                    }
                                    client.call("ak.wwise.core.object.delete", args)
                                    print_warning(
                                        sound_language_dict['name'] + "：容器引用删除")

                break

        # 不存在则创建
        if flag == 0:
            # 查找所在路径
            rnd_parent_id = find_obj_parent(rnd_name,
                                            {
                                                'waql': ' "%s" select descendants,this where type = "ActorMixer" ' %
                                                        os.path.join(
                                                            wwise_dict['Root'], system_name)})
            if rnd_parent_id:

                # 创建的rnd属性
                args = {
                    # 选择父级
                    "parent": rnd_parent_id,
                    # 创建类型机名称
                    "type": "RandomSequenceContainer",
                    "name": rnd_name,
                    "notes": event_descrip,
                    "@RandomOrSequence": 1,
                    "@NormalOrShuffle": 1,
                    "@RandomAvoidRepeatingCount": 1
                }
                rnd_container_object = client.call("ak.wwise.core.object.create", args)
                # 查找新创建的容器的路径
                _, _, rnd_path = find_obj(
                    {'waql': ' "%s"  ' % rnd_container_object['id']})
                if rnd_path:
                    print_warning(rnd_name + "：RandContainer创建")
                else:
                    print_error(rnd_name + "：RandContainer未能成功创建")

                # 声音容器创建及导入
                import_sound_sfx_and_media(media_name, rnd_path)

        return rnd_path, rnd_name


    """设置新的占位Sound资源超越父级为灰色"""


    def set_sound_color_default(obj_id):
        args_property = {
            "object": obj_id,
            "property": "OverrideColor",
            "value": True,
        }
        client.call("ak.wwise.core.object.setProperty", args_property)

        args_property = {
            "object": obj_id,
            "property": "Color",
            "value": 0,
        }
        client.call("ak.wwise.core.object.setProperty", args_property)


    """VO：复制Chinese下的文件到其他文件夹里"""


    def copy_reference_VO_to_other(media_name):
        wwise_vo_media_path_cn = wwise_vo_media_path + "\\Chinese" + "\\" + media_name + ".wav"
        wwise_vo_media_path_en = wwise_vo_media_path + "\\English" + "\\" + media_name + ".wav"
        wwise_vo_media_path_jp = wwise_vo_media_path + "\\Japanese" + "\\" + media_name + ".wav"
        wwise_vo_media_path_kr = wwise_vo_media_path + "\\Korean" + "\\" + media_name + ".wav"
        if os.path.isfile(wwise_vo_media_path_cn):
            # print(wwise_vo_media_path_cn)
            if not os.path.isfile(wwise_vo_media_path_en):
                source_path = shutil.copy2(wwise_vo_media_path_cn,
                                           wwise_vo_media_path_en)
                # print(wwise_vo_media_path_en)
            if not os.path.isfile(wwise_vo_media_path_jp):
                source_path = shutil.copy2(wwise_vo_media_path_cn,
                                           wwise_vo_media_path_jp)
                # print(wwise_vo_media_path_jp)
            if not os.path.isfile(wwise_vo_media_path_kr):
                source_path = shutil.copy2(wwise_vo_media_path_cn,
                                           wwise_vo_media_path_kr)
                # print(wwise_vo_media_path_kr)


    """媒体资源导入"""


    def import_media_in_rnd(source_path, media_name, rnd_path):
        # 导入语音
        if system_name == "VO":
            args_import = {
                "importOperation": "useExisting",
                "imports": [
                    {
                        "audioFile": source_path,
                        "objectPath": rnd_path + "\\<Sound Voice>" + media_name + "\\<AudioFileSource>" + "CN_" + media_name,
                        "importLanguage": "Chinese"
                    },
                    {
                        "audioFile": source_path,
                        "objectPath": rnd_path + "\\<Sound Voice>" + media_name + "\\<AudioFileSource>" + "EN_" + media_name,
                        "importLanguage": "English"
                    },
                    {
                        "audioFile": source_path,
                        "objectPath": rnd_path + "\\<Sound Voice>" + media_name + "\\<AudioFileSource>" + "JP_" + media_name,
                        "importLanguage": "Japanese"
                    },
                    {
                        "audioFile": source_path,
                        "objectPath": rnd_path + "\\<Sound Voice>" + media_name + "\\<AudioFileSource>" + "KR_" + media_name,
                        "importLanguage": "Korean"
                    }
                ]
            }
        # 导入音效
        else:
            args_import = {
                # createNew
                # useExisting：会增加一个新媒体文件但旧的不会删除
                # replaceExisting:会销毁Sound，上面做的设置都无了
                "importOperation": "replaceExisting",
                "default": {
                    "importLanguage": "SFX"
                },
                "imports": [
                    {
                        "audioFile": source_path,
                        "objectPath": rnd_path + '\\<Sound SFX>' + media_name,
                        "originalsSubFolder": system_name
                        #                                                         名为Test 0的顺序容器            名为My SFX 0 的音效
                        # "objectPath": "\\Actor-Mixer Hierarchy\\Default Work Unit\\<Sequence Container>Test 0\\<Sound SFX>My SFX 0"
                    }
                ]
            }

        # 定义返回结果参数，让其只返回 Windows 平台下的信息，信息中包含 GUID 和新创建的对象名
        opts = {
            "platform": "Windows",
            "return": [
                "id", "name"
            ]
        }
        client.call("ak.wwise.core.audio.import", args_import, options=opts)

        if system_name == "VO":
            copy_reference_VO_to_other(media_name)


    """创建事件的参数"""


    def get_event_args(event_parent_path, event_name, event_action, event_target, notes):
        args_new_event = {
            # 上半部分属性中分别为 Event 创建后存放的路径、类型、名称、遇到名字冲突时的处理方法
            "parent": event_parent_path,
            "type": "Event",
            "name": event_name,
            "onNameConflict": "merge",
            "children": [
                {
                    # 为其创建播放行为，名字留空，使用 @ActionType 定义其播放行为为 Play，@Target 为被播放的声音对象
                    "name": "",
                    "type": "Action",
                    "@ActionType": event_action,
                    "@Target": event_target
                }
            ],
            "notes": notes
        }
        return args_new_event


    """事件生成"""


    def create_event(rnd_name, rnd_path):
        parent_id = find_obj_parent(rnd_name, {
            'waql': ' "%s" select descendants where type = "WorkUnit" ' % os.path.join(wwise_dict['Event_Root'],
                                                                                       system_name)})
        event_name = "AKE_" + "Play_" + rnd_name
        event_stop_name = "AKE_" + "Stop_" + rnd_name
        # 查找事件是否已存在
        event_list, _, _ = find_obj(
            {'waql': ' "%s" select descendants where type = "Event" ' %
                     wwise_dict['Event_Root']})

        # 判断其random容器的父级是不是actor-mixer，如果不是就不生成event
        rnd_parent_list, _, _ = find_obj(
            {'waql': ' "%s" select parent ' %
                     rnd_path})
        if rnd_parent_list:
            if rnd_parent_list[0]['type'] == "ActorMixer":
                # pprint(event_list)
                flag = 0
                for event_dict in event_list:
                    if rnd_name in event_dict['name']:
                        # Play事件的notes重新设置
                        if event_dict['name'] == event_name:
                            flag = 1
                            set_obj_notes(event_dict['id'], event_descrip)
                            # print(event_dict['name'])
                        elif event_dict['name'] == event_stop_name:
                            flag = 2
                            set_obj_notes(event_dict['id'], event_descrip + "停止")
                            # print(event_dict['name'])
                    else:
                        # Play:没有重名但描述一致，判定为改名
                        if event_dict['notes'] == event_descrip:
                            # pprint(rnd_name + "：RandomContainer已存在，将不再导入")
                            flag = 3
                            change_name_by_wwise_content(event_dict['id'], event_name, event_dict['name'], "Event")
                        # Stop：没有重名但描述一致，判定为改名
                        elif event_dict['notes'] == event_descrip + "停止":
                            # pprint(rnd_name + "：RandomContainer已存在，将不再导入")
                            flag = 4
                            change_name_by_wwise_content(event_dict['id'], event_stop_name, event_dict['name'], "Event")

                if flag == 0:
                    event_args = get_event_args(parent_id, event_name, 1, rnd_path, event_descrip)
                    client.call("ak.wwise.core.object.create", event_args)
                    # {'children': [{'id': '{26A1FB06-3908-4F55-AB8A-E67264100F18}', 'name': ''}],
                    #  'id': '{F04A7B4F-9266-4DAB-8DD2-CAF6BFD6D78A}',
                    #  'name': 'AKE_Play_Mon_Boss_Body_B01_Atk5'}
                    print_warning(event_name + "事件创建")
                    if "_LP" in rnd_name:
                        event_name = "AKE_" + "Stop_" + rnd_name
                        event_args = get_event_args(parent_id, event_name, 2, rnd_path, event_descrip + "停止")
                        client.call("ak.wwise.core.object.create", event_args)
                        pprint(event_name + "事件创建")


    """媒体资源导入的总流程"""


    def create_wwise_content(media_name, system_name):

        # 随机容器创建
        rnd_path, rnd_name = create_rnd_container(media_name, system_name)

        # 事件自动生成
        create_event(rnd_name, rnd_path)


    def delete_wwise_content():
        # 随机容器引用的资产删除
        media_list, media_id, _ = find_obj(
            {'waql': ' "%s" select children where type = "Sound" ' % rnd_container['id']})
        if media_list:
            for media_dict in media_list:
                if os.path.isfile(media_dict['originalWavFilePath']):
                    os.remove(media_dict['originalWavFilePath'])
                    print_warning(media_dict['originalWavFilePath'] + "已删除")

        # 随机容器删除
        args = {
            "object": "%s" % rnd_container['id']
        }
        client.call("ak.wwise.core.object.delete", args)
        print_warning(rnd_container['name'] + ":RandomContainer已删除")

        # 事件删除
        for event_dict in event_list:
            if rnd_container['name'] in event_dict['name']:
                args = {
                    "object": "%s" % event_dict['id']
                }
                client.call("ak.wwise.core.object.delete", args)
                print_warning(event_dict['name'] + ":Event已删除")


    def delete_or_modify_wwise_content():
        for i in file_name_list:
            if ".xlsx" in i:
                # 拼接xlsx的路径
                file_path_xlsx = os.path.join(py_path, i)
                # 获取xlsx的workbook
                wb = openpyxl.load_workbook(file_path_xlsx)
                # 获取xlsx的所有sheet
                sheet_names = wb.sheetnames
                # 加载所有工作表
                for sheet_name in sheet_names:
                    sheet = wb[sheet_name]
                    # 获取工作表第一行数据
                    for cell in list(sheet.rows)[0]:
                        if 'Sample Name' in str(cell.value):
                            # 获取音效名下的内容
                            for cell_sound in list(sheet.columns)[cell.column - 1]:
                                # 空格和中文不检测
                                if cell_sound.value != None:
                                    if check_is_chinese(cell_sound.value) == False:

                                        """事件描述"""
                                        event_descrip = get_event_descrip()

                                        # 如果音效名与rnd名相同
                                        if rnd_container['name'] == cell_sound.value:
                                            # 如果状态标为取消的删除
                                            for status in list(sheet.rows)[cell_sound.row - 1]:
                                                if "cancel" == status.value:
                                                    delete_wwise_content()
                                            return
        # 其他为在表格中搜索不到的名字
        delete_wwise_content()


    # 删除语音不对应的rnd sound资源
    def delete_wwise_content_vo_sound():
        # Sound容器删除
        args = {
            "object": "%s" % sound_container['id']
        }
        client.call("ak.wwise.core.object.delete", args)
        print_warning(rnd_container['name'] + ":SoundContainer已删除")

        # Sound容器引用的资产删除
        if os.path.isfile(sound_container['originalWavFilePath']):
            os.remove(sound_container['originalWavFilePath'])
            print_warning(sound_container['originalWavFilePath'] + "已删除")


    # 删除语音不对应的rnd sound资源
    def delete_or_modify_wwise_content_vo_sound():
        for i in file_name_list:
            if ".xlsx" in i:
                # 拼接xlsx的路径
                file_path_xlsx = os.path.join(py_path, i)
                # 获取xlsx的workbook
                wb = openpyxl.load_workbook(file_path_xlsx)
                # 获取xlsx的所有sheet
                sheet_names = wb.sheetnames
                # 加载所有工作表
                for sheet_name in sheet_names:
                    sheet = wb[sheet_name]
                    # 获取工作表第一行数据
                    for cell in list(sheet.rows)[0]:
                        if 'Sample Name' in str(cell.value):
                            # 获取音效名下的内容
                            for cell_sound in list(sheet.columns)[cell.column - 1]:
                                # 空格和中文不检测
                                if cell_sound.value != None:
                                    if check_is_chinese(cell_sound.value) == False:

                                        """事件描述"""
                                        event_descrip = get_event_descrip()

                                        # 如果音效名与rnd名相同
                                        if sound_container['name'] == cell_sound.value:
                                            # 如果状态标为取消的删除
                                            for status in list(sheet.rows)[cell_sound.row - 1]:
                                                if "cancel" == status.value:
                                                    delete_wwise_content_vo_sound()
                                            return
        # 其他为在表格中搜索不到的名字
        delete_wwise_content_vo_sound()


    """*****************主程序处理******************"""
    # 撤销开始
    client.call("ak.wwise.core.undo.beginGroup")

    # 记录所有资源名称
    sound_name_list = []

    # 提取规则：只提取xlsx文件
    for i in file_name_list:
        if ".xlsx" in i:
            # 拼接xlsx的路径
            file_path_xlsx = os.path.join(py_path, i)
            # 获取xlsx的workbook
            wb = openpyxl.load_workbook(file_path_xlsx)
            # 获取xlsx的所有sheet
            sheet_names = wb.sheetnames
            # 加载所有工作表
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                require_module_column, second_module_column, require_name_column, status_column = get_descrip_and_status_column()
                # 获取工作表第一行数据
                for cell in list(sheet.rows)[0]:
                    if 'Sample Name' in str(cell.value):
                        # 获取音效名下的内容
                        for cell_sound in list(sheet.columns)[cell.column - 1]:
                            # 空格和中文不检测
                            if cell_sound.value:
                                if not check_is_chinese(cell_sound.value):
                                    if status_column:
                                        if sheet.cell(row=cell_sound.row,
                                                      column=status_column).value in status_list:

                                            """❤❤❤❤数据获取❤❤❤❤"""
                                            """检测表格内容"""
                                            is_pass = True

                                            """检查表格中是否有内容重复项"""
                                            if cell_sound.value in sound_name_list:
                                                print_error(cell_sound.value + "：表格中有重复项音效名，请检查")
                                            else:
                                                sound_name_list.append(cell_sound.value)
                                                """测试名称"""
                                                name = cell_sound.value
                                                """事件描述"""
                                                event_descrip = get_event_descrip()
                                                if (event_descrip in event_descrip_dict) and (not check_is_random()):
                                                    print_error(event_descrip + "：表格中有重复项描述，请检查")
                                                else:
                                                    event_descrip_dict[event_descrip] = cell_sound.value

                                                # print(event_descrip)

                                                if is_pass:
                                                    # 分隔符的大小写和长度检查
                                                    system_name = check_by_length_and_word()

                                                    # 检查LP是否在末尾
                                                    name = check_LP_in_last()
                                                    # print(name)

                                                    # 通用词汇检查
                                                    check_by_com_word(name)

                                                    # 检查一级系统并索引到不同的检测方式
                                                    check_first_system_name(name)
                                                    # print(name)

                                                    # 检查是否为随机样本
                                                    if check_is_random():
                                                        is_pass = False

                                                    # 生成Wwise内容
                                                    if is_pass:
                                                        create_wwise_content(cell_sound.value, system_name)

                                        # else:
                                        #     print_warning(cell_sound.value + "：该状态不会生成Wwise占位资源，请检查")

    """同步表中删除的内容"""
    # 查找所有Rnd和Event
    rnd_container_list, rnd_id, _ = find_obj(
        {'waql': ' "%s" select descendants where type = "RandomSequenceContainer" ' % wwise_dict['Root']})
    # [{'id': '{C1BFDDC1-CA6F-45BC-8B43-15AE866AA20A}',
    #   'name': 'Char_Skill_C01_Atk1',
    #   'notes': '',
    #   'path': '\\Actor-Mixer '
    #           'Hierarchy\\v1\\Char\\Char\\Char_Skill\\Char_Skill\\Char_Skill_C01\\Char_Skill_C01\\Char_Skill_C01_Atk1'}
    event_list, event_id, _ = find_obj(
        {'waql': ' "%s" select descendants where type = "Event" ' % wwise_dict['Event_Root']})

    for rnd_container in rnd_container_list:
        delete_or_modify_wwise_content()

    # 对于语音，需要查找Sound容器里的随机资源
    # 查找所有Sound
    sound_container_list, _, _ = find_obj(
        {'waql': ' "%s" select descendants where type = "Sound" ' % wwise_dict['VO_Game']})

    for sound_container in sound_container_list:
        delete_or_modify_wwise_content_vo_sound()

    # 撤销结束
    client.call("ak.wwise.core.undo.endGroup", displayName="rnd创建撤销")

    # 清除复制的媒体资源
    shutil.rmtree("New_Media")
    os.mkdir("New_Media")
    os.mkdir("New_Media/Chinese")
    os.mkdir("New_Media/English")
    os.mkdir("New_Media/Japanese")
    os.mkdir("New_Media/Korean")
