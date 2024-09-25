import openpyxl
import sys
from os.path import abspath, dirname
import os
from module.oi.oi_h import get_type_file_name_and_path
from pprint import pprint
import shutil
from waapi import WaapiClient
from openpyxl.styles import colors
import re

# 文件所在目录
py_path = ""
if hasattr(sys, 'frozen'):
    py_path = dirname(sys.executable)
elif __file__:
    py_path = dirname(abspath(__file__))

# wav文件目录
# wav_path = r"S:\chen.gong_DCC_Audio\Audio\SilverPalace_WwiseProject\Originals\SFX"
wav_path = r"C:\Users\happyelements\Desktop\Old"


# 容器根目录
# waapi_path = "\\Actor-Mixer Hierarchy\\v1"


# 打开工作簿并获取工作表
def excel_get_sheet(path, sheetname):
    # 打开工作簿
    wb = openpyxl.load_workbook(path)
    # print(wb)
    # 获取工作表
    sheet = wb[sheetname]
    # print(sheet)
    return sheet, wb


"""获取.xlsx文件"""

file_names = []
for i in os.walk(py_path):
    file_names.append(i)
file_name_list = file_names[0][2]
# pprint(file_name_list)

"""获取资源导入记录表"""
excel_media_import_path = "资源导入记录表.xlsx"
sheet_m, wb_m = excel_get_sheet(excel_media_import_path, 'Sheet1')
# print(sheet_m)

"""获取.wav文件"""
file_wav_dict = get_type_file_name_and_path('.wav', wav_path)
# pprint(file_wav_dict)

"""复制.wav文件的路径"""
copy_path = r"New_Media"

"""*****************功能检测区******************"""

"""检测字符串是否含有中文"""


def check_is_chinese(string):
    for ch in string:
        if u'\u4e00' <= ch <= u'\u9fff':
            return True
    return False


with WaapiClient() as client:
    """****************Wwise功能区****************"""
    """查找对象"""


    def find_obj(args):
        options = {
            'return': ['name']

        }
        obj_sub_list = client.call("ak.wwise.core.object.get", args, options=options)['return']
        obj_name_list = []
        for obj_sub_dict in obj_sub_list:
            obj_name_list.append(obj_sub_dict['name'])
        return obj_name_list


    # rnd_container_list = find_obj(
    #     {'waql': ' "%s" select descendants where type = "RandomSequenceContainer" ' % waapi_path})
    # pprint(rnd_container_list)

    """*****************主程序处理******************"""


    # 提取规则：只提取xlsx文件
    for i in file_name_list:
        if ".xlsx" in i:
            # 拼接xlsx的路径
            file_path_xlsx = os.path.join(py_path, i)
            # 获取xlsx的workbook
            wb = openpyxl.load_workbook(file_path_xlsx)
            # 获取xlsx的所有sheet
            sheet_names = wb.sheetnames
            # 加载所有工作表
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                # 获取工作表第一行数据
                for cell in list(sheet.rows)[0]:
                    if 'Sample Name' in str(cell.value):
                        # 获取音效名下的内容
                        for cell_sound in list(sheet.columns)[cell.column - 1]:
                            # 空格和中文不检测
                            if cell_sound.value != None:
                                if check_is_chinese(cell_sound.value) == False:
                                    """❤❤❤❤数据获取❤❤❤❤"""
                                    """测试名称"""
                                    new_name = cell_sound.value
                                    old_name = sheet.cell(row=cell_sound.row, column=cell_sound.column - 1).value
                                    # pprint(old_name)
                                    """资源导入记录表"""
                                    flag = 0
                                    for cell in list(sheet_m.columns)[1]:
                                        if cell.value != None:
                                            # print(cell.value)
                                            # 如果资源表中已存在该名称
                                            if str(cell.value) == str(new_name):
                                                flag = 1
                                                insert_row = cell.row
                                                break
                                    # 表中没有记录名字则导入
                                    if flag == 0:
                                        insert_row = sheet_m.max_row + 1
                                        sheet_m.cell(row=insert_row, column=1).value = old_name
                                        sheet_m.cell(row=insert_row, column=2).value = new_name
                                    # 表中有名字，但没有对应的wav文件
                                    if old_name and (old_name + '.wav' not in file_wav_dict):
                                        # pprint(old_name + ":目前没有表中对应的wav文件")
                                        # 为随机容器的资源
                                        for wav_name in file_wav_dict:
                                            if old_name in wav_name:
                                                # 获取词缀尾部
                                                no_wav_name = re.sub(".wav", "", wav_name)
                                                wav_tail = re.search(r"(\d{2,4})$", no_wav_name)
                                                if not wav_tail:
                                                    wav_tail = re.sub(r"(_R\d{2,4})$", "", no_wav_name)
                                                if wav_tail:
                                                    # print(wav_tail.group())
                                                    new_tail = "_R" + wav_tail.group()
                                                    new_wav_name = new_name + new_tail + ".wav"
                                                    # print(new_wav_name)
                                                    # 文件复制
                                                    shutil.copy2(file_wav_dict[wav_name]
                                                                 , os.path.join(copy_path, new_wav_name))
                                                    # 颜色修改
                                                    for cell_color in list(sheet_m.rows)[insert_row - 1]:
                                                        cell_color.fill = openpyxl.styles.PatternFill(
                                                            start_color='F1FC72',
                                                            end_color='F1FC72',
                                                            fill_type='solid')
                                                    sheet_m.cell(row=insert_row, column=3).value = 1

                                    else:
                                        # 有名字也有文件
                                        if old_name:
                                            # pprint(old_name)
                                            sheet_m.cell(row=insert_row, column=6).value = 1
                                            # 颜色修改
                                            for cell_color in list(sheet_m.rows)[insert_row - 1]:
                                                cell_color.fill = openpyxl.styles.PatternFill(start_color='88DB29',
                                                                                              end_color='88DB29',
                                                                                              fill_type='solid')
                                            # 文件复制
                                            shutil.copy2(file_wav_dict[old_name + '.wav']
                                                         , os.path.join(copy_path, new_name + ".wav"))
                                        # 表里没有旧名字，应该没资源
                                        else:
                                            sheet_m.cell(row=insert_row, column=7).value = 1
                                            # 颜色修改
                                            for cell_color in list(sheet_m.rows)[insert_row - 1]:
                                                cell_color.fill = openpyxl.styles.PatternFill(start_color='BFC4D7',
                                                                                              end_color='BFC4D7',
                                                                                              fill_type='solid')
wb_m.save(excel_media_import_path)
