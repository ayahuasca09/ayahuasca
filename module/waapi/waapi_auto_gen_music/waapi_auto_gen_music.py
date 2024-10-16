import openpyxl
import json
import sys
from os.path import abspath, dirname
import os
from pprint import pprint
import re
from waapi import WaapiClient
import shutil
from openpyxl.cell import MergedCell

# 文件所在目录
py_path = ""
if hasattr(sys, 'frozen'):
    py_path = dirname(sys.executable)
elif __file__:
    py_path = dirname(abspath(__file__))

"""获取.xlsx文件"""

file_names = []
for i in os.walk(py_path):
    file_names.append(i)
# pprint("输出文件夹下的文件名：")
file_name_list = file_names[0][2]

"""音乐资源名称列表"""
mus_name_list = []
value_desc_list = []

"""state folder path"""
state_path = '{59A9E7B9-61C9-461A-A96F-4997B541A59C}'
"""music path"""
music_path = "{E8A5F6A0-AB3E-4166-BB68-C7D7ECC92B29}"
stin_path = "{DD48832A-AAC1-49F7-B57B-86F5FACD8FB3}"
trig_path = "{94BBA60A-29EF-4328-83E8-075770BF2687}"
trig_event_path = "{EDD34B59-3938-4C22-A63D-838FD8DCD57E}"
"""媒体资源路径"""
original_path = r'S:\chen.gong_DCC_Audio\Audio\SilverPalace_WwiseProject\Originals\SFX'

"""*****************功能检测区******************"""
"""遍历文件目录获取文件名称和路径"""


def get_type_file_name_and_path(file_type, dir_path):
    file_dict = {}
    file_list = []
    # 遍历文件夹下的所有子文件
    # 绝对路径，子文件夹，文件名
    for root, dirs, files in os.walk(dir_path):
        # {'name': ['1111.akd', 'Creature Growls 4.akd', 'Sonic Salute_005_2_20.akd'],
        #  'path': 'S:\\chen.gong_DCC_Audio\\Audio\\SilverPalace_WwiseProject\\Originals\\SFX'}
        for file in files:
            if file_type in file:
                new_dict = {}
                file_dict[file] = os.path.join(
                    root, file)
                new_dict[file] = os.path.join(
                    root, file)
                file_list.append(new_dict)

    return file_dict, file_list


"""正则表达式检查字符，并输出除该检查字符之外的其他其他字符"""


# content:字符串  pattern：正则表达式

def check_by_re(content, pattern):
    # 使用 re.match 来尝试匹配字符串
    match = re.match(pattern, content)

    # 如果匹配成功，返回匹配组中的内容
    if match:
        return match.group(1)
    else:
        print_error(content + "：" + pattern + "匹配失败，请重新检查格式是否正确")
        return None


"""返回列表中字典中某个键值组成的新列表"""


def get_one_value_list(dict_list, dict_key):
    extract_key = lambda d: d[dict_key]
    result = list(map(extract_key, dict_list))
    # 还需要去除为None的数据
    for i in result:
        if i == 'None':
            result.remove(i)
    # print(result)
    return result


"""检查是否为合并单元格"""


def check_is_mergecell(cell):
    if isinstance(cell, MergedCell):  # 判断该单元格是否为合并单元格
        for merged_range in sheet.merged_cells.ranges:  # 循环查找该单元格所属的合并区域
            if cell.coordinate in merged_range:
                # 获取合并区域左上角的单元格作为该单元格的值返回
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    return cell.value


"""报错捕获"""


def print_error(error_info):
    global is_pass
    is_pass = False
    print("[error]" + error_info)


"""检测字符串是否含有中文"""


def check_is_chinese(string):
    for ch in string:
        if u'\u4e00' <= ch <= u'\u9fff':
            return True
    return False


"""字符串长度检查"""


def check_by_str_length(str1, length):
    if len(str1) > length:
        if "_" in str1:
            print_error(cell_sound.value + "：描述后缀名过长，限制字符数为" + str(length))
        else:
            print_error(cell_sound.value + "：通过_切分的某个单词过长，每个单词长度不应超过10个字符，需要再拆分")


"""获取表格中的事件描述和状态所在的列"""


def get_descrip_and_status_column():
    require_module_column = None
    second_module_column = None
    require_name_column = None
    status_column = None
    if list(sheet.rows)[0]:
        for cell in list(sheet.rows)[0]:
            if cell.value:
                if '资源名称' == str(cell.value):
                    require_name_column = cell.column
                    # print(require_name_column)
                elif '资源描述' == str(cell.value):
                    status_column = cell.column
                    # print(status_column)

    return require_name_column, status_column


"""分隔符的大小写和长度检查"""


def check_by_length_and_word(name):
    # pprint(cell_sound.value)
    word_list = name.split("_")
    is_title = True
    for word in word_list:
        """判定_的每个都不能超过10个"""
        check_by_str_length(word, 10)
        """判定_的每个开头必须大写"""
        if len(word) > 0:
            # 只需要检查第一个字符，因为istitle会导致检查字符串只能首字母大写
            if word[0].istitle() == False and word[0].isdigit() == False:
                is_title = False
                break
    if is_title == False:
        print_error(name + "：通过”_“分隔的每个单词开头都需要大写")
    if word_list:
        return word_list[-1]
    else:
        return None


with WaapiClient() as client:
    """*****************Wwise功能区******************"""

    """媒体资源导入"""


    def import_media_in_track(source_path, obj_id):
        args_import = {
            # createNew
            # useExisting：会增加一个新媒体文件但旧的不会删除
            # replaceExisting:会销毁Sound，上面做的设置都无了
            "importOperation": "replaceExisting",
            "imports": [
                {
                    "audioFile": source_path,
                    "objectPath": obj_id,
                    "originalsSubFolder": sheet_name
                }
            ]
        }
        client.call("ak.wwise.core.audio.import", args_import)


    """警告捕获"""


    def print_warning(warn_info):
        print("[warning]" + warn_info)


    """查找对象"""


    def find_obj(args):
        options = {
            'return': ['name', 'id', 'path', 'notes', 'originalWavFilePath', 'type', 'parent']

        }
        obj_sub_list = client.call("ak.wwise.core.object.get", args, options=options)['return']
        if not obj_sub_list:
            obj_sub_id = ""
            obj_sub_path = ""
        else:
            obj_sub_id = obj_sub_list[0]['id']
            obj_sub_path = obj_sub_list[0]['path']
        return obj_sub_list, obj_sub_id, obj_sub_path


    """对象列表查找"""


    def find_obj_list(obj_path, obj_type):
        obj_list, _, _ = find_obj(
            {'waql': ' "%s" select descendants where type = "%s" ' % (obj_path, obj_type)})
        return obj_list


    """设置obj的notes"""


    def set_obj_notes(obj_id, notes_value):
        args = {
            'object': obj_id,
            'value': notes_value
        }
        client.call("ak.wwise.core.object.setNotes", args)
        # print_warning(obj_name + "描述更改为：" + notes_value)


    """修改Wwise内容的名字"""


    def change_name_by_wwise_content(obj_id, name, old_name, obj_type):
        args = {
            "objects": [
                {

                    "object": obj_id,
                    "name": name,
                }
            ]
        }
        client.call("ak.wwise.core.object.set", args)
        print_warning(old_name + "(" + obj_type + ")改名为：" + name)


    """通用内容创建"""


    def create_obj_content(state_group_parent_path, state_group_type,
                           state_group_name, state_group_desc, state_id):
        flag = 0
        state_group_id = ""
        state_group_path = ""
        # 查找此state_group是否存在
        state_group_list, _, _ = find_obj(
            {'waql': ' "%s" select descendants where type = "%s"' % (
                state_group_parent_path, state_group_type)})
        # pprint(state_group_list)
        # state_group已存在
        for state_group_dict in state_group_list:
            if state_group_dict['name'] == state_group_name:
                flag = 1
                # 设置notes
                set_obj_notes(state_group_dict['id'], state_group_desc)
                state_group_id = state_group_dict['id']
                state_group_path = state_group_dict['path']
                break
            # 改名了但描述一致
            elif state_group_dict['notes'] == state_group_desc:
                flag = 2
                # 改名
                change_name_by_wwise_content(state_group_dict['id'], state_group_name, state_group_dict['name'],
                                             state_group_type)
                state_group_id = state_group_dict['id']
                state_group_path = state_group_dict['path']
                break
        # state_group不存在，需要创建
        # print(flag)
        # 不存在则创建
        if flag == 0:
            if state_group_type == "Event":
                # 专属trig用
                args = {
                    # 选择父级
                    "parent": state_group_parent_path,
                    # 创建类型名称
                    "type": state_group_type,
                    "name": state_group_name,
                    "notes": state_group_desc,
                    "children": [
                        {
                            "name": "",
                            "type": "Action",
                            "@ActionType": 35,
                            "@Target": state_id
                        }]
                }
            else:
                # 创建的rnd属性
                args = {
                    # 选择父级
                    "parent": state_group_parent_path,
                    # 创建类型名称
                    "type": state_group_type,
                    "name": state_group_name,
                    "notes": state_group_desc,
                }
            state_group_object = client.call("ak.wwise.core.object.create", args)
            # print(state_group_object)
            state_group_id = state_group_object['id']
            _, _, state_group_path = find_obj(
                {'waql': ' "%s" ' % state_group_id})
            print_warning(state_group_object['name'] + ":" + state_group_type + "已创建")

        return state_group_id, state_group_path


    """音乐资源导入"""


    def import_media(mus_name, mus_track_path):
        flag = 0
        # 音乐资源导入
        for media_info_name in media_info_dict:
            if mus_name in media_info_name:
                print_warning((mus_name + ":媒体资源已导入"))
                media_path = os.path.join(py_path, media_info_dict[media_info_name])
                import_media_in_track(media_path, mus_track_path)
                flag = 1
                break
        # if flag == 0:
        #     print_error(mus_name + ":该名称的媒体资源未找到，无法导入")


    """创建新的mus"""


    def create_mus_content(mus_name, mus_desc):
        """music segment创建"""
        # 查找music segment的playlist container父级
        for music_playlist_container_dict in music_playlist_container_list:
            if music_playlist_container_dict['name'] in mus_name:
                mus_segment_id, _ = create_obj_content(music_playlist_container_dict['id'], "MusicSegment", mus_name,
                                                       mus_desc, "")
                mus_track_id, mus_track_path = create_obj_content(mus_segment_id, "MusicTrack", mus_name,
                                                                  mus_desc, "")
                import_media(mus_name, mus_track_path)


    """创建新的singer及trigger指派"""


    def create_stin_content(stin_name, stin_desc):
        stin_segment_id, _ = create_obj_content(stin_path,
                                                "MusicSegment",
                                                stin_name,
                                                stin_desc, "")
        stin_track_id, stin_track_path = create_obj_content(stin_segment_id,
                                                            "MusicTrack",
                                                            stin_name,
                                                            stin_desc, "")
        import_media(cell_sound.value, stin_track_path)

        # Trig创建及指派
        trig_name = stin_name.replace("Stin", "Trig")
        trig_id, create_trig_path = create_obj_content(trig_path,
                                                       "Trigger",
                                                       trig_name,
                                                       stin_desc, "")
        """Trigger应该是直接在音乐中设置，不需要创建Event"""
        # # Event创建
        # trig_event_name = 'AKE_Set_' + trig_name
        # trig_event, _ = create_obj_content(trig_event_path,
        #                                    "Event",
        #                                    trig_event_name,
        #                                    stin_desc, trig_id)
        #
        # # Trigger指派
        # # 23才支持


    """*****************主程序处理******************"""
    # 撤销开始
    client.call("ak.wwise.core.undo.beginGroup")

    # 记录所有资源名称
    sound_name_list = []

    # 获取所有State名称列表
    state_list = find_obj_list(state_path, "State")
    state_name_list = get_one_value_list(state_list, 'name')
    # pprint(state_name_list)

    # 获取所有Music Playlist Container
    music_playlist_container_list = find_obj_list(music_path, "MusicPlaylistContainer")
    music_playlist_container_name_list = get_one_value_list(music_playlist_container_list, 'name')

    # 获取new media的信息
    media_info_dict, _ = get_type_file_name_and_path('.wav', 'New_Media')
    # pprint(media_info)

    # pprint(music_playlist_container_name_list)

    # 提取规则：只提取xlsx文件
    for i in file_name_list:
        if ".xlsx" in i:
            # 拼接xlsx的路径
            file_path_xlsx = os.path.join(py_path, i)
            # 获取xlsx的workbook
            wb = openpyxl.load_workbook(file_path_xlsx)
            # 获取xlsx的所有sheet
            sheet_names = wb.sheetnames
            # 加载所有工作表
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                value, value_desc = get_descrip_and_status_column()
                # 获取工作表第一行数据
                for cell in list(sheet.rows)[0]:
                    if '资源名称' == str(cell.value):
                        # 获取音效名下的内容
                        for cell_sound in list(sheet.columns)[cell.column - 1]:
                            # 空格和中文不检测
                            if cell_sound.value:
                                if not check_is_chinese(cell_sound.value):

                                    """❤❤❤❤数据获取❤❤❤❤"""
                                    """检测表格内容"""
                                    is_pass = True

                                    """检查表格中是否有内容重复项"""
                                    if cell_sound.value in sound_name_list:
                                        print_error(cell_sound.value + "：表格中有重复项Group名称 ，请检查")
                                    else:
                                        sound_name_list.append(cell_sound.value)

                                        """❤❤❤通用内容检查❤❤❤"""
                                        """资源名称"""
                                        value = cell_sound.value
                                        mus_name_list.append(value)

                                        """资源描述"""
                                        value_desc_value = sheet.cell(row=cell_sound.row,
                                                                      column=value_desc).value
                                        """资源描述检查"""
                                        if value_desc_value:
                                            if value_desc_value not in value_desc_list:
                                                value_desc_list.append(value_desc_value)
                                            else:
                                                print_error(value_desc_value + "：表格中有重复项描述，请检查")
                                        else:
                                            print_error(cell_sound.value + "：描述不能为空，请补充")

                                        """资源名称检查"""
                                        # 分隔符的大小写和长度检查，取后缀名
                                        suffix_name = check_by_length_and_word(value)

                                        """检查是否为分轨"""
                                        is_subtrack = False

                                        if sheet_name == "Mus":
                                            """❤❤❤Mus资源名称检查❤❤❤"""
                                            # 检查是否以Mus_开头
                                            check_by_re(value, r'^Mus_(.*)')
                                            if suffix_name not in state_name_list:
                                                is_subtrack = True
                                                # 分轨的话需要把后缀去掉
                                                value = value.replace("_" + suffix_name, "")

                                            # 检查播放列表里是否有这个名字
                                            flag = 0
                                            for music_container_name in music_playlist_container_name_list:
                                                if music_container_name in value:
                                                    flag = 1
                                                    break
                                            if flag == 0:
                                                print_error(
                                                    cell_sound.value + "：没有相应的MusicPlaylistContainer，请在Wwise中创建")

                                            if is_pass:
                                                create_mus_content(cell_sound.value,
                                                                   value_desc_value)

                                            """❤❤❤Stin资源名称检查❤❤❤"""
                                        elif sheet_name == "Stin":
                                            check_by_re(value, r'^Stin_(.*)')
                                            if is_pass:
                                                create_stin_content(cell_sound.value, value_desc_value)

    """******************对象删除清理********************"""
    """对象删除"""


    def delete_obj(obj_id, obj_name, obj_type):
        args = {
            "object": "%s" % obj_id
        }
        client.call("ak.wwise.core.object.delete", args)
        print_warning(obj_name + "(" + obj_type + ")" + ":已删除")


    """删除状态引用的事件"""


    def delete_state_refer_event(state_id):
        args = {
            'waql': '"%s" select referencesTo' % state_id
        }
        refer_list, _, _ = find_obj(args)
        # print(refer_list)
        if refer_list:
            # pprint(refer_list)
            for refer_dict in refer_list:
                if 'parent' in refer_dict:
                    if 'AKE_Set_' in refer_dict['parent']['name']:
                        # print(refer_dict['parent'])
                        delete_obj(refer_dict['parent']['id'], refer_dict['parent']['name'], "Event")


    """内容删除"""


    def delete_state(wwise_obj_list, excel_list, obj_type, trig_list):
        for wwise_obj_dict in wwise_obj_list:
            # pprint(wwise_obj_dict['name'])
            if wwise_obj_dict['name'] != "None":
                if wwise_obj_dict['name'] not in excel_list:
                    # Segment删除
                    delete_obj(wwise_obj_dict['id'], wwise_obj_dict['name'], obj_type)
                    original_path_media = os.path.join(original_path, "Stin", wwise_obj_dict['name'], '.wav')
                    # print(original_path_media)
                    # 媒体资源删除
                    original_path_media = ""
                    if "Stin" in wwise_obj_dict['name']:
                        original_path_media = os.path.join(original_path, "Stin", wwise_obj_dict['name'] + '.wav')
                        # Trig删除
                        for trig_dict in trig_list:
                            trig_name = wwise_obj_dict['name'].replace("Stin", "Trig")
                            if trig_dict['name'] == trig_name:
                                delete_obj(trig_dict['id'], trig_name, "Trigger")

                    elif "Mus" in wwise_obj_dict['name']:
                        original_path_media = os.path.join(original_path, "Mus", wwise_obj_dict['name'] + '.wav')
                    if os.path.isfile(original_path_media):
                        os.remove(original_path_media)
                        print_warning("[文件清理]" + original_path_media + "已删除")


    """同步表中删除的内容"""

    # Wwise中的内容列表获取
    mus_segment_list = find_obj_list(music_path, "MusicSegment")
    stin_segment_list = find_obj_list(stin_path, "MusicSegment")
    trig_list = find_obj_list(trig_path, "Trigger")

    # # 查找state/switch，再跟表格比对有没有，没有就删除资源及事件引用
    delete_state(mus_segment_list, mus_name_list, "MusicSegment",trig_list)

    # 撤销结束
    client.call("ak.wwise.core.undo.endGroup", displayName="rnd创建撤销")

    # 清除复制的媒体资源
    shutil.rmtree("New_Media")
    os.mkdir("New_Media")

# 以Music Clip为最小单位，从Music Clip找有没有表中的资源，没有才生成
