from waapi import WaapiClient, CannotConnectToWaapiException
import openpyxl
import json
from os.path import abspath, dirname
import os
from openpyxl.styles import Alignment
import sys


# 打开工作簿并获取工作表
def excel_get_sheet(path, sheetname):
    # 打开工作簿
    wb = openpyxl.load_workbook(path)
    # print(wb)
    # 获取工作表
    sheet = wb[sheetname]
    # print(sheet)
    return sheet, wb


# 文件所在目录
py_path = ""
if hasattr(sys, 'frozen'):
    py_path = dirname(sys.executable)
elif __file__:
    py_path = dirname(abspath(__file__))

# 获取excel文件路径
excel_name = '生成的wem文件查询.xlsx'
excel_path = os.path.join(py_path, excel_name)
# 获取json文件路径
json_name = 'GenerateSoundbank_Log.json'
json_path = os.path.join(py_path, json_name)

# 创建openpyxl对齐方式：左对齐
align = Alignment(horizontal='left')

# 清空excel表的内容
sheet, wb = excel_get_sheet(excel_path, 'Sheet1')
for cell in list(sheet.columns)[0]:
    if cell.row != 1:
        sheet.delete_rows(cell.row)

wem_id_list = []

# Python 的异常处理 try…except…else… 语句
try:
    client = WaapiClient()

except CannotConnectToWaapiException:
    print("Could not connect to Waapi: Is Wwise running and Wwise Authoring API enabled?")

else:

    # 建立 on_name_changed() 准备在订阅中作为回调函数，用来接收字典形式的返回参数
    def soundbank_generated(*args, **kwargs):

        bankinfo_list = kwargs.get("bankInfo")
        # print("*****生成的单个SoundBank*****")
        # pprint(bankinfo_list)

        for bankinfo_dict in bankinfo_list:
            # print("*****生成的单个BankInfo*****")
            # pprint(bankinfo_dict)
            if 'Media' in bankinfo_dict:
                for media_dict in bankinfo_dict['Media']:
                    if media_dict['Id'] not in wem_id_list:
                        wem_id_list.append(media_dict['Id'])
                        insert_row = sheet.max_row + 1
                        sheet.cell(row=insert_row, column=1).value = int(media_dict['Id'])
                        sheet.cell(row=insert_row, column=1).alignment = align
                        sheet.cell(row=insert_row, column=2).value = media_dict['ShortName']

            # print("*****生成的单个BankInfo结束*****")

        # 数据解析
        # bankinfo_wem = kwargs.get("bankInfo", {}).get("Media")

        # print("*****生成的单个SoundBank结束*****\n")

        # 执行完成后断开 WAMP 连接，当然，要是想一直监控信息也可以不断开
        client.disconnect()


    # 订阅所需主题，传入回调函数，使用选项 type 以让名称修改时传回的字典里直接有被修改的对象类型
    new_data = client.subscribe("ak.wwise.core.soundbank.generated", soundbank_generated, infoFile=True)

    args = {
        "soundbanks": [
            {"name": "AKE_Play_Mus_Global"}
        ],
        "writeToDisk": True,
        # "clearAudioFileCache": True
    }

    gen_log = client.call("ak.wwise.core.soundbank.generate", args)
    # pprint(gen_log)

    # 将log写入json
    gen_log_json = json.dumps(gen_log, indent=4)
    with open(json_path, 'w') as file:
        file.write(gen_log_json)

# wb.save(excel_path)
# os.system(excel_path)
