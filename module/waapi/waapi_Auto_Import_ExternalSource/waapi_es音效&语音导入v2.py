# xml读取库
import shutil
import pandas as pd
import os
from waapi import WaapiClient
from pprint import pprint
import sys
from os.path import abspath, dirname
import openpyxl
import re

# xml写入库
from xml.dom.minidom import Document, parse

# 自定义库
import module.excel.excel_h as excel_h
import module.config as config
import module.oi.oi_h as oi_h
import module.waapi.waapi_h as waapi_h
import module.csv.csv_h as csv_h

"""****************数据获取******************"""

# 获取文件所在目录
py_path = ""
if hasattr(sys, 'frozen'):
    py_path = dirname(sys.executable)
elif __file__:
    py_path = dirname(abspath(__file__))

# 获取.csv文件
csv_mediainfo_path = os.path.join(py_path, config.csv_mediainfo_path)
csv_wwisecookie_path = os.path.join(py_path, config.csv_wwisecookie_path)
csv_DT_AudioPlotInfo_path = os.path.join(py_path, config.csv_DT_AudioPlotInfo_path)
csv_DT_AudioPlotSoundInfo_path = os.path.join(py_path, config.csv_DT_AudioPlotSoundInfo_path)

# 获取.xlsx文件
excel_mediainfo_path = os.path.join(py_path, config.excel_mediainfo_path)
excel_wwisecookie_path = os.path.join(py_path, config.excel_wwisecookie_path)
excel_DT_AudioPlotInfo_path = os.path.join(py_path, config.excel_DT_AudioPlotInfo_path)
excel_DT_AudioPlotSoundInfo_path = os.path.join(py_path, config.excel_DT_AudioPlotSoundInfo_path)

# 获取相应的excel表信息
sheet_mediainfo, wb_mediainfo = excel_h.excel_get_sheet(excel_mediainfo_path, 'Sheet1')
sheet_wwisecookie, wb_wwisecookie = excel_h.excel_get_sheet(excel_wwisecookie_path, 'Sheet1')
sheet_DT_AudioPlotInfo, wb_DT_AudioPlotInfo = excel_h.excel_get_sheet(excel_DT_AudioPlotInfo_path, 'Sheet1')
sheet_DT_AudioPlotSoundInfo, wb_DT_AudioPlotSoundInfo = excel_h.excel_get_sheet(excel_DT_AudioPlotSoundInfo_path,
                                                                                'Sheet1')

# 获取媒体资源文件列表
wav_path = os.path.join(py_path, "New_Media")

# excel表标题列
excel_es_title_list = config.excel_es_title_list

# 获取当前目录下要遍历的表路径列表
file_name_list = excel_h.excel_get_path_list(os.path.join(py_path, "Excel"))

# 标题列索引获取
mediainfo_title_dict = excel_h.excel_get_all_sheet_title_column(sheet_mediainfo)
wwisecookie_title_dict = excel_h.excel_get_all_sheet_title_column(sheet_wwisecookie)
DT_AudioPlotInfo_title_dict = excel_h.excel_get_all_sheet_title_column(sheet_DT_AudioPlotInfo)
DT_AudioPlotSoundInfo_dict = excel_h.excel_get_all_sheet_title_column(sheet_DT_AudioPlotSoundInfo)

# 复制列
copy_wwisecookie_list = [wwisecookie_title_dict['MediaInfoId'], wwisecookie_title_dict['MediaInfoId'],
                         wwisecookie_title_dict['ExternalSourceName'], wwisecookie_title_dict['MediaInfoId']]
copy_DT_AudioPlotInfo_list = [DT_AudioPlotSoundInfo_dict['Row'], DT_AudioPlotSoundInfo_dict['PlotID'],
                              DT_AudioPlotSoundInfo_dict['Extern Source Name'],
                              DT_AudioPlotSoundInfo_dict['Extern Source ID']]

copy_DT_AudioPlotSoundInfo_list = copy_DT_AudioPlotInfo_list

_, wem_list = oi_h.get_type_file_name_and_path('.wem', config.external_output_path)

"""**************写xml**************"""
# 创建一个对象
doc = Document()
# 创建根节点
root = doc.createElement("ExternalSourcesList")
root.setAttribute("SchemaVersion", "1")

# 如果保存在Wwise工程下，此Root为Wwise工程的相对路径，则不需要输入绝对路径，仅需输入.wav的名字
# 否则取文件的绝对路径
# root.setAttribute("Root", "UngeneratedExternalSources")
# 添加载Document对象上
doc.appendChild(root)

"""*****************功能检测区******************"""

"""xml创建元素"""


def xml_create_element(media_path):
    source = doc.createElement("Source")
    source.setAttribute("Path", media_path)
    if "CG" in media_path:
        source.setAttribute("Conversion", "CG")
    elif "VO" in media_path:
        source.setAttribute("Conversion", "VO")
    root.appendChild(source)


"""写入media_info excel表"""


def write_media_info_excel(vo_id, cell_sound):
    flag = 0
    for cell in list(sheet_mediainfo.columns)[0]:
        # 在media_info表中找到该ID，则替换
        if sheet_mediainfo.cell(row=cell.row, column=1).value == vo_id:
            flag = 1
            sheet_mediainfo.cell(row=cell.row,
                                 column=mediainfo_title_dict['MediaName']).value = cell_sound.value + ".wem"

    # 在media_info表中未找到该ID，则新增
    if flag == 0:
        # 插入为空的行
        insert_row = sheet_mediainfo.max_row + 1
        value_dict = {
            "Name": vo_id,
            'ExternalSourceMediaInfoId': vo_id,
            'MediaName': cell_sound.value + ".wem",
            'CodecID': 4,
            'bIsStreamed': 'TRUE',
            'bUseDeviceMemory': "FALSE",
            'MemoryAlignment': 0,
            'PrefetchSize': 0
        }
        for cell in list(sheet_mediainfo.rows)[0]:
            if cell.value in value_dict:
                sheet_mediainfo.cell(row=insert_row, column=cell.column).value = value_dict[cell.value]


"""写入wwise_cookie excel表"""


def write_wwise_cookie_excel(vo_id, cell_sound, external_sound_dict):
    flag = 0
    for cell in list(sheet_wwisecookie.columns)[0]:
        # 在media_info表中找到该ID，则替换
        if sheet_wwisecookie.cell(row=cell.row, column=1).value == vo_id:
            flag = 1
            sheet_wwisecookie.cell(row=cell.row,
                                   column=wwisecookie_title_dict["MediaName"]).value = cell_sound.value + ".wem"
            sheet_wwisecookie.cell(row=cell.row, column=wwisecookie_title_dict["ExternalSourceCookie"]).value = \
                external_sound_dict['shortId']
            sheet_wwisecookie.cell(row=cell.row, column=wwisecookie_title_dict["ExternalSourceName"]).value = \
                external_sound_dict['name']

    # 在media_info表中未找到该ID，则新增
    if flag == 0:
        # 插入为空的行
        insert_row = sheet_wwisecookie.max_row + 1
        value_dict = {
            "Name": vo_id,
            'ExternalSourceCookie': external_sound_dict['shortId'],
            'ExternalSourceName': external_sound_dict['name'],
            'MediaInfoId': vo_id,
            'MediaName': cell_sound.value + ".wem"
        }
        for cell in list(sheet_wwisecookie.rows)[0]:
            if cell.value in value_dict:
                sheet_wwisecookie.cell(row=insert_row, column=cell.column).value = value_dict[cell.value]


"""删除相应的.wem文件"""


def delete_cancel_wem(media_name):
    # [{'VO_C01_15_World.wem': 'S:\\Ver_1.0.0\\Project\\Content\\Audio\\GeneratedExternalSources\\Android\\VO_C01_15_World.wem'},
    #  {'VO_C01_33_Battle.wem': 'S:\\Ver_1.0.0\\Project\\Content\\Audio\\GeneratedExternalSources\\Android\\VO_C01_33_Battle.wem'},
    #  {'VO_C02_06_Menu.wem': 'S:\\Ver_1.0.0\\Project\\Content\\Audio\\GeneratedExternalSources\\Android\\VO_C02_06_Menu.wem'},
    #  {'VO_C02_33_Battle.wem': 'S:\\Ver_1.0.0\\Project\\Content\\Audio\\GeneratedExternalSources\\Android\\VO_C02_33_Battle.wem'},]
    for wem_dict in wem_list:
        for wem_name in wem_dict:
            # print(wem_name)
            if media_name in wem_name:
                if os.path.isfile(wem_dict[wem_name]):
                    os.remove(wem_dict[wem_name])
                    oi_h.print_warning("wem文件删除：" + str(wem_dict[wem_name]))


"""移除标记为cancel的所有内容"""


def delete_cancel_content(table_name_list):
    # pprint(table_name_list)
    # 删除excel表的相应内容

    for row in range(2, sheet_mediainfo.max_row + 1):
        cell = sheet_mediainfo.cell(row=row, column=mediainfo_title_dict['MediaName'])
        # pprint(cell.value)
        flag = 0
        for table_name in table_name_list:
            # 在media_info表中找到该ID，则删除
            if table_name in cell.value:
                flag = 1
                break
        if flag == 0:
            sheet_mediainfo.delete_rows(cell.row)
            sheet_wwisecookie.delete_rows(cell.row)
            oi_h.print_warning(cell.value + ":在mediainfo和wwisecookie中删除相关信息")
            delete_cancel_wem(cell.value)


"""检查是否以CG_External_或VO_External_开头"""


# 返回值为True或False

def check_prefix(string):
    pattern = r'^(CG_External_|VO_External_)'
    return re.match(pattern, string) is not None


"""资源命名规范检查"""


def check_name():
    table_name_list = []
    is_pass = True
    for i in file_name_list:
        if (".xlsx" in i) and ("MediaInfoTable" not in i) and ("ExternalSourceDefaultMedia" not in i):
            # 拼接xlsx的路径
            file_path_xlsx = os.path.join(os.path.join(py_path, "Excel"), i)
            # 获取xlsx的workbook
            wb = openpyxl.load_workbook(file_path_xlsx)
            # 获取xlsx的所有sheet
            sheet_names = wb.sheetnames
            # 加载所有工作表
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]

                title_colunmn_dict = excel_h.excel_get_sheet_title_column(sheet, excel_es_title_list)
                if title_colunmn_dict:
                    # 获取音效名下的内容
                    for cell_sound in list(sheet.columns)[
                        excel_h.sheet_title_column("文件名", title_colunmn_dict) - 1]:
                        if cell_sound.value and (cell_sound.value != "文件名"):
                            """资源命名规范检查"""
                            check_list = [

                            ]
                            _, _, is_pass = oi_h.check_name_all(cell_sound.value)
                            if is_pass:
                                # 检查是否以CG_External_或VO_External_开头
                                if not check_prefix(cell_sound.value):
                                    is_pass = oi_h.print_error(
                                        cell_sound.value + "：请检查名称是否以CG_External_或VO_External_开头")
                                if is_pass and sheet.cell(row=cell_sound.row,
                                                          column=excel_h.sheet_title_column("State",
                                                                                            title_colunmn_dict)).value != 'cancel':
                                    if cell_sound.value not in table_name_list:
                                        table_name_list.append(cell_sound.value)
                                    else:
                                        is_pass = oi_h.print_error(
                                            cell_sound.value + "：在表中有重复，请检查")
    # 文件清理
    delete_cancel_content(table_name_list)
    # pprint(table_name_list)

    return is_pass


"""自动化生成ES数据"""


def auto_gen_es_file(file_wav_dict):
    # 查找要导入的媒体文件里有没有对应的
    for file_wav_name in file_wav_dict:
        flag = 0
        # 读excel表
        # 提取规则：只提取xlsx文件
        for i in file_name_list:
            # print(i)
            if (".xlsx" in i) and ("MediaInfoTable" not in i) and ("ExternalSourceDefaultMedia" not in i):
                # 拼接xlsx的路径
                file_path_xlsx = os.path.join(os.path.join(py_path, "Excel"), i)
                # 获取xlsx的workbook
                wb = openpyxl.load_workbook(file_path_xlsx)
                # 获取xlsx的所有sheet
                sheet_names = wb.sheetnames
                # 加载所有工作表
                for sheet_name in sheet_names:
                    sheet = wb[sheet_name]

                    title_colunmn_dict = excel_h.excel_get_sheet_title_column(sheet, excel_es_title_list)
                    if title_colunmn_dict:
                        # 获取音效名下的内容
                        for cell_sound in list(sheet.columns)[
                            excel_h.sheet_title_column("文件名", title_colunmn_dict) - 1]:
                            if cell_sound.value and (cell_sound.value != "文件名"):

                                """查找文件"""
                                if cell_sound.value in file_wav_name:
                                    flag = 1
                                    if sheet.cell(row=cell_sound.row, column=excel_h.sheet_title_column("State",
                                                                                                        title_colunmn_dict)).value != 'cancel':
                                        # print(file_wav_dict[file_wav_name])
                                        # 在表格中找到此音效才将其写入xml
                                        xml_create_element(file_wav_dict[file_wav_name])

                                        # 在状态处自动标记完成
                                        sheet.cell(row=cell_sound.row,
                                                   column=excel_h.sheet_title_column("State",
                                                                                     title_colunmn_dict)).value = 'done'
                                        is_wwise_have = 0
                                        # 查找Wwise中有无相应的Sound
                                        for external_sound_dict in external_sound_list:
                                            if external_sound_dict['parent']['name'] == sheet.cell(
                                                    row=cell_sound.row,
                                                    column=excel_h.sheet_title_column("External_Type",
                                                                                      title_colunmn_dict)).value:
                                                # pprint(external_sound_dict['shortId'])
                                                # 查找mediainfo表的id有没有，没有则添加，有则修改内容
                                                vo_id = excel_h.create_id(config.es_id_config, i, sheet_mediainfo)

                                                # 写入media_info excel表
                                                write_media_info_excel(vo_id, cell_sound)
                                                # # 写入wwise_cookie excel表
                                                write_wwise_cookie_excel(vo_id, cell_sound, external_sound_dict)

                                                oi_h.print_warning(file_wav_name + "：已导入")
                                                is_wwise_have = 1
                                                break
                                        if is_wwise_have == 0:
                                            oi_h.print_error(
                                                file_wav_name + "：在Wwise中未建立相应容器，请检查填写是否正确或Wwise中容器是否存在")

                                    break

                wb.save(file_path_xlsx)

        if flag == 0:
            oi_h.print_error(file_wav_name + "：在语音需求表中不存在，请检查名称是否正确或在表格中补充该名字")


with WaapiClient() as client:
    """自动生成externalsource"""


    def gen_external(language):
        if language == "SFX":
            language = ""
        args = {
            "sources": [
                {
                    "input": config.external_input_path,
                    "platform": "Windows",
                    "output": os.path.join(config.external_output_win_path, language)
                },
                {
                    "input": config.external_input_path,
                    "platform": "Android",
                    "output": os.path.join(config.external_output_android_path, language)
                },
                {
                    "input": config.external_input_path,
                    "platform": "iOS",
                    "output": os.path.join(config.external_output_ios_path, language)
                }
            ]
        }

        gen_log = client.call("ak.wwise.core.soundbank.convertExternalSources", args)


    # 查找External下的VO
    obj_list = client.call("ak.wwise.core.object.get",
                           waapi_h.waql_by_type("ExternalSource", config.vo_external_path),
                           options=config.options)['return']
    external_vo_list, _, _ = waapi_h.find_obj(obj_list)
    # pprint(external_sound_list)

    # 查找External下的CG
    obj_list = client.call("ak.wwise.core.object.get",
                           waapi_h.waql_by_type("ExternalSource", config.cg_external_path),
                           options=config.options)['return']
    external_sfx_list, _, _ = waapi_h.find_obj(obj_list)
    # pprint(external_cg_list)

    external_sound_list = external_vo_list + external_sfx_list

    """*******************主程序*******************"""

    # 命名规范检查
    if check_name():
        # 语音ES生成
        for language in config.language_list:
            wav_language_path = os.path.join(py_path, "New_Media", language)
            file_wav_language_dict, _ = oi_h.get_type_file_name_and_path('.wav', wav_language_path)
            auto_gen_es_file(file_wav_language_dict)
            # # xml文件写入
            # pprint(file_wav_dict)
            with open(config.es_xml_path, 'w+') as f:
                # 按照格式写入
                f.write(doc.toprettyxml())
                f.close()
            # pprint(doc.toprettyxml())
            # 复制xml为wsources
            shutil.copy2(os.path.join(py_path, config.es_xml_path),
                         os.path.join(py_path, config.es_wsources_path))

            gen_external(language)

            # 初始化xml
            doc = Document()
            root = doc.createElement("ExternalSourcesList")
            root.setAttribute("SchemaVersion", "1")
            doc.appendChild(root)

            # 将删除的xml内容写入wsources
            shutil.copy2(os.path.join(py_path, config.es_xml_path),
                         os.path.join(py_path, config.es_wsources_path))

        # 删除xml的内容
        doc = parse(config.es_xml_path)
        parent_node = doc.getElementsByTagName('ExternalSourcesList')[0]
        while parent_node.hasChildNodes():
            parent_node.removeChild(parent_node.firstChild)

        # 保存excel表的信息
        wb_mediainfo.save(excel_mediainfo_path)
        wb_wwisecookie.save(excel_wwisecookie_path)

        # 复制资源表的信息到剧情ID表
        # 遍历每一行并复制符合条件的行
        for row in sheet_wwisecookie.iter_rows(min_row=1, max_row=sheet_wwisecookie.max_row):
            # 检查第 1 列的值是否在 1 到 50000 之间
            first_col_value = row[0].value
            if isinstance(first_col_value, (int, float)) and 1 <= first_col_value <= 50000:
                for src_col, tgt_col in zip(copy_wwisecookie_list, copy_DT_AudioPlotInfo_list):
                    sheet_DT_AudioPlotSoundInfo.cell(row=row[0].row, column=tgt_col).value = row[src_col - 1].value

        # 保存目标 Excel 文件
        wb_DT_AudioPlotSoundInfo.save(excel_DT_AudioPlotSoundInfo_path)
        wb_DT_AudioPlotInfo.save(excel_DT_AudioPlotInfo_path)

        # 将excel转为csv
        # 指定CSV文件名和编码格式
        csv_h.excel_to_csv(excel_mediainfo_path, csv_mediainfo_path)
        csv_h.excel_to_csv(excel_wwisecookie_path, csv_wwisecookie_path)
        csv_h.excel_to_csv(excel_DT_AudioPlotInfo_path, csv_DT_AudioPlotInfo_path)
        csv_h.excel_to_csv(excel_DT_AudioPlotSoundInfo_path, csv_DT_AudioPlotSoundInfo_path)

    # else:
    #     print(1)

    # # 清除复制的媒体资源
    shutil.rmtree("New_Media")
    os.mkdir("New_Media")
    os.mkdir("New_Media/Chinese")
    os.mkdir("New_Media/English")
    os.mkdir("New_Media/Japanese")
    os.mkdir("New_Media/Korean")
    os.mkdir("New_Media/SFX")
