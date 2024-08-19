# xml读取库
import shutil
import pandas as pd
import sys
from os.path import abspath, dirname
import os
import warnings
import openpyxl
from waapi import WaapiClient
# xml写入库
from xml.dom.minidom import Document

"""获取文件所在目录"""
py_path = ""
if hasattr(sys, 'frozen'):
    py_path = dirname(sys.executable)
elif __file__:
    py_path = dirname(abspath(__file__))

"""*****************可变配置******************"""
vo_external_path = "\\Actor-Mixer Hierarchy\\VO"
# 获取.csv文件
media_info_path = os.path.join(py_path, "MediaInfoTable.csv")
external_cookie_path = os.path.join(py_path, "ExternalSourceDefaultMedia.csv")
# external的输入输出路径
external_input_path = "F:\\pppppy\\SP\\module\\waapi\\waapi_Auto_Import_ExternalSource\\ExternalSource.wsources"
external_output_win_path = "F:\\SP\\UE53AndWwise2218\\UE53AndWwise2218\\Content\\GeneratedExternalSources\\Windows"
external_output_mac_path = "F:\\SP\\UE53AndWwise2218\\UE53AndWwise2218\\Content\\GeneratedExternalSources\\Mac"
external_output_path = "F:\\SP\\UE53AndWwise2218\\UE53AndWwise2218\\Content\\GeneratedExternalSources"

"""**************写xml**************"""
# 创建一个对象
doc = Document()
# 创建根节点
root = doc.createElement("ExternalSourcesList")
root.setAttribute("SchemaVersion", "1")

# 如果保存在Wwise工程下，此Root为Wwise工程的相对路径，则不需要输入绝对路径，仅需输入.wav的名字
# 否则取文件的绝对路径
# root.setAttribute("Root", "UngeneratedExternalSources")
# 添加载Document对象上
doc.appendChild(root)

"""获取.xlsx文件"""
file_names = []
for i in os.walk(py_path):
    file_names.append(i)
# pprint("输出文件夹下的文件名：")
file_name_list = file_names[0][2]

"""读csv"""
media_info_data = pd.read_csv(media_info_path)
external_cookie_data = pd.read_csv(external_cookie_path)
# 报错消除
warnings.simplefilter(action='ignore', category=FutureWarning)
# 新加数据
media_info_initial_row = media_info_data.shape[0]
media_info_row_max = media_info_data.shape[0] + 1
external_cookie_initial_row = external_cookie_data.shape[0]
external_cookie_row_max = external_cookie_data.shape[0] + 1

is_pass = True

"""*****************功能检测区******************"""

"""报错捕获"""


def print_warning(warn_info):
    print("[warning]" + warn_info)


def print_error(error_info):
    global is_pass
    is_pass = False
    print("[error]" + error_info)


"""遍历文件目录获取文件名称和路径"""


def get_type_file_name_and_path(file_type, dir_path):
    file_dict = {}
    file_list = []
    # 遍历文件夹下的所有子文件
    # 绝对路径，子文件夹，文件名
    for root, dirs, files in os.walk(dir_path):
        # {'name': ['1111.akd', 'Creature Growls 4.akd', 'Sonic Salute_005_2_20.akd'],
        #  'path': 'S:\\chen.gong_DCC_Audio\\Audio\\SilverPalace_WwiseProject\\Originals\\SFX'}
        for file in files:
            if file_type in file:
                new_dict = {}
                file_dict[file] = os.path.join(
                    root, file)
                new_dict[file] = os.path.join(
                    root, file)
                file_list.append(new_dict)

    return file_dict, file_list


"""创建子节点"""


def xml_create_element(media_path):
    source = doc.createElement("Source")
    source.setAttribute("Path", media_path)
    source.setAttribute("Conversion", "VO")
    root.appendChild(source)


"""获取语音需求表每列的内容"""


def get_vo_excel_column(sheet):
    # 语音ID
    vo_id_column = None
    # 文件名
    file_name_column = None
    external_type_column = None
    state_column = None
    for cell in list(sheet.rows)[0]:
        if '文件名' in str(cell.value):
            file_name_column = cell.column
        elif '语音ID' in str(cell.value):
            vo_id_column = cell.column
        elif 'External_Type' in str(cell.value):
            external_type_column = cell.column
        elif 'State' in str(cell.value):
            state_column = cell.column
    return vo_id_column, file_name_column, external_type_column, state_column


"""写入media_info csv表"""


def write_media_info_csv():
    # 在media_info表中未找到该ID，则新增
    if media_info_data[media_info_data['ExternalSourceMediaInfoId'] == vo_id].empty:
        # 作初始化操作
        if media_info_initial_row == 0:
            new_data = pd.DataFrame([{
                "Name": vo_id,
                'ExternalSourceMediaInfoId': vo_id,
                'MediaName': cell_sound.value + ".wem",
                'CodecID': 4,
                'bIsStreamed': 'TRUE',
                'bUseDeviceMemory': "FALSE",
                'MemoryAlignment': 0,
                'PrefetchSize': 0
            }])
            new_data.to_csv(media_info_path, mode='a', header=False,
                            index=False)

        # 作新增操作
        else:
            global media_info_row_max
            new_data = [vo_id,
                        vo_id,
                        cell_sound.value + ".wem",
                        4,
                        'TRUE',
                        "FALSE",
                        0,
                        0
                        ]
            media_info_data.loc[media_info_row_max] = new_data
            media_info_data.to_csv(media_info_path, mode='a', header=False,
                                   index=False)
            media_info_row_max = media_info_row_max + 1

    # 在media_info表中找到该ID，则替换
    else:
        media_info_data[media_info_data['ExternalSourceMediaInfoId'] == vo_id] = [vo_id,
                                                                                  vo_id, cell_sound.value + ".wem", 4,
                                                                                  'TRUE', "FALSE", 0, 0]
        media_info_data.to_csv(media_info_path, index=False)


"""写入wwise_cookie csv表"""


def write_external_cookie_csv():
    # 在media_info表中未找到该ID，则新增
    if external_cookie_data[external_cookie_data['MediaInfoId'] == vo_id].empty:
        # 作初始化操作
        if external_cookie_initial_row == 0:
            new_data = pd.DataFrame({
                "Name": vo_id,
                'ExternalSourceCookie': external_sound_dict['shortId'],
                'ExternalSourceName': external_sound_dict['name'],
                'MediaInfoId': vo_id,
                'MediaName': cell_sound.value + ".wem"
            }, index=[0])
            # index无效，但不加会报错
            new_data.to_csv(external_cookie_path, mode='a+', header=False,
                            index=False)
            # 作新增操作
        else:
            global external_cookie_row_max
            new_data = [vo_id,
                        external_sound_dict['shortId'],
                        external_sound_dict['name'],
                        vo_id,
                        cell_sound.value + ".wem"
                        ]
            external_cookie_data.loc[external_cookie_row_max] = new_data
            external_cookie_data.to_csv(external_cookie_path, mode='a', header=False,
                                        index=False)
            external_cookie_row_max = external_cookie_row_max + 1

    # 在media_info表中找到该ID，则替换
    else:
        external_cookie_data[external_cookie_data['MediaInfoId'] == vo_id] = [vo_id,
                                                                              external_sound_dict['shortId'],
                                                                              external_sound_dict['name'],
                                                                              vo_id,
                                                                              cell_sound.value + ".wem"]
        external_cookie_data.to_csv(external_cookie_path, index=False)


"""删除相应的.wem文件"""


def delete_cancel_wem(media_name):
    _, wem_list = get_type_file_name_and_path('.wem', external_output_path)
    # [{'VO_C01_15_World.wem': 'S:\\Ver_1.0.0\\Project\\Content\\Audio\\GeneratedExternalSources\\Android\\VO_C01_15_World.wem'},
    #  {'VO_C01_33_Battle.wem': 'S:\\Ver_1.0.0\\Project\\Content\\Audio\\GeneratedExternalSources\\Android\\VO_C01_33_Battle.wem'},
    #  {'VO_C02_06_Menu.wem': 'S:\\Ver_1.0.0\\Project\\Content\\Audio\\GeneratedExternalSources\\Android\\VO_C02_06_Menu.wem'},
    #  {'VO_C02_33_Battle.wem': 'S:\\Ver_1.0.0\\Project\\Content\\Audio\\GeneratedExternalSources\\Android\\VO_C02_33_Battle.wem'},]
    for wem_dict in wem_list:
        for wem_name in wem_dict:
            # print(wem_name)
            if media_name in wem_name:
                if os.path.isfile(wem_dict[wem_name]):
                    os.remove(wem_dict[wem_name])
                    print_warning("wem文件删除：" + str(wem_dict[wem_name]))


"""移除标记为cancel的行"""


def drop_row_by_cancel():
    # 移除空白行
    media_info_data.dropna(how='all', inplace=True)
    media_info_data.to_csv('MediaInfoTable.csv', index=False)
    external_cookie_data.dropna(how='all', inplace=True)
    external_cookie_data.to_csv('MediaInfoTable.csv', index=False)

    # 读excel表
    # 提取规则：只提取xlsx文件
    for i in file_name_list:
        if ".xlsx" in i:
            # 拼接xlsx的路径
            file_path_xlsx = os.path.join(py_path, i)
            # 获取xlsx的workbook
            wb = openpyxl.load_workbook(file_path_xlsx)
            # 获取xlsx的所有sheet
            sheet_names = wb.sheetnames
            # 加载所有工作表
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                vo_id_column, file_name_column, external_type_column, state_column = get_vo_excel_column(sheet)
                if state_column:
                    # 获取状态下的内容
                    for state in list(sheet.columns)[state_column - 1]:
                        if state.value and state.value == "cancel":
                            if file_name_column:
                                # print(sheet.cell(state.row, column=vo_id_column).value)
                                media_name = (sheet.cell(state.row, column=file_name_column).value)
                                # 删除相应的.wem文件
                                delete_cancel_wem(media_name)

                                # 删除media_info表中的内容
                                row_index_list = media_info_data.index[
                                    media_info_data['MediaName'] == (media_name + '.wem')].tolist()

                                for row_index in row_index_list:
                                    media_info_data.drop(row_index, inplace=True)
                                media_info_data.to_csv(media_info_path, index=False)

                                # 删除external_cookie的内容
                                row_index_list = external_cookie_data.index[
                                    external_cookie_data['MediaName'] == (media_name + '.wem')].tolist()
                                print(row_index_list)
                                for row_index in row_index_list:
                                    external_cookie_data.drop(row_index, inplace=True)
                                    external_cookie_data.to_csv(external_cookie_path, index=False)


# 获取媒体资源文件列表
wav_path = os.path.join(py_path, "New_Media")
file_wav_dict, _ = get_type_file_name_and_path('.wav', wav_path)
# pprint(file_wav_dict)
# {'VO_Lorin_01.wav': 'F:\\pppppy\\SP\\module\\waapi\\waapi_Auto_Import_ExternalSource\\New_Media\\VO_Lorin_01.wav',
#  'VO_Lorin_02.wav': 'F:\\pppppy\\SP\\module\\waapi\\waapi_Auto_Import_ExternalSource\\New_Media\\VO_Lorin_02.wav',
#  'VO_Lorin_03.wav': 'F:\\pppppy\\SP\\module\\waapi\\waapi_Auto_Import_ExternalSource\\New_Media\\VO_Lorin_03.wav',}


with WaapiClient() as client:
    """查找对象"""


    def find_obj(args):
        options = {
            'return': ['name', 'id', 'path', 'notes', 'shortId', 'parent']

        }
        obj_sub_list = client.call("ak.wwise.core.object.get", args, options=options)['return']
        if not obj_sub_list:
            obj_sub_id = ""
            obj_sub_path = ""
        else:
            obj_sub_id = obj_sub_list[0]['id']
            obj_sub_path = obj_sub_list[0]['path']
        return obj_sub_list, obj_sub_id, obj_sub_path


    """自动生成externalsource"""


    def gen_external():
        args = {
            "sources": [
                {
                    "input": external_input_path,
                    "platform": "Windows",
                    "output": external_output_win_path
                },
                {
                    "input": external_input_path,
                    "platform": "Mac",
                    "output": external_output_mac_path
                }

            ]
        }

        gen_log = client.call("ak.wwise.core.soundbank.convertExternalSources", args)


    # 查找External下的所有Sound
    external_sound_list, _, _ = find_obj(
        {'waql': ' "%s" select descendants where type = "ExternalSource" ' % vo_external_path})
    # pprint(external_sound_list)
    # {'id': '{2D2C454C-38BE-414D-8712-780F570DC12D}',
    #   'name': 'vn2',
    #   'notes': '',
    #   'parent': {'id': '{1F7E9E8F-3BC4-484A-945F-45A3F9457BA6}',
    #              'name': 'VO_External_NPC_2D'},
    #   'path': '\\Actor-Mixer '
    #           'Hierarchy\\v1\\VO\\VO\\VO_External\\VO_External\\VO_External_NPC\\VO_External_NPC_2D\\vn2',
    #   'shortId': 560471565}

    """*******************主程序*******************"""
    # 移除标记为cancel的行
    drop_row_by_cancel()

    # 查找要导入的媒体文件里有没有对应的
    for file_wav_name in file_wav_dict:
        flag = 0
        # 读excel表
        # 提取规则：只提取xlsx文件
        for i in file_name_list:
            if ".xlsx" in i:
                # 拼接xlsx的路径
                file_path_xlsx = os.path.join(py_path, i)
                # 获取xlsx的workbook
                wb = openpyxl.load_workbook(file_path_xlsx)
                # 获取xlsx的所有sheet
                sheet_names = wb.sheetnames
                # 加载所有工作表
                for sheet_name in sheet_names:
                    sheet = wb[sheet_name]
                    vo_id_column, file_name_column, external_type_column, state_column = get_vo_excel_column(sheet)
                    if file_name_column:
                        # 获取音效名下的内容
                        for cell_sound in list(sheet.columns)[file_name_column - 1]:
                            if cell_sound.value and (cell_sound.value != "文件名"):

                                if cell_sound.value in file_wav_name:
                                    if sheet.cell(row=cell_sound.row, column=state_column).value != 'cancel':
                                        # 在表格中找到此音效才将其写入xml
                                        xml_create_element(file_wav_dict[file_wav_name])
                                        flag = 1

                                        # 在状态处自动标记完成
                                        sheet.cell(row=cell_sound.row, column=state_column).value = 'done'

                                        # 查找Wwise中有无相应的Sound
                                        for external_sound_dict in external_sound_list:
                                            if external_type_column:
                                                if external_sound_dict['parent']['name'] == sheet.cell(
                                                        row=cell_sound.row,
                                                        column=external_type_column).value:

                                                    # pprint(external_sound_dict['shortId'])
                                                    # 查找mediainfo表的id有没有，没有则添加，有则修改内容
                                                    if vo_id_column:
                                                        vo_id = sheet.cell(
                                                            row=cell_sound.row,
                                                            column=vo_id_column).value
                                                        # 写入media_info csv表
                                                        write_media_info_csv()
                                                        # 写入wwise_cookie表
                                                        write_external_cookie_csv()
                                                        break
                                    break
                wb.save(file_path_xlsx)
        if flag == 0:
            print_error(file_wav_name + "：在语音需求表中不存在，请检查名称是否正确或在表格中补充该名字")

    # xml文件写入
    with open('ExternalSource.xml', 'w+') as f:
        # 按照格式写入
        f.write(doc.toprettyxml())
        f.close()

    # 复制xml为wsources
    source_path = shutil.copy2(os.path.join(py_path, 'ExternalSource.xml'),
                               os.path.join(py_path, 'ExternalSource.wsources'))

    # 自动生成externalsource
    gen_external()
