import openpyxl
import os
import module.oi.oi_h as oi_h

"""打开工作簿并获取工作表"""


def excel_get_sheet(path, sheetname):
    # 打开工作簿
    wb = openpyxl.load_workbook(path)
    # print(wb)
    # 获取工作表
    sheet = wb[sheetname]
    # print(sheet)
    return sheet, wb


"""获取表头数据所在列"""


def excel_get_sheet_title_column(sheet, title_list):
    # 标题所在列的字典映射
    title_colunmn_dict = {}
    # 遍历第一行
    for cell in list(sheet.rows)[0]:
        for title_name in title_list:
            if title_name in str(cell.value):
                title_colunmn_dict[title_name] = cell.column
    return title_colunmn_dict


"""获取表头所有数据所在列"""


def excel_get_all_sheet_title_column(sheet):
    # 标题所在列的字典映射
    title_colunmn_dict = {}
    # 遍历第一行
    for cell in list(sheet.rows)[0]:
        title_colunmn_dict[cell.value] = cell.column
    return title_colunmn_dict


"""获取当前目录下要遍历的表路径列表"""


def excel_get_path_list(py_path):
    file_names = []
    for i in os.walk(py_path):
        file_names.append(i)
    # pprint("输出文件夹下的文件名：")
    file_name_list = file_names[0][2]
    return file_name_list


"""获取标题所在列"""


def sheet_title_column(title_name, title_colunmn_dict):
    file_name = ""
    title_colunmn = ""
    if title_name in title_colunmn_dict.keys():
        # 读取测试
        title_colunmn = title_colunmn_dict[title_name]
    else:
        oi_h.print_error(title_name + ":标题列不存在，请检查表格中是否有该标题的列表")

    return title_colunmn


"""遍历第n列内容并生成列表,列表生成的是单列"""


def get_colunmn_one_list(column_index, sheet):
    # 初始化列表
    column_data = []

    # 遍历第 8 列，从第二行开始
    for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True):
        # row 是一个元组，只有一个元素
        if row not in column_data:
            column_data.append(row[0])

    return column_data


"""遍历第n列内容并生成列表,列表生成的是元组"""


def get_colunmn_multi_list(column_min, column_max, sheet):
    # 初始化列表
    column_data = []

    # 遍历第 8 列，从第二行开始
    for row in sheet.iter_rows(min_row=2, min_col=column_min, max_col=column_max, values_only=True):
        # row 是一个元组，只有一个元素
        if row not in column_data:
            column_data.append(row)

    return column_data


# 测试
# sheet, wb = excel_get_sheet(r"F:\pppppy\SP\module\waapi\waapi_Auto_Import_ExternalSource\MediaInfoTable.xlsx", "Sheet1")
# column_data = get_colunmn_one_list(1, sheet)
# print(column_data)

"""创建表格ID"""


# id_list：获取当前id表中的所有id
# id_type_dict:获取当前类型的id区间段列表
# id_type_name：要获取的区间段类型
# id_sheet：id表的sheet
def create_id(id_type_dict, id_type_name, id_sheet):
    # 获取目前已有的ID列表
    id_list = get_colunmn_one_list(1, id_sheet)
    for id_type in id_type_dict:
        if id_type in id_type_name:
            # print(id_type)
            id_min = id_type_dict[id_type]['min']
            id_max = id_type_dict[id_type]['max']
            for i in range(id_min, id_max):
                if i not in id_list:
                    return i


# 测试
# sheet_mediainfo, wb_mediainfo = excel_get_sheet(
#     r"F:\pppppy\SP\module\waapi\waapi_Auto_Import_ExternalSource\MediaInfoTable.xlsx", 'Sheet1')
# id_l = create_id(config.es_id_config,
#                  r"F:\pppppy\SP\module\waapi\waapi_Auto_Import_ExternalSource\B类（LevelSequence）剧情语音.xlsx",
#                  sheet_mediainfo)
# print(id_l)


"""将表1的多列复制到表二的多列"""


# source_columns_list：源需要复制的列的索引列表
# target_columns_list：目标需要复制的列的索引列表

def copy_excel1_to_excel2_n_column(sheet1, sheet2, wb2, source_columns_list, target_columns_list, excel2_path):
    # 定义源列和目标列的对应关系

    # 遍历每一行并复制指定列
    for row in sheet1.iter_rows(min_row=1, max_row=sheet1.max_row):
        for src_col, tgt_col in zip(source_columns_list, target_columns_list):
            sheet2.cell(row=row[0].row, column=tgt_col).value = row[src_col - 1].value

    # 保存目标 Excel 文件
    wb2.save(excel2_path)


