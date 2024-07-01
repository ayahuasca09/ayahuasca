import unreal
import openpyxl
import json
from waapi import WaapiClient
import re

"""ID表内容写入"""
# 获取ID表路径
# 打开工作簿
excel_path = r"F:\pppppy\SP\module\ue\ue_ID表生成\Audio.xlsx"
wb = openpyxl.load_workbook(excel_path)
# 获取工作表
sheet = wb['audio']
# unreal.log(wb)
# unreal.log(sheet)

"""Wwise路径"""
event_root = "\\Events\\v1"

"""Event_Json_Info读取"""
# json_path = r"F:\pppppy\SP\module\ue\ue_ID表生成\js_event_info.json"
# with open(json_path, 'r', encoding='UTF-8') as jsfile:
#     js_list = json.load(jsfile)
# # unreal.log(js_list)
# # 读取notes和name
# for js_dict in js_list:
#     unreal.log(js_dict['name'])
#     unreal.log(js_dict['notes'])

range_min = 0
range_max = 0
# ID
temp = 0

event_id_config = {
    "Amb":
        {
            "min": 1,
            "max": 3000
        },
    "Char":
        {
            "min": 3001,
            "max": 7000
        },
    "Imp":
        {
            "min": 7001,
            "max": 8000
        },
    "Mon":
        {
            "min": 8001,
            "max": 13000
        },
    "Mus":
        {
            "min": 13001,
            "max": 14000
        },
    "Sys":
        {
            "min": 14001,
            "max": 15000
        },

}


def get_module_range(event_name):
    global range_min, range_max
    flag = 0
    for module_name in event_id_config:
        if str(module_name) in str(event_name):
            range_min = event_id_config[module_name]['min']
            range_max = event_id_config[module_name]['max']
            flag = 1
            return
    if flag == 0:
        range_min = 0
        range_max = 0
        unreal.log_error("找不到" + event_name + "的系统名")


def set_event_path(insert_row):
    flag = 0
    # Event路径赋值
    for asset_path in allAssets:
        if event_dict['name'] in asset_path:
            asset_refer_path = "/Script/AkAudio.AkAudioEvent\'" + asset_path + "\'"
            sheet.cell(row=insert_row, column=2).value = asset_refer_path
            flag = 1
            break
    if flag == 0:
        sheet.cell(row=insert_row, column=2).value = ""


# # 表的属性值修改
# def modify_event_property(insert_row, type_event_dict):
#     # 事件描述赋值
#     sheet.cell(row=insert_row, column=4).value = type_event_dict['notes']
#     # Event路径赋值
#     for asset_path in allAssets:
#         if event_dict['name'] in asset_path:
#             asset_refer_path = "/Script/AkAudio.AkAudioEvent\'" + asset_path + "\'"
#             sheet.cell(row=insert_row, column=2).value = asset_refer_path


# ID生成
def create_id():
    global temp
    # 获取当前的最小序列
    # 记录累加的数
    temp = range_min
    # 获取第0列
    for cell in list(sheet.columns)[0]:
        if cell.value != None:
            # print(cell.value)
            for i in range(range_min, range_max):
                if cell.value == i:
                    temp = temp + 1


# 表的值设置
def set_value():
    global temp
    create_id()
    flag = 0
    for cell in list(sheet.columns)[2]:
        if event_dict['name'] == cell.value:
            flag = 1
            # 事件描述赋值
            sheet.cell(row=cell.row, column=4).value = event_dict['notes']
            set_event_path(cell.row)
            break
        elif event_dict['notes'] == sheet.cell(cell.row, column=4).value:
            flag = 2
            # 事件名称复制
            sheet.cell(row=cell.row, column=3).value = event_dict['name']
            set_event_path(cell.row)
    if flag == 0:
        # 插入为空的行
        insert_row = sheet.max_row + 1
        # pprint(insert_row)
        # id编号赋值
        sheet.cell(row=insert_row, column=1).value = temp
        temp = temp + 1
        # 事件名称赋值
        sheet.cell(row=insert_row, column=3).value = event_dict['name']
        # 事件描述赋值
        sheet.cell(row=insert_row, column=4).value = event_dict['notes']
        set_event_path(insert_row)


"""******************UE功能区**********************"""

# unreal.log(path_event)
assetlib = unreal.EditorAssetLibrary()
# 查找有引用的资产
# 遍历Event
event_path = "/Game/Audio/WwiseAudio/Events"
allAssets = assetlib.list_assets(event_path, True, False)
# for asset_path in allAssets:
#     # unreal.log(allAssets)
#     excel_path_original = asset_path
#     # excel_path = asset_path.replace("/Game/WwiseAsset/AK_Event/", "")
#     unreal.log(excel_path_original)

"""从Wwise读Event_Info"""
with WaapiClient() as client:
    def find_obj(args):
        options = {
            'return': ['name', 'id', 'path', 'notes']

        }
        obj_sub_list = client.call("ak.wwise.core.object.get", args, options=options)['return']
        if not obj_sub_list:
            obj_sub_id = ""
            obj_sub_path = ""
        else:
            obj_sub_id = obj_sub_list[0]['id']
            obj_sub_path = obj_sub_list[0]['path']
        return obj_sub_list, obj_sub_id, obj_sub_path


    """*************主程序*************"""
    event_list, event_id, _ = find_obj(
        {'waql': ' "%s" select descendants where type = "Event" ' % event_root})
    for event_dict in event_list:
        # 获取值域范围
        get_module_range(event_dict['name'])
        if (range_min != 0) and (range_max != 0):
            set_value()
    # 如果表里找到匹配的Event

wb.save(excel_path)
