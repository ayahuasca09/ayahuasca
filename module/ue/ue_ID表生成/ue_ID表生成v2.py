import openpyxl
from waapi import WaapiClient
import os
import shutil
import sys
import re
import pandas as pd
import configparser
import tkinter as tk
from tkinter import messagebox
from os.path import abspath, dirname
from pprint import pprint

# 文件所在目录
py_path = ""
if hasattr(sys, 'frozen'):
    py_path = dirname(sys.executable)
elif __file__:
    py_path = dirname(abspath(__file__))

"""ID表内容写入"""
# 获取ID表路径
# 打开工作簿
excel_path = os.path.join(py_path, "Audio.xlsx")
wb = openpyxl.load_workbook(excel_path)
# 获取工作表
sheet = wb['audio']

"""csv表路径"""
csv_path = os.path.join(py_path, "Audio.csv")

"""ini文件"""
# 创建 ConfigParser 对象
config = configparser.ConfigParser()
# 读取 ini 文件
config.read(os.path.join(py_path, "Config.ini"))

"""UE工程路径"""
ue_audio_path = config['DEFAULT']['UE_Audio_Path']
# ue_audio_path = r'S:\Ver_1.0.0\Project\Content\Audio\WwiseAudio\Events\v1'

"""Wwise路径"""
event_root = "\\Events\\v1"

range_min = 0
range_max = 0
# ID
temp = 0
rowname_index = 1
audioid_index = 2
audioname_index = 3
audioevent_index = 4
desc_index = 5
isloop_index = 6

event_id_config = {
    "Amb":
        {
            "min": 1,
            "max": 3000
        },
    "Char":
        {
            "min": 3001,
            "max": 7000
        },
    "Imp":
        {
            "min": 7001,
            "max": 8000
        },
    "Mon":
        {
            "min": 8001,
            "max": 13000
        },
    "Mus":
        {
            "min": 13001,
            "max": 14000
        },
    "Sys":
        {
            "min": 14001,
            "max": 15000
        },
    "VO":
        {
            "min": 15001,
            "max": 30000
        },
    "Set_State":
        {
            "min": 30001,
            "max": 32000
        },
    "Set_Switch":
        {
            "min": 32001,
            "max": 34000
        },
    "Set_Trigger":
        {
            "min": 34001,
            "max": 36000
        },
    "CG":
        {
            "min": 36001,
            "max": 40000
        },

}

"""ue路径转换"""


def convert_ue_event_path(old_path):
    new_path = ""
    # 获取从\Audio开始的内容
    result = re.search('(?<=\\\\Content).*', old_path).group()
    # \Audio\WwiseAudio\Events\v1\VO\VO_Game\VO_Game_C01\VO_Game_Battle_C01\AKE_Play_VO_Game_Battle_C01_01.uasset
    # /替换\
    result = result.replace("\\", "/")
    # /Audio/WwiseAudio/Events/v1/VO/VO_Game/VO_Game_C01/VO_Game_Battle_C01/AKE_Play_VO_Game_Battle_C01_01.uasset
    # 去除.uasset
    result = result.replace(".uasset", "")
    # /Audio/WwiseAudio/Events/v1/VO/VO_Game/VO_Game_C01/VO_Game_Battle_C01/AKE_Play_VO_Game_Battle_C01_01
    # 获取后缀
    result_list = result.split("/")
    # 拼接字符串
    if result_list[-1]:
        new_path = '/Script/AkAudio.AkAudioEvent\'/Game' + result + "." + result_list[-1] + '\''
        # print(result)
    return new_path


"""报错捕获"""


def print_error(error_info):
    global is_pass
    is_pass = False
    print("[error]" + error_info)


def print_warning(warn_info):
    print("[warning]" + warn_info)


"""遍历文件目录获取文件名称和路径"""


def get_type_file_name_and_path(file_type, dir_path):
    file_dict = {}
    file_list = []
    # 遍历文件夹下的所有子文件
    # 绝对路径，子文件夹，文件名
    for root, dirs, files in os.walk(dir_path):
        # {'name': ['1111.akd', 'Creature Growls 4.akd', 'Sonic Salute_005_2_20.akd'],
        #  'path': 'S:\\chen.gong_DCC_Audio\\Audio\\SilverPalace_WwiseProject\\Originals\\SFX'}
        for file in files:
            if file_type in file:
                new_dict = {}
                file_dict[file] = os.path.join(
                    root, file)
                new_dict[file] = os.path.join(
                    root, file)
                file_list.append(new_dict)

    return file_dict, file_list


# 获取ID区间
def get_module_range(event_name):
    global range_min, range_max
    flag = 0
    for module_name in event_id_config:
        if str(module_name) in str(event_name):
            range_min = event_id_config[module_name]['min']
            range_max = event_id_config[module_name]['max']
            flag = 1
    if flag == 0:
        range_min = 0
        range_max = 0
        print_error("找不到" + event_name + "的系统名")


# 设置事件路径
def set_event_path(insert_row):
    flag = 0
    # Event路径赋值
    for asset_path in ue_audio_path_list:
        if event_dict['name'] in asset_path:
            sheet.cell(row=insert_row, column=audioevent_index).value = asset_path
            flag = 1
            break
    if flag == 0:
        sheet.cell(row=insert_row, column=audioevent_index).value = ""


# ID生成
def create_id():
    global temp
    # 获取当前的最小序列
    # 记录累加的数
    i = range_min
    # 获取第0列
    for cell in list(sheet.columns)[0]:
        if cell.value and (type(cell.value) is int):
            # print(cell.value)
            if range_min < cell.value < range_max:
                if cell.value > i:
                    i = cell.value
    temp = i + 1


# 获取是否循环并赋值
def get_is_loop():
    is_loop = False
    # 查找事件引用的对象
    args = {
        'waql': '"%s" select descendants where type= "Action" ' % (
            event_dict['id'])
    }

    options = {
        'return': ['Target', 'ActionType']

    }
    obj_sub_list = client.call("ak.wwise.core.object.get", args, options=options)['return']

    if obj_sub_list:
        for obj_dict in obj_sub_list:
            if is_loop:
                break
            else:
                # 为play的事件才继续查
                if obj_dict['ActionType'] == 1:
                    # print("nice")
                    args = {
                        'waql': '"%s" select descendants where type= "Sound" ' % (
                            obj_dict['Target']['id'])
                    }
                    sound_list, _, _ = find_obj(args)
                    # pprint(sound_list)
                    if sound_list:
                        for sound_dict in sound_list:
                            if sound_dict['IsLoopingEnabled']:
                                is_loop = True
                                break

                    # 音乐循环查询
                    args = {
                        'waql': '"%s" select descendants where type= "MusicPlaylistContainer" ' % (
                            obj_dict['Target']['id'])
                    }
                    sound_list, _, _ = find_obj(args)
                    # pprint(sound_list)
                    if sound_list:
                        for sound_dict in sound_list:
                            if sound_dict['id'] in loop_musicplaylist_container_id_list:
                                is_loop = True
                                # print(event_dict['name'])
                                break

    return is_loop


# 表的值设置
def set_value():
    global temp
    create_id()
    flag = 0
    for cell in list(sheet.columns)[2]:
        if event_dict['name'] == cell.value:
            flag = 1
            # 事件描述赋值
            sheet.cell(row=cell.row, column=desc_index).value = event_dict['notes']
            set_event_path(cell.row)
            # 获取是否循环并赋值
            sheet.cell(row=cell.row, column=isloop_index).value = get_is_loop()
            break
        elif (event_dict['notes'] == sheet.cell(cell.row, column=desc_index).value) and event_dict['notes'] \
                and sheet.cell(cell.row, column=desc_index).value:
            flag = 2
            # 事件名称复制
            sheet.cell(row=cell.row, column=audioname_index).value = event_dict['name']
            set_event_path(cell.row)
            # 获取是否循环并赋值
            sheet.cell(row=cell.row, column=isloop_index).value = get_is_loop()
        # 是否循环赋值
    if flag == 0:
        # 插入为空的行
        insert_row = sheet.max_row + 1
        # pprint(insert_row)
        # rowname赋值
        sheet.cell(row=insert_row, column=rowname_index).value = temp
        # id编号赋值
        sheet.cell(row=insert_row, column=audioid_index).value = temp
        # 事件名称赋值
        sheet.cell(row=insert_row, column=audioname_index).value = event_dict['name']
        # 事件描述赋值
        sheet.cell(row=insert_row, column=desc_index).value = event_dict['notes']
        set_event_path(insert_row)
        # 获取是否循环并赋值
        sheet.cell(row=insert_row, column=isloop_index).value = get_is_loop()


"""从Wwise读Event_Info"""
with WaapiClient() as client:
    def find_obj(args):
        options = {
            'return': ['name', 'id', 'path', 'notes', 'IsLoopingEnabled', 'Inclusion',
                       'musicPlaylistRoot', 'LoopCount', 'PlaylistItemType', 'owner', 'parent']

        }
        obj_sub_list = client.call("ak.wwise.core.object.get", args, options=options)['return']
        if not obj_sub_list:
            obj_sub_id = ""
            obj_sub_path = ""
        else:
            obj_sub_id = obj_sub_list[0]['id']
            obj_sub_path = obj_sub_list[0]['path']
        return obj_sub_list, obj_sub_id, obj_sub_path


    """判定音乐是否循环"""


    def find_music_lp_list():
        # PS：MusicPlaylistItem的根节点只有owner，
        # MusicPlaylistItem的其他子节点只有parent，parent为根节点
        args = {
            'waql': 'from type MusicPlaylistItem '
        }
        musicplaylistitem_list, _, _ = find_obj(args)
        # pprint(refer_list)

        # 使用一个列表保存为loop的musicplaylist container id
        loop_musicplaylist_container_id_list = []

        # 查找为loop的、非根节点的子节点
        for musicplaylistitem_dict in musicplaylistitem_list:
            if musicplaylistitem_dict['LoopCount'] == 0:
                # 获取非根子节点
                if (musicplaylistitem_dict['PlaylistItemType'] == 1):
                    # pprint(musicplaylistitem_dict)
                    # print()
                    # 获取子节点的parent，即根节点
                    root_id = musicplaylistitem_dict['parent']['id']
                    # print(root_id)
                    # 获取根节点的owner，即musicplaylist容器的id
                    args = {
                        'waql': ' from object "%s" ' % root_id
                    }
                    root_list, _, _ = find_obj(args)

                # 获取根节点
                else:
                    # 获取根节点的owner，即musicplaylist容器的id
                    args = {
                        'waql': ' from object "%s" ' % musicplaylistitem_dict['id']
                    }
                    root_list, _, _ = find_obj(args)
                    # pprint(root_list)
                for root_dict in root_list:
                    # 添加非重复元素，set为去重，然后再转为list
                    # 获取的是为loop的非根子节点的父级
                    loop_musicplaylist_container_id_list = list(
                        set(loop_musicplaylist_container_id_list + [root_dict['owner']['id']]))
        return loop_musicplaylist_container_id_list


    """*************主程序*************"""

    # 获取为循环的音乐列表
    loop_musicplaylist_container_id_list = find_music_lp_list()
    # pprint(loop_musicplaylist_container_id_list)

    # 存放UE的Event路径
    ue_audio_path_list = []
    # 获取ue工程下的audio并转换路径
    ue_audio_dict, _ = get_type_file_name_and_path('.uasset', ue_audio_path)
    for ue_audio in ue_audio_dict:
        if convert_ue_event_path(ue_audio_dict[ue_audio]):
            ue_audio_path_list.append(convert_ue_event_path(ue_audio_dict[ue_audio]))

    event_list, event_id, _ = find_obj(
        {'waql': ' "%s" select descendants where type = "Event" ' % event_root})
    for event_dict in event_list:
        # 事件需要勾选才生成info信息
        if event_dict['Inclusion']:
            # 获取值域范围
            get_module_range(event_dict['name'])
            if (range_min != 0) and (range_max != 0):
                set_value()

    # 同步删除Wwise中没有的事件
    # 创建Event Name列表
    extract_key = lambda d: d['name']
    event_name_list = list(map(extract_key, event_list))
    # pprint(event_name_list)
    # 获取AudioName列的内容
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value == 'AudioName':
            # 返回该列的值（除了标题行）
            for cell in col[1:]:
                if cell.value not in event_name_list:
                    sheet.delete_rows(cell.row, 1)
                    print_warning(cell.value + "：在Wwise Event中未被找到，该ID及数据已删除")

wb.save(excel_path)

# excel转csv
df = pd.read_excel(excel_path)
# 指定CSV文件名和编码格式
encoding = 'utf-8'
df.to_csv(csv_path, encoding=encoding, index=False)

# # 复制一份excel文件
# excel_no_desc = "Audio.xlsx"
# # shutil.copy2(excel_path, excel_no_desc)
# # 删除描述列
# wb = openpyxl.load_workbook(excel_no_desc)
# # 获取工作表
# sheet = wb['audio']
# # 找到第一行中值为 "Desc" 的列
# for col in sheet.iter_cols(1, sheet.max_column):
#     if col[0].value == "Desc":
#         col_letter = col[0].column_letter
#         sheet.delete_cols(col[0].column)
#         break
# # 保存修改后的工作簿
# wb.save(excel_no_desc)

# 转为csv
# csv_no_desc = "Audio.csv"
# df = pd.read_excel(excel_no_desc)
# df.to_csv(csv_no_desc, encoding=encoding, index=False)

# 应用程序弹窗
root = tk.Tk()
root.withdraw()
messagebox.showinfo("Info", "ID表已更新")
