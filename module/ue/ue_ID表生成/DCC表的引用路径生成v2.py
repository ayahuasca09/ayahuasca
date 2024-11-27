import re
import json
import openpyxl
from pprint import pprint
import warnings

# 自定义库
import module.excel.excel_h as excel_h
import config as config
import module.oi.oi_h as oi_h

# 新特性：excel表分页
# 新特性：excel表分页新增除GE外的其他配置路径

"""**********************数据获取*************************"""
excel_type = '.xlsx'
wb_dcc = openpyxl.load_workbook(config.excel_dcc_dt_audio_page_path)
# 剧情表路径
excel_plot_path_dict, _ = oi_h.get_type_file_name_and_path(excel_type, config.excel_plot_path)
# pprint(excel_plot_path_dict)

"""*****************功能函数**********************"""
"""查找使用剧情表的ID的引用路径"""


def find_plot_ID_refer_path(excel_plot_path_dict):
    # 存储id所在的引用映射路径
    id_refer_dict = {}
    for excel_plot_name in excel_plot_path_dict:
        # 忽略特定类型的警告
        warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')
        wb = openpyxl.load_workbook(excel_plot_path_dict[excel_plot_name])
        # 加载所有工作表
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            column_index = 0
            for cell in sheet[4]:
                if cell.value == "AudioAt2DID":
                    # 需要遍历的列的索引
                    column_index = cell.column
                    # 遍历第索引列的内容
                    for row in sheet.iter_cols(min_col=column_index, max_col=column_index, min_row=5,
                                               max_row=sheet.max_row):
                        for row_cell in row:
                            if row_cell.value:
                                if row_cell.value != 0 and row_cell.value != "0":
                                    if row_cell.value not in id_refer_dict:
                                        id_refer_dict[row_cell.value] = excel_plot_path_dict[excel_plot_name]
                                    # print(row_cell.value)

                    # print(str(column_index) + "：" + excel_plot_name)
                    break
    # pprint(id_refer_dict)
    return id_refer_dict


"""查找使用GE的ID的引用路径"""


def find_GE_ID_refer_path(json_path):
    file_dict, _ = oi_h.get_type_file_name_and_path('.json', json_path)
    # 存储id所在的引用映射路径
    id_refer_dict = {}
    # pprint(file_dict)
    for json_name in file_dict:
        with open(file_dict[json_name], 'r') as jsfile:
            # 这是json转换后的dict，可在python里处理了
            js_dict = json.load(jsfile)
            if "taskConfigData" in js_dict:

                if js_dict["taskConfigData"]:
                    for state in js_dict["taskConfigData"]:
                        state_dict = js_dict["taskConfigData"][state]
                        # pprint(state_dict)
                        if "configList" in state_dict:
                            if state_dict["configList"]:
                                for config_dict in state_dict["configList"]:
                                    # pprint(config_dict)
                                    if "_ClassName" in config_dict:
                                        ue_file_name = re.sub(r'_Skill\.json$', '', json_name)
                                        if config_dict["_ClassName"] == "/Script/SPGameFramework.ANGEConfigPlayAudio":
                                            if (config_dict["eventId"] != 0) and (config_dict["eventId"] != -1):
                                                id_refer_dict[config_dict["eventId"]] = (
                                                    state, ue_file_name)
                                            if "endEventId" in config_dict:
                                                if (config_dict["endEventId"] != 0) and (
                                                        config_dict["endEventId"] != -1):
                                                    id_refer_dict[config_dict["endEventId"]] = (
                                                        state, ue_file_name)
                                                # print(config_dict["endEventId"])

    # pprint(id_refer_dict)
    return id_refer_dict


"""设置配置方式和路径"""


# first_col_value:event id
# config_type：配置类型
# id_refer_dict相应方式的id索引表
def set_config_type_and_path(first_col_value, config_type, id_refer_dict):
    # 如果id为string
    if str(first_col_value) in id_refer_dict:
        if '配置路径' in title_colunmn_dict:
            # 将元组转str
            sheet_dcc.cell(row=row, column=title_colunmn_dict['配置路径']).value = id_refer_dict[str(first_col_value)]
        if '配置方式' in title_colunmn_dict:
            sheet_dcc.cell(row=row, column=title_colunmn_dict['配置方式']).value = config_type
    # 如果id为数字
    if first_col_value in id_refer_dict:
        if '配置路径' in title_colunmn_dict:
            # 将元组转str
            sheet_dcc.cell(row=row, column=title_colunmn_dict['配置路径']).value = id_refer_dict[first_col_value]
        if '配置方式' in title_colunmn_dict:
            sheet_dcc.cell(row=row, column=title_colunmn_dict['配置方式']).value = config_type


"""*****************主程序**********************"""
# 查找ID的引用路径
ge_id_refer_dict = find_GE_ID_refer_path(config.ue_ge_path)
# find_GE_ID_refer_path(r"S:\Ver_1.0.0\Project\Content\SPSkill\BossAshley_Attack1")
# {21: ('state1', 'Candleman_01_c_Stun_HitR'),
#  183: ('state1', 'GratiaShieldATK_CharSkin'),
#  195: ('state1', 'CandlemanMeleeCharSkin'),}
plot_id_refer_dict = find_plot_ID_refer_path(excel_plot_path_dict)
# pprint(plot_id_refer_dict)

sheet_names = wb_dcc.sheetnames
# 加载所有工作表
for sheet_name in sheet_names:
    sheet_dcc = wb_dcc[sheet_name]
    # 获取表头所有数据所在列
    title_colunmn_dict = excel_h.excel_get_all_sheet_title_column(sheet_dcc)
    # print(title_colunmn_dict)
    # 遍历每一行（跳过标题行）
    for row in range(2, sheet_dcc.max_row + 1):  # 假设第一行是标题行
        # 然后需要把数据写入dcc的表
        # 遍历第一列并更新第八列
        first_col_value = sheet_dcc.cell(row=row, column=1).value
        # GE索引检查
        if first_col_value in ge_id_refer_dict:
            if '配置路径' in title_colunmn_dict:
                # 将元组转str
                sheet_dcc.cell(row=row, column=title_colunmn_dict['配置路径']).value = ','.join(
                    ge_id_refer_dict[first_col_value])
            if '配置方式' in title_colunmn_dict:
                sheet_dcc.cell(row=row, column=title_colunmn_dict['配置方式']).value = 'GE'
        # 剧情索引检查
        set_config_type_and_path(first_col_value, '剧情', plot_id_refer_dict)

wb_dcc.save(config.excel_dcc_dt_audio_page_path)
